import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from CriticalFlowPath.keys.secrets import RSLTS_DIR

class WbsFramework:
    def __init__(self, input_file_path, input_file_basename, input_file_extension, 
                 input_worksheet_name, project_ordered_dict):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.worksheet_name = input_worksheet_name
        self.ordered_dict = project_ordered_dict

        #Module Attributes
        self.wbs_start_row = 1
        self.wbs_start_col = 'A'
        self.default_hex_font_color = "00FFFFFF"
        self.default_hex_fill_color = "00FFFF00"

        #Structures
        self.entry_statuses = {
            "added": "#b5f26b",
            "new": "#f2b86b",
            "updated": "#9f6bf2", 
            "modified": "#9f6bf2",
            "matching": "5dc1e8", 
            "removed": "f2706b", 
            "duplicate": "c2110b", 
            "invalid": "#f2ee6b",
            "n/a": "c1c9cc"
        }
    
    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, input_file_extension=None, 
             input_worksheet_name=None, project_ordered_dict=None):
        if auto:
            project = WbsFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_worksheet_name, 
                project_ordered_dict,
            )
        else:
            project = WbsFramework.generate_ins()

        if project:
            table = project.design_json_table(project.ordered_dict)
            proc_table, color_list = project.restructure_for_workable_table(table)
            project.create_wbs_table(proc_table)
            project.process_wbs_column('Activity Code', color_list)
            project.process_wbs_column('Color', color_list) 
            project.style_categorized_column("activity category")

    @staticmethod
    def generate_ins():
        WbsFramework.ynq_user_interaction(
            "Run as Module as stand-alone? "
        )

        setup_cls = setup.Setup.main()
        
        ins = WbsFramework(
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            "Compared - WBS Table",
            setup_cls.obj["project"]["modules"]["MODULE_3"].get("content"),
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension,
                          input_worksheet_name, project_ordered_dict):
        ins = WbsFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_worksheet_name,
            project_ordered_dict
        )
        
        return ins

    @staticmethod
    def display_directory_files(file_list:list) -> int:
        selection_idx = 0
        if len(file_list) == 0:
            print("Error. No files found")
            return -1
        
        if len(file_list) > 1:
            print(f"-- {len(file_list)} files found:")
            idx = 0
            for file in file_list:
                idx += 1
                print(f"{idx}. {file}")

            selection_idx = input("\nPlease enter the index number to select the one to process: ") 
        else:
            print(f"Single file found: {file_list[0]}")
            print("Will go ahead and process")

        return int(selection_idx) - 1

    @staticmethod
    def clear_directory(directory_path:str) -> None:
        if os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            for file in file_list:
                file_path = os.path.join(directory_path, file)
                os.remove(file_path)

    def design_json_table(self, json_dict:dict):
        df_table = pd.DataFrame(json_dict)
        df_table.fillna("N/A", inplace=True)

        reset_table = df_table.reset_index()

        proc_table = pd.pivot_table(
            reset_table,
            index=["phase", "location", "area", "trade", "color", "activity_code"],
            values=["wbs_code", "activity_name", "activity_category", "start", "finish", "difference"],
            aggfunc="first",
            observed=True
        )

        return proc_table
    
    def restructure_for_workable_table(self, table):
        filtered_table = table[~table["activity_category"].str.startswith("INVALID")]

        proc_table = pd.pivot_table(
            filtered_table,
            index=["phase", "location", "area", "trade", "color", "activity_code"],
            values=["wbs_code", "activity_name", "activity_category", "start", "finish", "difference"],
            aggfunc="first",
            observed=True
        )

        proc_table = proc_table.reset_index(drop=False)
        color_list = proc_table["color"].tolist()

        column_header_list = proc_table.columns.tolist()
        column_order = self._check_column_order(
            column_header_list, 
            self._bring_column_to_front(column_header_list, "wbs_code")
        )
        column_order = self._check_column_order(
            column_order, 
            self._send_column_to_back(column_order, "finish")
        )
        column_order = self._check_column_order(
            column_order, 
            self._send_column_to_back(column_order, "difference")
        )

        proc_table = proc_table[column_order]

        column_renames = {
            "phase": "Phase",
            "location": "Location",
            "area": "Area",
            "trade": "Trade",
            "color": "Color",
            "activity_code": "Activity Code",
            "wbs_code": "WBS Code",
            "activity_name": "Activity Name",
            "activity_category": "Activity Category",
            "start": "Start",
            "finish": "Finish",
            "difference": "Difference"
        }
        proc_table.rename(columns=column_renames, inplace=True)

        return proc_table, color_list

    def _check_column_order(self, column_header_list:list, func) -> list:
        if column_header_list[-1] != "finish":
            ordered_header_list = func
        else:
            ordered_header_list = column_header_list

        return ordered_header_list

    def _bring_column_to_front(self, column_list:list, target_column:str) -> list:
        for col in column_list:
            if col == target_column:
                target_idx = column_list.index(target_column)
                target = column_list.pop(target_idx)
                ordered_list = [target] + column_list
                break
        
        return ordered_list
    
    def _send_column_to_back(self, column_list:list, target_column:str) -> list:
        for idx, col in enumerate(column_list):
            if col == target_column:
                temp = column_list[-1]
                column_list[-1] = col
                column_list[idx] = temp
                break
        
        return column_list
    
    def create_wbs_table(self, proc_table) -> None:
        self._write_data_to_excel(proc_table)

    def _write_data_to_excel(self, proc_table) -> None:
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            basename = self.input_basename + '.' + self.input_extension
            file = os.path.join(self.input_path, basename)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name=self.worksheet_name, 
                        startrow=self.wbs_start_row - 1, 
                        startcol=column_index_from_string(self.wbs_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {self.worksheet_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")

    def process_wbs_column(self, col_header:str, color_list:list) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        wb = load_workbook(file)
        ws = wb[self.worksheet_name]

        header_idx = self._find_column_idx(ws, col_header)
        self._fill_color_col(ws, header_idx, color_list)

        wb.save(file)

    def style_categorized_column(self, col_header:str) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        wb = load_workbook(file)
        ws = wb[self.worksheet_name]

        header_idx = self._find_column_idx(ws, col_header)
        
        for row in ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                max_row=ws.max_row, 
                                min_col=header_idx,
                                max_col=header_idx):
            for cell in row:
                color = self.entry_statuses[cell.value.lower()] if cell.value else "n/a" 
                proc_color = self._process_hex_val(color)
                self._style_cell(cell, proc_color)

        wb.save(file)

    def _find_column_idx(self, active_worksheet, column_header:str) -> int|None:
        ws = active_worksheet
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.replace(" ", "_").lower()

        for row in ws.iter_rows(min_row=self.wbs_start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def _fill_color_col(self, active_worksheet, col_idx:int, col_list:list) -> None:
        ws = active_worksheet

        if not col_list:
            print("Error: Color list is empty.")
            return

        for idx, row in enumerate(ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                                max_row=ws.max_row, 
                                                min_col=col_idx,
                                                max_col=col_idx)):
            for cell in row:
                color = col_list[idx]
                proc_color = self._process_hex_val(color)
                try:
                    cell.fill = PatternFill(start_color=proc_color, 
                                            end_color=proc_color, 
                                            fill_type="solid")
                except Exception as e:
                    print(f"Color hex not found: {proc_color}. Error: {e}")
                    cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                            end_color=self.default_hex_fill_color, 
                                            fill_type="solid")

        column_letter = get_column_letter(col_idx)
        print(f"Column ({column_letter}) styled successfully")

    def _style_cell(self, cell, color:str) -> None:
        try:
            cell.fill = PatternFill(start_color=color, 
                                    end_color=color, 
                                    fill_type="solid")
        except Exception as e:
            print(f"Color hex not found: {color}. Error: {e}")
            cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                    end_color=self.default_hex_fill_color, 
                                    fill_type="solid")

    def _process_hex_val(self, hex_val:str) -> str:
        return hex_val.replace('#', "00")


if __name__ == "__main__":
    WbsFramework.main(False)
