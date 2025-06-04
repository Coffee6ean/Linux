import os
import re
import json
import cv2 as cv
import numpy as np
import pytesseract
from datetime import datetime
from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image

import sys
sys.path.append("../")
from CpmProcessor.keys.secrets import TEST_PDF_DIR

class PdfDataIngestion:
    allowed_extensions = ["pdf", "xlsx"]
    project_pages = 1

    def __init__(self, input_file_path, input_file_basename, 
                 input_file_extension, output_file_dir) -> None:
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.output_dir = output_file_dir

        #Module Attributes
        self.sorted_image_list = []

        #Structures
        self.project_content_headers = {
            "entry": None, 
            "phase": ["phase"],
            "area": ["area"],
            "zone": ["zone"], 
            "trade": ["trade"], 
            "color": ["color"],
            "parent_id": None, 
            "activity_code": ["activity_code", "code", "task_code", "act_code"],
            "activity_id": None,
            "activity_name": ["activity_name", "act_name"], 
            "activity_status": ["activity_status", "status", "task_status"], 
            "activity_ins": None, 
            "start": ["start", "start_date", "start_dates"], 
            "finish": ["finish", "finish_date", "finish_dates", "end", "end_date"], 
            "total_float": ["total_float"],
            "activity_successor_id": ["successor"],
        }

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, 
             input_file_extension=None, output_file_dir=None):
        if auto:
            project = PdfDataIngestion.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension, 
                output_file_dir
            )
        else:
            project = PdfDataIngestion.generate_ins()

        module_data = {
            "details": {
                "file_path": None,
                "file_basename": None,
            },
            "logs": {
                "start": PdfDataIngestion.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            },
            "content": {}
        }

        if project:
            file, output_folder = project.create_project_directory()
            module_data["content"] = project.handle_pdf(file, output_folder)

        module_data["logs"]["finish"] = PdfDataIngestion.return_valid_date()
        module_data["logs"]["run-time"] = PdfDataIngestion.calculate_time_duration(
            module_data["logs"].get("start"), 
            module_data["logs"].get("finish")
        )

        return module_data

    @staticmethod
    def generate_ins():
        input_file = PdfDataIngestion.return_valid_file(
            input("Please enter the path to the file or directory: ").strip()
        )

        ins = PdfDataIngestion(
            input_file.get("path"), 
            input_file.get("basename"), 
            input_file.get("extension")
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, output_file_dir):
        ins = PdfDataIngestion(
            input_file_path,
            input_file_basename,
            input_file_extension,
            output_file_dir
        )

        return ins

    @staticmethod
    def return_valid_file(input_file_dir:str) -> dict|int:
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = PdfDataIngestion._handle_dir(input_file_dir)
        elif os.path.isfile(input_file_dir):
            file_dict = PdfDataIngestion._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str="r") -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = PdfDataIngestion._display_directory_files(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = os.path.dirname(input_file_dir), 
            basename = os.path.basename(input_file_basename).split(".")[0],
            extension = os.path.basename(input_file_basename).split(".")[-1],
        )
    
    @staticmethod
    def _display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def _handle_file(input_file_dir:str) -> dict|int:
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in PdfDataIngestion.allowed_extensions):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return -1

    @staticmethod
    def return_valid_path(prompt_message:str) -> (str|None):
        while(True):   
            value = input(prompt_message).strip()
            try:
                if os.path.isdir(value):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    @staticmethod
    def get_type_files(input_dir:str, file_type:str) -> list:
        file_list = []

        if os.path.isfile(input_dir):
            if input_dir.endswith(file_type):
                file_list.append(os.path.abspath(input_dir))
            else:
                print(f"Error: The file '{input_dir}' does not match the specified type '{file_type}'.")

        elif os.path.isdir(input_dir):
            files = os.listdir(input_dir)
            for file in files:
                if file.endswith(file_type):
                    file_path = os.path.join(input_dir, file)
                    file_list.append(os.path.abspath(file_path))

        else:
            print(f"Error: '{input_dir}' is neither a valid file nor a directory.")

        return file_list

    @staticmethod
    def repackage_file(input_dir:str, package_name:str=None) -> str:
        if os.path.isfile(input_dir):
            file_path = os.path.dirname(input_dir)
            file_basename = os.path.basename(input_dir)
            file_extension = os.path.splitext(file_basename)[1]
            new_directory_basename = os.path.splitext(file_basename)[0]

            if isinstance(package_name, str):
                new_directory = os.path.join(file_path, package_name)
            else:
                new_directory = os.path.join(file_path, new_directory_basename)

            if not os.path.exists(new_directory):
                os.mkdir(new_directory)

            new_file_name = f"packaged_{new_directory_basename}{file_extension}"
            new_file_path = os.path.join(new_directory, new_file_name)

            try:
                os.rename(input_dir, new_file_path)
            except Exception as e:
                print(f"Error moving the file {input_dir} to {new_file_path}: {e}")
        else:
            print(f"{input_dir} is not a valid file.")

        return new_file_path

    @staticmethod
    def slice_based_on_color_change(image_path:str, output_dir:str, image_extension:str, 
                                     color_change_threshold:int=50) -> None:
        os.makedirs(output_dir, exist_ok=True)
        image = cv.imread(image_path)

        if image is None:
            print(f"Error: Could not read image '{image_path}'")
            return

        # Convert image to Lab color space (perceptually uniform)
        lab_image = cv.cvtColor(image, cv.COLOR_BGR2Lab)

        # Get image dimensions
        height, width, _ = lab_image.shape

        # List to store slice boundaries
        slice_boundaries = []

        # Loop through rows and calculate color differences between consecutive rows
        for y in range(1, height):
            row1 = lab_image[y - 1, :, :]
            row2 = lab_image[y, :, :]

            # Calculate the mean color for each row
            mean_row1 = cv.mean(row1)[:3]
            mean_row2 = cv.mean(row2)[:3]

            # Calculate the Euclidean distance between the mean colors of the rows
            color_difference = np.sqrt(sum((mean_row2[i] - mean_row1[i]) ** 2 for i in range(3)))

            # If the color difference exceeds the threshold, mark it as a slice boundary
            if color_difference > color_change_threshold:
                slice_boundaries.append(y)

        # Slice the image at detected color change boundaries
        task_slices = []
        previous_boundary = 0

        for boundary in slice_boundaries:
            task_slice = image[previous_boundary:boundary, :]  # Slice from previous boundary to current boundary
            task_slices.append(task_slice)
            previous_boundary = boundary

        # Add the last slice (from the last boundary to the bottom of the image)
        if previous_boundary < height:
            task_slices.append(image[previous_boundary:, :])

        # Save each slice with a unique name
        for idx, task_slice in enumerate(task_slices):
            save_path = os.path.join(output_dir, f"image_slice_{idx+1}.{image_extension}")
            cv.imwrite(save_path, task_slice)

    @staticmethod
    def extract_data_from_image(image_path:str, output_path:str, output_basename:str, 
                                file_extension:str, tesseract_config:str="--psm 4") -> str:
        try:
            if os.stat(image_path).st_size >= 5000:
                image = Image.open(image_path)
                text = pytesseract.image_to_string(image, config=tesseract_config)

                os.makedirs(output_path, exist_ok=True)
                output_file = os.path.join(output_path, f"{output_basename}.{file_extension}")
                
                cleaned_text = PdfDataIngestion._normalize_file_text(text)

                return cleaned_text
        except Exception as e:
            print(f"Error: Unable to write to file at '{output_file}'. Exception: {e}")

    @staticmethod
    def _normalize_file_text(text_body:str) -> str:
        empty_lines = re.compile('^\s*$', re.MULTILINE)
        removed_empty_lines = re.sub(empty_lines, "", text_body)

        whitespace_lines = re.compile('^\n', re.MULTILINE)
        removed_whitespace_lines = re.sub(whitespace_lines, "", removed_empty_lines)

        removed_whitespace_lines = removed_whitespace_lines.strip()
        
        return removed_whitespace_lines

    @staticmethod
    def jsonify_project_data(text_body:str) -> dict:
        text_lines = text_body.split("\n")
        proc_dict = {
            "header": {"page": PdfDataIngestion.project_pages},
            "body": {
                "details": {"activities": {"total": 1}},
                "logs": [],
                "content": [],
            },
        }

        def extract_match(line):
            match = PdfDataIngestion.standard_pattern(line) or PdfDataIngestion.special_pattern(line)
            if match:
                return {
                    "activity_id": match.group("ActivityID").strip() if match.group("ActivityID") else "",
                    "activity_name": match.group("ActivityName").strip() if match.group("ActivityName") else "",
                    "duration": match.group("Duration").strip() if match.group("Duration") else "",
                    "dates": match.group("Dates").strip().split() if match.group("Dates") else []
                }
            return None

        for i, line in enumerate(text_lines):
            entry = i + 1
            match_data = extract_match(line)

            if match_data:
                dates = [re.sub(r"[A*]$", "", date) for date in match_data["dates"]]
                start_date, finish_date = dates[0], dates[1] if len(dates) > 1 else ""

                proc_dict["body"]["content"].append({
                    "entry": entry,
                    "parent_id": match_data["activity_id"],
                    "parent_name": match_data["activity_name"],
                    "parent_duration": match_data["duration"],
                    "parent_start": start_date,
                    "parent_finish": finish_date,
                })
                proc_dict["body"]["details"]["activities"]["total"] += 1
            else:
                proc_dict["body"]["logs"].append({"task": entry, "content": line})

        PdfDataIngestion.project_pages += 1
        return proc_dict

    @staticmethod
    def standard_pattern(text:str):
        special_chars = re.compile('[+=—@_!#$%^&*<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', text)

        pattern = re.compile(
            r"(?P<ActivityID>[A-Z]+(?:\d+(?:-\d+)*|(?:-[A-Z0-9]+)+))"
            r"\s+(?P<ActivityName>.*?)\s+"
            r"(?P<Duration>\d*d?)\s*"
            r"(?P<Dates>\d{2}-[A-Za-z]{3}-\d{2}[A*]?(?:\s+\d{2}-[A-Za-z]{3}-\d{2}[A*]?)?)",
            re.IGNORECASE
        )

        return pattern.match(remove_special_chars)

    @staticmethod
    def special_pattern(text:str):
        special_chars = re.compile('[+=—@_!#$%^&*<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', text)

        pattern = re.compile(
            r"(?P<ActivityID>(?:[A-Z]{2,}(?:-[A-Z0-9]+)*-\d+|[A-Z]+(?:-[A-Z0-9]+){2,})|)"
            r"(?P<ActivityName>.*?)\s+"
            r"(?P<Duration>(\d+d)|)\s*"
            r"(?P<Dates>(?:\d{2}-[A-Za-z]{3}-\d{2}[A*]?(?:\s+\d{2}-[A-Za-z]{3}-\d{2}[A*]?)?))",
            re.DOTALL | re.IGNORECASE
        )

        return pattern.match(remove_special_chars)

    @staticmethod
    def write_dict_to_json(json_dict:dict, file_name:str, mode:str='w'):
        if not json_dict:
            print("Error: Dictionary is empty. No data to write.\n")
            return
        
        basename = file_name.split(".")[0] if "." in file_name else file_name
        new_directory = os.path.join(
            f"{TEST_PDF_DIR}/{PdfDataIngestion.return_valid_date()}", 
            f"{basename}.json"
        )

        try:
            with open(new_directory, mode) as file_writer:
                json.dump(json_dict, file_writer)

            print(f"Successfully saved Dictionary to JSON:\nFile: {new_directory}\n")
        except Exception as e:
            print(f"An unexpected error occurred while writing to Excel: {e}\n")

    @staticmethod
    def write_text_to_txt(text:str, file_name:str, mode:str='w'):
        if not text:
            print("Error: Text is empty. No data to write.\n")
            return
        
        basename = file_name.split(".")[0] if "." in file_name else file_name
        new_directory = os.path.join(
            f"{TEST_PDF_DIR}/{PdfDataIngestion.return_valid_date()}", 
            f"{basename}.txt"
        )

        try:
            with open(new_directory, mode) as writer:
                writer.write(text)

            print(f"Successfully saved Text to TXT:\nFile: {new_directory}\n")
        except Exception as e:
            print(f"An unexpected error occurred while writing to Excel: {e}\n")

    def return_excel_workspace(self, worksheet_name:str="Sheet1"):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{worksheet_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = PdfDataIngestion._display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.\n")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def create_project_directory(self):
        file_basename = f"{self.input_basename}.{self.input_extension}" 
        file = os.path.join(self.input_path, file_basename) 
        output_folder = f"{TEST_PDF_DIR}/{PdfDataIngestion.return_valid_date()}"
        os.makedirs(output_folder, exist_ok=True)

        print("Directory successfully created")
        return file, output_folder

    def pdf_to_images(self, input_file:str, output_directory:str, image_extension:str="png") -> None:
        pdf = input_file

        if pdf:
            images = convert_from_path(pdf)

            for i, image in enumerate(images):
                image_path = os.path.join(output_directory, f'page_{i + 1}.{image_extension}')
                image.save(image_path, 'PNG')

            print(f"Images saved to: {output_directory}")
        else:
            print("Error: Invalid PDF file.")

    def resize_images(self, image_path:str, image_extension:str="png") -> None:
        image_list = self._bubble_sort_str_list(os.listdir(image_path))

        for i, image in enumerate(image_list):
            image_file = f"{image_path}/{image}"
            
            with Image.open(image_file) as img:
                new_width, new_height = img.width * 2, img.height * 2
                resized_img = img.resize((new_width, new_height))
                
                resized_image_path = f"{image_path}/resized_page_{i+1}.{image_extension}"
                self.sorted_image_list.append(resized_image_path)
                resized_img.save(resized_image_path)
                os.remove(image_file)

        print("Successfully resized images")

    def _bubble_sort_str_list(self, img_dir:list) -> list:
        n = len(img_dir)
        
        for i in range(n):
            swap = False

            for j in range(n - 1 - i):

                img_id = re.search(r'\d+', img_dir[j])
                one_plus_img_id = re.search(r'\d+', img_dir[j + 1])

                img_id = int(img_id.group()) if img_id else 0
                one_plus_img_id = int(one_plus_img_id.group()) if one_plus_img_id else 0

                if img_id > one_plus_img_id:
                    img_dir[j], img_dir[j + 1] = img_dir[j + 1], img_dir[j]
                    swap = True

            if not swap:
                break

        return img_dir

    def handle_pdf(self, input_file:str, image_path:str, roi_msg:str="Select ROI", image_extension:str="png") -> None:
        self.pdf_to_images(input_file, image_path)
        self.resize_images(image_path)
        self._select_wbs_region(image_path, roi_msg, image_extension)
        self._repackage_images("package")
        #self._slice_images("slices", image_extension)
        json_dict = self._return_json_obj(image_path)

        return json_dict
        
    def _select_wbs_region(self, image_path:str, roi_msg:str, image_extension:str) -> None:
        if not roi_msg:
            roi_msg = "Select the area"

        if self.sorted_image_list:
            self._crop_images(image_path, roi_msg, image_extension)
            cv.destroyWindow(roi_msg)
        else:
            print("Error: No images found")
        
        print("WBS section successfully defined and images cropped")

    def _crop_images(self, image_path:str, roi_msg:str, image_extension:str) -> None:
        new_images = []
        first_image_path = self.sorted_image_list[0]
        first_image = cv.imread(first_image_path)

        if first_image is None:
            print(f"Error: Could not read image '{first_image_path}'")
            return

        r = cv.selectROI(roi_msg, first_image)

        for i, img_name in enumerate(self.sorted_image_list):
            image = cv.imread(img_name)

            if image is None:
                print(f"Warning: Could not read image '{img_name}' - Skipping.")
                continue

            cropped_image = image[int(r[1]):int(r[1] + r[3]), int(r[0]):int(r[0] + r[2])]

            output_basename = f"cropped_img_{i+1}.{image_extension}"
            output_path = os.path.join(image_path, output_basename)
            
            cv.imwrite(output_path, cropped_image)
            new_images.append(output_path)
            os.remove(img_name)

        self.sorted_image_list.clear()
        self.sorted_image_list = new_images

    def _repackage_images(self, package_name:str) -> None:
        new_images = []
        for i, image in enumerate(self.sorted_image_list):
            output_path = PdfDataIngestion.repackage_file(image, f"{package_name}_{i+1}")
            new_images.append(output_path)

        self.sorted_image_list.clear()
        self.sorted_image_list = new_images
        print("Images successfully packaged")

    def _slice_images(self, new_directory:str, image_extension:str) -> None:
        for image in self.sorted_image_list:
            output_dir = os.path.join(os.path.dirname(image), f"{new_directory}")
            PdfDataIngestion.slice_based_on_color_change(image, output_dir, image_extension)

        print("Images succesfully sliced")

    def _return_json_obj(self, output_path:str, target_directory:str="", 
                         file_extension:str="json", mode:str='a') -> None:
        project_dict = {}
    
        for i, package in enumerate(self.sorted_image_list):
            images_path = os.path.dirname(package) + '/' + target_directory
            sliced_images = self._bubble_sort_str_list(os.listdir(images_path))

            for image in sliced_images:
                image_dir = os.path.join(images_path, image)

                normalized_text = PdfDataIngestion.extract_data_from_image(
                    image_dir, 
                    output_path, 
                    "project_text",
                )
                project_dict[i+1] = PdfDataIngestion.jsonify_project_data(normalized_text)
            
            #PdfDataIngestion.write_text_to_txt(normalized_text, "project_text", mode)
            
        output_file = os.path.join(output_path, f"project_dict.{file_extension}")
        #PdfDataIngestion.write_dict_to_json(project_dict, "project_dict")
        
        print("Project json file successfully created")
        return output_file


if __name__ == "__main__":
    PdfDataIngestion.main(False)
