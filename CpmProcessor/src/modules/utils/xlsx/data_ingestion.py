import os
import re
import json
from dateutil import parser
from openpyxl import load_workbook
from datetime import datetime, date
from openpyxl.utils import get_column_letter, column_index_from_string 

import sys
sys.path.append("../")
from CpmProcessor.keys.secrets import TEST_XLSX_DIR, TEST_JSON_DIR

class XlsxDataIngestion:
    allowed_extensions = ["pdf", "xlsx"]
    project_pages = 1

    def __init__(self, input_file_path, input_file_basename, 
                 input_file_extension, output_file_dir) -> None:
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.output_dir = output_file_dir

        #Module Attributes
        self.start_col_idx = 1
        self.start_row_idx = 1
        self.ws_name_list = []

        #Structures
        self.activity_headers = {
            "phase": ["phase"],
            "area": ["area", "location"], 
            "zone": ["zone"],
            "trade": ["trade"], 
            "color": ["color"],
            "activity_code": ["activity_code", "code", "task_code", "act_code"],
            "wbs_code": ["wbs_code", "activity_id"],
            "activity_name": ["activity_name", "act_name", "task_name"], 
            "activity_status": ["activity_status", "status", "task_status", "status_code"], 
            "activity_duration": ["activity_duration", "duration", "remaining_duration"],
            "start": ["start", "start_date", "start_dates"], 
            "finish": ["finish", "finish_date", "finish_dates", "end", "end_date"], 
            "difference": ["difference"],
            "total_float": ["total_float"],
            "successor_code": ["successor"],
            "predecessor_code": ["predecessor"],
        }
        self.json_struct_categories = ["phase", "location", "area", "trade", "activity_code"]

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, 
             input_file_extension=None, output_file_dir=None):
        if auto:
            project = XlsxDataIngestion.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension, 
                output_file_dir
            )
        else:
            project = XlsxDataIngestion.generate_ins()

        if project:
            module_data = project.process_workbook(project)

        #XlsxDataIngestion.write_dict_to_json(module_data, "multiple_tabs_test", TEST_JSON_DIR)

        return module_data
    
    @staticmethod
    def process_workbook(project) -> dict:
        if project:
            xlsx_processed = {}
            xlsx_extracted = project.handle_xlsx()

            for sheet in project.ws_name_list:
                project.process_json(xlsx_processed, xlsx_extracted[sheet], sheet)
        
        return xlsx_processed

    @staticmethod
    def generate_ins():
        input_file = XlsxDataIngestion.return_valid_file(
            input("Please enter the path to the file or directory: ").strip()
        )

        ins = XlsxDataIngestion(
            input_file.get("path"), 
            input_file.get("basename"), 
            input_file.get("extension"),
            TEST_XLSX_DIR
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, output_file_dir):
        ins = XlsxDataIngestion(
            input_file_path,
            input_file_basename,
            input_file_extension,
            output_file_dir
        )

        return ins

    @staticmethod
    def return_valid_file(input_file_dir:str) -> dict|int:
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = XlsxDataIngestion._handle_dir(input_file_dir)
        elif os.path.isfile(input_file_dir):
            file_dict = XlsxDataIngestion._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str="r") -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = XlsxDataIngestion._display_options(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = os.path.dirname(input_file_dir), 
            basename = os.path.basename(input_file_basename).split(".")[0],
            extension = os.path.basename(input_file_basename).split(".")[-1],
        )
    
    @staticmethod
    def _display_options(file_list:list) -> list:
        if not file_list:
            print('Error: No elements found.')
            return []

        print(f'-- {len(file_list)} elements found:')
        for idx, file in enumerate(file_list, start=1):
            print(f'{idx}. {file}')

        result = []
        selection_input = input('\nEnter index numbers (comma-separated) to select elements to process: ').split(',')

        for selection in selection_input:
            selection = selection.strip()
            if not selection.isdigit():
                print(f'Error: Invalid input "{selection}", skipping.')
                continue

            index = int(selection)
            if 1 <= index <= len(file_list):
                result.append(index)
            else:
                print(f'Error: "{index}" is out of range (1 to {len(file_list)}).')

        return result

    @staticmethod
    def _handle_file(input_file_dir:str) -> dict|int:
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in XlsxDataIngestion.allowed_extensions):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return -1

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    @staticmethod
    def _normalize_file_text(text_body:str) -> str:
        empty_lines = re.compile('^\s*$', re.MULTILINE)
        removed_empty_lines = re.sub(empty_lines, "", text_body)

        whitespace_lines = re.compile('^\n', re.MULTILINE)
        removed_whitespace_lines = re.sub(whitespace_lines, "", removed_empty_lines)

        removed_whitespace_lines = removed_whitespace_lines.strip()
        
        return removed_whitespace_lines

    @staticmethod
    def jsonify_project_data(text_body:str) -> dict:
        text_lines = text_body.split("\n")
        proc_dict = {
            "header": {"page": XlsxDataIngestion.project_pages},
            "body": {
                "details": {"activities": {"total": 1}},
                "logs": [],
                "content": [],
            },
        }

        def extract_match(line):
            match = XlsxDataIngestion._standard_pattern(line) or XlsxDataIngestion._special_pattern(line)
            if match:
                return {
                    "activity_id": match.group("ActivityID").strip() if match.group("ActivityID") else "",
                    "activity_name": match.group("ActivityName").strip() if match.group("ActivityName") else "",
                    "duration": match.group("Duration").strip() if match.group("Duration") else "",
                    "dates": match.group("Dates").strip().split() if match.group("Dates") else []
                }
            return None

        for i, line in enumerate(text_lines):
            entry = i + 1
            match_data = extract_match(line)

            if match_data:
                dates = [re.sub(r"[A*]$", "", date) for date in match_data["dates"]]
                start_date, finish_date = dates[0], dates[1] if len(dates) > 1 else ""

                proc_dict["body"]["content"].append({
                    "entry": entry,
                    "parent_id": match_data["activity_id"],
                    "parent_name": match_data["activity_name"],
                    "parent_duration": match_data["duration"],
                    "parent_start": start_date,
                    "parent_finish": finish_date,
                })
                proc_dict["body"]["details"]["activities"]["total"] += 1
            else:
                proc_dict["body"]["logs"].append({"task": entry, "content": line})

        XlsxDataIngestion.project_pages += 1
        return proc_dict

    @staticmethod
    def _standard_pattern(text:str):
        special_chars = re.compile('[+=—@_!#$%^&*<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', text)

        pattern = re.compile(
            r"(?P<ActivityID>[A-Z]+(?:\d+(?:-\d+)*|(?:-[A-Z0-9]+)+))"
            r"\s+(?P<ActivityName>.*?)\s+"
            r"(?P<Duration>\d*d?)\s*"
            r"(?P<Dates>\d{2}-[A-Za-z]{3}-\d{2}[A*]?(?:\s+\d{2}-[A-Za-z]{3}-\d{2}[A*]?)?)",
            re.IGNORECASE
        )

        return pattern.match(remove_special_chars)

    @staticmethod
    def _special_pattern(text:str):
        special_chars = re.compile('[+=—@_!#$%^&*<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', text)

        pattern = re.compile(
            r"(?P<ActivityID>(?:[A-Z]{2,}(?:-[A-Z0-9]+)*-\d+|[A-Z]+(?:-[A-Z0-9]+){2,})|)"
            r"(?P<ActivityName>.*?)\s+"
            r"(?P<Duration>(\d+d)|)\s*"
            r"(?P<Dates>(?:\d{2}-[A-Za-z]{3}-\d{2}[A*]?(?:\s+\d{2}-[A-Za-z]{3}-\d{2}[A*]?)?))",
            re.DOTALL | re.IGNORECASE
        )

        return pattern.match(remove_special_chars)

    @staticmethod
    def write_dict_to_json(json_dict:dict, file_name:str, output_folder:str, mode:str='w') -> None:
        if not json_dict:
            print("Error: Dictionary is empty. No data to write.\n")
            return
        
        basename = file_name.split(".")[0] if "." in file_name else file_name
        new_directory = os.path.join(
            output_folder, 
            f"{basename}.json"
        )

        try:
            with open(new_directory, mode) as file_writer:
                json.dump(json_dict, file_writer)

            print(f"Successfully saved Dictionary to JSON:\nFile: {new_directory}\n")
        except Exception as e:
            print(f"An unexpected error occurred while writing to Excel: {e}\n")

    def create_project_directory(self):
        file_basename = f"{self.input_basename}.{self.input_extension}" 
        file = os.path.join(self.input_path, file_basename) 
        output_folder = f"{self.output_dir}/{XlsxDataIngestion.return_valid_date()}"
        os.makedirs(output_folder, exist_ok=True)

        print("Directory successfully created")
        return file, output_folder

    def handle_xlsx(self, ws_name:str="") -> dict:
        wb, ws_list = self._return_excel_workspace(ws_name)

        xlsx_results = {ws:{} for ws in ws_list}

        for ws in ws_list: 
            start = XlsxDataIngestion.return_valid_date()
            file_headers = self._xlsx_return_header(wb[ws], self.start_row_idx, self.start_col_idx)

            json_header = self._json_fill_header(file_headers)
            json_obj, entry_counter = self._json_fill_body(wb[ws], json_header)
            reworked_json = self._fill_missing_dates(json_obj)
            earliest_start, latest_finish = self._project_dates(reworked_json)
            nested_json = self._build_nested_dic(reworked_json)
            
            xlsx_results[ws] = {
                "proc_start": start,
                "proc_finish": XlsxDataIngestion.return_valid_date(),
                "flattend_json": reworked_json,
                "nested_json": nested_json,
                "entry_count": entry_counter,
                "earliest_start": earliest_start,
                "latest_finish": latest_finish,
            }
            
        return xlsx_results

    def _return_excel_workspace(self, worksheet_name:str="Sheet1") -> tuple:
        basename = f"{self.input_basename}.{self.input_extension}"
        file_path = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file_path)
            worksheet_list = [workbook[worksheet_name]]
            return workbook, worksheet_list
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            return self._handle_missing_worksheet(workbook, worksheet_name)
        
    def _handle_missing_worksheet(self, workbook, worksheet_name:str) -> tuple:
        while True:
            user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
            
            if user_answer == 'y':
                return self._create_new_worksheet(workbook, worksheet_name)
            elif user_answer == 'n':
                return self._select_existing_worksheets(workbook)
            elif user_answer == 'q':
                print("Quitting without changes.\n")
                return workbook, []
            else:
                print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")

    def _create_new_worksheet(self, workbook, worksheet_name:str) -> tuple:
        worksheet_list = [workbook.create_sheet(worksheet_name)]
        self.ws_name_list = worksheet_list
        print(f"New worksheet '{worksheet_name}' created.\n")
        return workbook, worksheet_list

    def _select_existing_worksheets(self, workbook) -> tuple:
        ws_list = workbook.sheetnames
        selected = XlsxDataIngestion._display_options(ws_list)
        
        if not selected:
            print("Invalid selection. Returning without changes.\n")
            return workbook, []
        
        worksheet_list = [workbook[ws_list[idx - 1]].title for idx in selected]
        self.ws_name_list = worksheet_list
        print(f"Worksheets selected: {self.ws_name_list}")
        return workbook, worksheet_list
        
    def _xlsx_return_header(self, active_worksheet, start_row_idx:int, start_col_idx:int) -> list:
        ws = active_worksheet
        header_list = []

        for row in ws.iter_rows(min_row=start_row_idx, max_row=start_row_idx, 
                                min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    normalized_str = self.normalize_string(cell.value)
                    header_tuple = (cell.coordinate, self._verify_header(normalized_str))
                    header_list.append(header_tuple)

        return header_list
    
    def _verify_header(self, entry_str:str) -> str:
        flat_allowed_headers = {
            value: key for key, values in self.activity_headers.items() if values is not None for value in values
        }

        if entry_str in flat_allowed_headers:
            result = flat_allowed_headers[entry_str]
        else:
            print(
                f"Warning. Header '{entry_str}' not found in declared dictionary"
            )
            result = entry_str
        
        return result
    
    def _json_fill_header(self, file_headers:list) -> dict:
        json_obj = {
            "start": None,
            "finish": None,
            "header": {key: None for key, _ in self.activity_headers.items()},
            "body": [],
        }

        for coordinates, data in file_headers:
            header_dict = json_obj["header"]

            header_dict[data] = coordinates

        return json_obj
    
    def _json_fill_body(self, active_worksheet, json_obj:dict) -> dict:
        ws = active_worksheet
        body_dict = json_obj["body"]
        header_dict = json_obj["header"]

        normalized_header = sorted(
            {k: v for k, v in header_dict.items() if v is not None}.items(), 
            key=lambda item: item[1]
        )

        header_coordinates_list = [dict_tuple[1] for dict_tuple in normalized_header]
        header_key_list = [dict_tuple[0] for dict_tuple in normalized_header]

        first_header_col_letter, first_header_row = self._get_column_coordinates(
            header_dict, normalized_header[0][0]
        )
        last_header_col_letter, _ = self._get_column_coordinates(header_dict, normalized_header[-1][0])
        first_header_col = column_index_from_string(first_header_col_letter)
        last_header_col = column_index_from_string(last_header_col_letter)

        entry_counter = 1
        for row in ws.iter_rows(min_row=first_header_row + 1, max_row=ws.max_row, 
                                min_col=first_header_col, max_col=last_header_col):
            json_activity = {key: None for key in header_dict.keys()}
            json_activity["entry"] = entry_counter
            for cell in row:
                parent_header = next((val for val in header_coordinates_list if get_column_letter(cell.column) in val), None)
                if cell.value:        
                    position = header_coordinates_list.index(parent_header)
                    key = header_key_list[position]
                    json_activity[key] = self._normalize_data_value(key, cell.value)
                else:
                    position = header_coordinates_list.index(parent_header)
                    key = header_key_list[position]

                    if key == "area":
                        json_activity[key] = "N/A"
                    else:
                        json_activity[key] = ""

            body_dict.append(json_activity)
            entry_counter += 1

        return json_obj, entry_counter
    
    def _get_column_coordinates(self, header_dict:dict, header_key:str) -> tuple:
        coordinates = header_dict[header_key]

        row_match = re.search(r'\d+', coordinates)
        column_match = re.search(r'[a-zA-Z]+', coordinates)

        if not row_match or not column_match:
            raise ValueError(f"Invalid coordinates format: {coordinates}")

        row = int(row_match.group(0))
        column = column_match.group(0)

        return column, row 

    def _normalize_data_value(self, header_column:str, cell_value) -> str|None:
        if isinstance(cell_value, str):
            if header_column == 'start' or header_column == 'finish':
                result = self._format_date_string(cell_value)
            else:
                result = cell_value.strip()

            return result
        elif isinstance(cell_value, date):
            return cell_value.strftime('%d-%b-%Y')
        elif isinstance(cell_value, int) or isinstance(cell_value, float):
            if cell_value == 0:
                cell_value = "0"
            return str(cell_value)
        else:
            return None
        
    def _format_date_string(self, date_string:str, output_format="%d-%b-%Y") -> str|None:
        try:
            special_chars = re.compile('[@_!#$%^&*()<>?\|}{~:]')

            if re.search(special_chars, date_string):
                cleaned_date_string = re.sub(special_chars, '', date_string)
            else:
                cleaned_date_string = re.sub("[a-zA-Z]$", '', date_string)
            

            parsed_date = parser.parse(cleaned_date_string.strip())
            formatted_date = parsed_date.strftime(output_format)

            return formatted_date
        except ValueError as e:
            print(f"Error parsing date: {e}")
            return None

    def _fill_missing_dates(self, json_obj:dict) -> dict:
        body_dict = json_obj["body"]

        for item in body_dict:
            start = item.get("start")
            finish = item.get("finish")

            if start == "" and finish == "":
                print(f"Both 'start' and 'finish' are missing for entry: {item['entry']}")
            elif start == "":
                item["start"] = finish
            elif finish == "":
                item["finish"] = start
        
        json_obj["body"] = body_dict

        return json_obj

    def _project_dates(self, json_obj:dict) -> tuple:
        body_dict = json_obj["body"]
        
        start_list = [date["start"] for date in body_dict]
        finish_list = [date["finish"] for date in body_dict]

        earliest_start = self._bubble_sort_dates(start_list)[0]        
        latest_finish = self._bubble_sort_dates(finish_list)[-1]        

        return earliest_start, latest_finish

    def _bubble_sort_dates(self, unsorted_list:list) -> list:
        n = len(unsorted_list)

        for i in range(n):
            for j in range(n-i-1):
                value_j = unsorted_list[j]
                value_j1 = unsorted_list[j+1]

                try:
                    if isinstance(value_j, str) and isinstance(value_j1, str):
                        value_j = datetime.strptime(value_j, "%d-%b-%Y").date()
                        value_j1 = datetime.strptime(value_j1, "%d-%b-%Y").date()

                    elif isinstance(value_j1, datetime):
                        value_j = value_j.date()
                        value_j1 = value_j1.date()

                except Exception as e:
                    print(f"Error: {e}")
                else:
                    if value_j > value_j1:
                        unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]
        
        return unsorted_list

    def _build_nested_dic(self, json_obj:dict) -> dict:
        header_dict = json_obj["header"]
        body_dict = json_obj["body"]

        nested_dict = {
            "header": header_dict,
            "body": {},
        }

        def dict_management(obj, key_list, recursive_dict):
            if not key_list:
                return recursive_dict

            current_key = key_list.pop(0)
            obj_key = obj.get(current_key)

            if obj_key is not None and obj_key != "":
                if len(key_list) < 1:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = []
                    recursive_dict[obj_key].append(obj)
                else:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = {}
                    dict_management(obj, key_list, recursive_dict[obj_key])
            else:
                if "M_DATA" not in recursive_dict:
                    recursive_dict["M_DATA"] = []
                recursive_dict["M_DATA"].append(obj)

            return recursive_dict


        for obj in body_dict:
            keys = [category for category in self.json_struct_categories if obj.get(category) is not None]
            nested_dict["body"] = dict_management(
                obj, keys, nested_dict["body"]
            )
        
        return nested_dict

    def process_json(self, new_dict:dict, xlsx_dict:dict, xlsx_sheet:str="") -> None:
        basename = self.input_basename + '.' + self.input_extension 
        run_time = XlsxDataIngestion.calculate_time_duration(
            xlsx_dict.get("proc_start", ""), 
            xlsx_dict.get("proc_finish", "")
        )

        new_dict[xlsx_sheet] = {
            "details": {
                "workbook": os.path.join(self.input_path, basename),
                "worksheet": xlsx_sheet,
                "start_date": xlsx_dict.get("earliest_start", ""),
                "finish_date": xlsx_dict.get("latest_finish", ""),
                "entry_count": xlsx_dict.get("entry_count", ""),
            },
            "logs": {
                "start": xlsx_dict.get("proc_start", ""),
                "finish": xlsx_dict.get("proc_finish", ""),
                "run-time": run_time,
                "status": [],
            },
            "content": xlsx_dict.get("flattend_json", "")
        }


if __name__ == "__main__":
    XlsxDataIngestion.main(False)
