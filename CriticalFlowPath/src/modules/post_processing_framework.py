import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from CriticalFlowPath.keys.secrets import RSLTS_DIR

class PostProcessingFramework():
    def __init__(self, input_file_path, input_file_basename, input_file_extension, 
                 input_file_workweek,project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                 project_lead_struct, project_duration_processed, project_start_date, project_finish_date, time_scale):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.input_workweek = input_file_workweek
        self.worksheet_name = project_worksheet_name
        self.table = project_table
        self.ordered_dict = project_ordered_dict
        self.phase_order = project_phase_order
        self.lead_struct = project_lead_struct
        self.duration_processed = project_duration_processed
        self.start_date = project_start_date
        self.finish_date = project_finish_date
        self.time_scale = time_scale

        #Module Attributes
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'

        #Structures
        self.time_scale_options = ["d", "w"]
        self.json_struct_categories = [
            "phase", 
            "area", 
            "zone", 
            "trade", 
            "activity_code"
        ]
        self.wbs_final_categories = {
            "phase": "thin",
            "area": "dashed", 
            "zone": "dotted",
        }
        self.calendar_weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, input_file_extension=None, 
             input_file_workweek=None, project_worksheet_name=None, project_table=None, project_ordered_dict=None, project_phase_order=None, 
             project_lead_struct=None, project_duration_processed=None, project_start_date=None, project_finish_date=None, time_scale=None):
        if auto:
            project = PostProcessingFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_file_workweek, 
                project_worksheet_name,
                project_table, 
                project_ordered_dict, 
                project_phase_order, 
                project_lead_struct,
                project_duration_processed,
                project_start_date, 
                project_finish_date,
                time_scale
            )
        else:
            project = PostProcessingFramework.generate_ins()


        if project:
            if not project.time_scale:
                print("Please choose the time scale to process the schedule: ")
                scale_idx = PostProcessingFramework.display_options(project.time_scale_options)
                project.time_scale = project.time_scale_options[scale_idx[0] - 1]

            active_workbook, active_worksheet = project.return_excel_workspace(project.worksheet_name)

            if active_workbook and active_worksheet:
                if not project.time_scale:
                    project.update_schedule_size(
                        active_workbook, 
                        active_worksheet, 
                        project.ordered_dict, 
                        project.table, 
                        project.lead_struct,
                        project.time_scale
                    )

                project.update_schedule_style(
                    active_workbook, 
                    active_worksheet,
                    project.lead_struct, 
                    project.json_struct_categories
                )

    @staticmethod
    def generate_ins():
        PostProcessingFramework.ynq_user_interaction(
            "Run as Module as stand-alone? "
        )

        setup_cls = setup.Setup.main()

        ins = PostProcessingFramework(
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            setup_cls.obj["project"]["modules"]["MODULE_3"]["details"].get("workweek"),
            "CFA - WBS Table",
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("table"),
            setup_cls.obj["project"]["modules"]["MODULE_3"].get("content"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("custom_phase_order"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("lead_schedule_struct"),
            setup_cls.obj["project"]["modules"]["MODULE_3"]["details"]["calendar"]["processed"]["days"].get("total"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("start_date"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("finish_date"),
            None
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, 
                          input_file_workweek, project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                          project_lead_struct, project_duration_processed, project_start_date, project_finish_date, time_scale):
        ins = PostProcessingFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_file_workweek,
            project_worksheet_name,
            project_table, 
            project_ordered_dict, 
            project_phase_order, 
            project_lead_struct,
            project_duration_processed,
            project_start_date, 
            project_finish_date,
            time_scale
        )

        return ins

    @staticmethod
    def ynq_user_interaction(prompt_message):
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def binary_user_interaction(prompt_message:str) -> bool:
        valid_responses = {'y', 'n'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                if user_input == 'y':
                    return True 
                else:
                    return False 
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No]\n")

    @staticmethod
    def display_options(file_list:list) -> list:
        if not file_list:
            print('Error: No elements found.')
            return []

        print(f'-- {len(file_list)} elements found:')
        for idx, file in enumerate(file_list, start=1):
            print(f'{idx}. {file}')

        result = []
        selection_input = input('\nEnter index numbers (comma-separated) to select elements to process: ').split(',')

        for selection in selection_input:
            selection = selection.strip()
            if not selection.isdigit():
                print(f'Error: Invalid input "{selection}", skipping.')
                continue

            index = int(selection)
            if 1 <= index <= len(file_list):
                result.append(index)
            else:
                print(f'Error: "{index}" is out of range (1 to {len(file_list)}).')

        return result

    def return_excel_workspace(self, worksheet_name):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.worksheet_name = worksheet_name
                    print(f"New worksheet '{self.worksheet_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = PostProcessingFramework.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.worksheet_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.worksheet_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def update_schedule_size(self, active_workbook, active_worksheet, custom_ordered_dict:dict, 
                              proc_table, lead_schedule_struct:str) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        overlap_results = self._overlapping_dates(
            custom_ordered_dict, 
            lead_schedule_struct,
            self.time_scale
        )
        entry_dict_available = {}
        
        ref_point = list(proc_table.index.get_level_values(lead_schedule_struct))[0]
        count_ins = 0

        for area in proc_table.index.get_level_values(lead_schedule_struct):
            if area == ref_point:
                count_ins += 1
            else:
                entry_dict_available.setdefault(ref_point, []).append(count_ins)
                ref_point = area
                count_ins = 1

        entry_dict_available.setdefault(ref_point, []).append(count_ins)
        self._process_file(
            active_worksheet, 
            lead_schedule_struct, 
            overlap_results, 
            entry_dict_available
        )

        active_workbook.save(filename=file)
        print("Schedule Frame successfully updated")
        active_workbook.close()

    def _overlapping_dates(self, custom_ordered_dict:dict, lead_schedule_struct:str, 
                           alloted_space:int=2, time_scale:str='d') -> dict:
        lead_based_lists = []
        nested_list = []
        ref_lead = self._generate_compound_category_name(custom_ordered_dict[0])

        for item in custom_ordered_dict:
            comp_lead_cat_title = self._generate_compound_category_name(item)

            if comp_lead_cat_title == ref_lead:
                nested_list.append(item)
            else:
                lead_based_lists.append(nested_list)
                nested_list = [item]
                ref_lead = comp_lead_cat_title

        if nested_list:
            lead_based_lists.append(nested_list)

        sorted_entries_list = []
        for location_list in lead_based_lists:
            sorted_list = self._bubble_sort_entries_by_dates(location_list)
            sorted_entries_list.append(sorted_list)

        overlap_results = self._calculate_overlap(
            sorted_entries_list, 
            lead_schedule_struct, 
            alloted_space,
            time_scale
        )
        return overlap_results

    def _generate_compound_category_name(self, item: dict) -> str:
        category_names = []

        for category in self.wbs_final_categories.keys():
            cat_name = item.get(category)
            if cat_name:
                category_names.append(cat_name)

        return "|".join(category_names)

    def _bubble_sort_entries_by_dates(self, unsorted_list:list) -> list:
        n = len(unsorted_list)

        for i in range(n):
            for j in range(0, n-i-1):
                date_typed_start = datetime.strptime(unsorted_list[j]["start"], "%d-%b-%Y")
                date_typed_start_n1 = datetime.strptime(unsorted_list[j+1]["start"], "%d-%b-%Y")

                if date_typed_start > date_typed_start_n1:
                    unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]

                elif date_typed_start == date_typed_start_n1:
                    date_typed_end = datetime.strptime(unsorted_list[j]["finish"], "%d-%b-%Y")
                    date_typed_end_n1 = datetime.strptime(unsorted_list[j+1]["finish"], "%d-%b-%Y")

                    if date_typed_end > date_typed_end_n1:
                        unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]
        
        sorted_list = unsorted_list
        return sorted_list

    def _calculate_overlap(self, location_based_lists:list, lead_schedule_struct:str, 
                           alloted_space:int, time_scale:str) -> dict:
        overlap = {}

        if time_scale == 'd':
            overlap = self._day_based_overlap(location_based_lists, lead_schedule_struct, alloted_space)
        elif time_scale == 'w':
            overlap = self._week_based_overlap(location_based_lists, lead_schedule_struct, alloted_space)            

        return overlap
        
    def _day_based_overlap(self, location_based_lists:list, lead_schedule_struct:str, 
                           alloted_space:int) -> dict:
        overlap_results = []

        for lead_list in location_based_lists:
            active_overlaps = []
            max_overlap = 0
            ref_point = lead_list[0][lead_schedule_struct]

            for activity in lead_list:
                start = datetime.strptime(activity["start"], "%d-%b-%Y")
                finish = datetime.strptime(activity["finish"], "%d-%b-%Y")

                # Remove activities that have ended
                active_overlaps = [
                    (active_start, active_finish) 
                    for active_start, active_finish in active_overlaps 
                    if finish >= active_start and start <= active_finish
                ]

                # Add the current activity to active overlaps
                active_overlaps.append((start, finish))

                # Update the maximum overlap count
                max_overlap = max(max_overlap, len(active_overlaps))

            overlap_results.append((ref_point, max_overlap))

        max_rows_dict = {}
        for lead_cat, max_row in overlap_results:
            if lead_cat in max_rows_dict:
                max_rows_dict[lead_cat].append(max_row + alloted_space)
            else:
                max_rows_dict[lead_cat] = [max_row + alloted_space]

        return max_rows_dict
    
    def _week_based_overlap(self, location_based_lists: list, 
                           lead_schedule_struct: str, 
                           alloted_space: int) -> dict:
        overlap_results = []

        for lead_list in location_based_lists:
            if not lead_list:
                continue

            active_overlaps = []
            max_overlap = 0
            ref_point = lead_list[0].get(lead_schedule_struct, "UNKNOWN")

            for activity in lead_list:
                try:
                    start_date = datetime.strptime(activity["start"], "%d-%b-%Y")
                    finish_date = datetime.strptime(activity["finish"], "%d-%b-%Y")
                except (KeyError, ValueError) as e:
                    print(f"Invalid date format in activity: {e}")
                    continue

                week_start, week_finish = self._calculate_week(start_date, finish_date)

                # Filter out non-overlapping activities
                active_overlaps = [
                    (active_start, active_finish) 
                    for active_start, active_finish in active_overlaps 
                    if week_finish >= active_start and week_start <= active_finish
                ]

                active_overlaps.append((week_start, week_finish))
                max_overlap = max(max_overlap, len(active_overlaps))

            if ref_point != "UNKNOWN":
                overlap_results.append((ref_point, max_overlap))

        # Process results into output dictionary
        max_rows_dict = {}
        for lead_cat, max_row in overlap_results:
            max_rows_dict.setdefault(lead_cat, []).append(max_row + alloted_space)

        return max_rows_dict
    
    def _calculate_week(self, start_date:datetime, 
                       finish_date:datetime) -> tuple[datetime, datetime]:

        def adjust_to_weekday(target_date, target_weekday):
            while self.calendar_weekdays[target_date.weekday()] != target_weekday:
                if self.calendar_weekdays[target_date.weekday()] < target_weekday:
                    target_date += timedelta(days=1)
                else:
                    target_date -= timedelta(days=1)
            return target_date

        week_start = adjust_to_weekday(start_date, self.input_workweek[0])
        week_finish = adjust_to_weekday(finish_date, self.input_workweek[-1])

        if week_start > finish_date:
            week_start = adjust_to_weekday(finish_date, self.input_workweek[0])
        if week_finish < start_date:
            week_finish = adjust_to_weekday(start_date, self.input_workweek[-1])

        return (week_start, week_finish)
    
    def _process_file(self, active_worksheet, lead_schedule_struct:str, 
                     overlap_results:dict, con_available_ins:dict) -> None:
        ws = active_worksheet

        print(f"Processing Excel file: {self.input_basename}")
        
        wbs_start_col = column_index_from_string(self.wbs_start_col)
        start_col_idx = self._find_column_idx(ws, lead_schedule_struct, self.wbs_start_row) + 1
        end_col_idx = self._return_first_column_idx(ws, 1, 1)

        self._unmerge_columns(ws, wbs_start_col, end_col_idx)
        self._delete_excess_rows(
            ws, 
            lead_schedule_struct, 
            overlap_results, 
            con_available_ins
        )
        self._delete_columns(ws, start_col_idx, end_col_idx)
        self._insert_columns(ws, start_col_idx, end_col_idx - start_col_idx)
    
    def _find_column_idx(self, active_worksheet, column_header:str, start_row:int) -> int|None:
        ws = active_worksheet
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.strip().replace(" ", "_").lower()
        
        for row in ws.iter_rows(
            min_row=start_row,
            min_col=start_col_idx,
            max_col=active_worksheet.max_column
        ):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.strip().replace(" ", "_").lower()
                    if normalized_header == normalized_cell_value:
                        return cell.column
        
        if PostProcessingFramework.binary_user_interaction(
            f"Column header '{column_header}' not found. Use default column 'A'?: "
        ):
            return 0
        return None

    def _return_first_column_idx(self, active_worksheet, start_row_idx:int, start_col_idx:int):
        ws = active_worksheet

        for row in ws.iter_rows(min_row=start_row_idx, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    return cell.column
        
        print(f"Error. No valid value encountered in specified row [{start_row_idx}, {start_row_idx}]")
        return -1

    def _unmerge_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        for merged_cell in list(ws.merged_cells.ranges):
            if (merged_cell.bounds[0] >= start_column_idx and 
                merged_cell.bounds[2] <= end_column_idx):
                ws.unmerge_cells(str(merged_cell))
        
        start_col_letter = get_column_letter(start_column_idx)
        end_col_letter = get_column_letter(end_column_idx)

        print(f"Column unmerging applied successfully: [{start_col_letter}, {end_col_letter}]")

    def _delete_excess_rows(self, active_worksheet, column_header:str, 
                            overlap_results:dict, con_available_ins:dict) -> None:
        ws = active_worksheet
        header_col = self._find_column_idx(active_worksheet, column_header, self.wbs_start_row)
        all_current_locations = list(overlap_results.keys())

        rows_to_delete = []
        init_count = 0
        max_rows = 0
        for row in ws.iter_rows(min_row=self.wbs_start_row + 1,
                                max_row=ws.max_row,
                                min_col=header_col,
                                max_col=header_col):
            for cell in row:
                if cell.value is not None:  
                    init_count = 0
                    if cell.value in all_current_locations:
                        if len(overlap_results[cell.value]) > 1:
                            if overlap_results[cell.value][0] > con_available_ins[cell.value][0]:
                                max_rows = con_available_ins[cell.value][0]
                                con_available_ins[cell.value].pop(0)
                            else:
                                max_rows = overlap_results[cell.value][0]
                                overlap_results[cell.value].pop(0)
                        else:
                            if overlap_results[cell.value][0] > con_available_ins[cell.value][0]:
                                max_rows = con_available_ins[cell.value][0]
                            else:
                                max_rows = overlap_results[cell.value][0]

                if init_count > max_rows:
                    rows_to_delete.append(cell.row)
                
            init_count += 1

        for row in sorted(set(rows_to_delete), reverse=True):
            ws.delete_rows(row)
        
        print("Excess rows removed successfully.")

    def _delete_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        amount_to_delete = end_column_idx - start_column_idx
        ws.delete_cols(start_column_idx, amount_to_delete)

        start_col_letter = get_column_letter(start_column_idx)
        end_col_letter = get_column_letter(end_column_idx)

        print(f"Columns deleted successfully: [{start_col_letter}, {end_col_letter}]")

    def _insert_columns(self, active_worksheet, start_column_idx:int, num_cols:int) -> None:
        ws = active_worksheet

        for _ in range(num_cols):
            ws.insert_cols(start_column_idx)

        print(f"Columns ({num_cols}) added successfully")

    def update_schedule_style(self, active_workbook, active_worksheet, lead_schedule_struct:str, json_struct_categories:list) -> None:
        wb = active_workbook
        ws = active_worksheet
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        point_col_idx = self._find_column_idx(
            ws, 
            lead_schedule_struct, 
            self.wbs_start_row
        )
        wbs_keys = list(self.wbs_final_categories.keys())
        lead_cat_columns = wbs_keys.index(lead_schedule_struct)
        relevant_keys = wbs_keys[:lead_cat_columns + 1]
        
        column_idxs = {
            item: self._find_column_idx(ws, item, self.wbs_start_row)
            for item in json_struct_categories
            if item in relevant_keys
        }
        start_col_idx = point_col_idx + 1
        end_col_idx = self._return_first_column_idx(ws, 1, 1)

        self._delete_columns(ws, start_col_idx, end_col_idx)

        for _, value in column_idxs.items():
            self._merge_until_different_value(
                ws, value, self.wbs_start_row
            )

        self._style_file(ws, lead_schedule_struct, start_col_idx)

        wb.save(filename=file)
        print("WBS Table successfully updated")
        wb.close()

    def _merge_until_different_value(self, active_worksheet, starting_col_idx:int, 
                                     starting_row_idx:int) -> None:
        ws = active_worksheet

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        current_merge_range = []
        merge_ranges = []

        for row_idx in range(starting_row_idx, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=starting_col_idx)

            if cell.value is not None:
                if current_merge_range:
                    merge_ranges.append(current_merge_range)
                current_merge_range = [cell]
            else:
                current_merge_range.append(cell)

        if current_merge_range:
            merge_ranges.append(current_merge_range)

        for merge_range in merge_ranges:
            if len(merge_range) > 1:
                start_row = merge_range[0].row
                end_row = merge_range[-1].row

                ws.merge_cells(
                    start_row=start_row,
                    start_column=starting_col_idx,
                    end_row=end_row,
                    end_column=starting_col_idx
                )

                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=starting_col_idx)
                    cell.border = thin_border

    def _style_file(self, active_worksheet, lead_schedule_struct:str, 
                    start_col:int, start_row:int=1) -> None:
        ws = active_worksheet
        
        starting_day_of_the_week = ws[get_column_letter(start_col)+str(self.wbs_start_row)].value

        year_list, year_row = self._same_week_based_values(
            ws, 
            start_col, 
            start_row, 
            starting_day_of_the_week
        )
        for nested_list in year_list:
            self._merge_same_value_cells(
                ws, 
                list(nested_list.keys()), 
                list(nested_list.values()), 
                year_row
            ) 

        month_list, month_row = self._same_week_based_values(
            ws, 
            start_col, 
            start_row + 1, 
            starting_day_of_the_week
        )
        for nested_list in month_list:
            self._merge_same_value_cells(
                ws, 
                list(nested_list.keys()), 
                list(nested_list.values()), 
                month_row
            )  

        self._apply_color_format(
            ws,
            start_col, 
            start_row
        )
        self._apply_post_styling(ws)
        
        wbs_keys = list(self.wbs_final_categories.keys())
        lead_cat_columns = wbs_keys.index(lead_schedule_struct)
        relevant_keys = wbs_keys[:lead_cat_columns + 1]

        for categroy in reversed(relevant_keys):
            cat_value = self.wbs_final_categories.get(categroy)
            self._apply_horizontal_sectioning(
                ws, 
                categroy, 
                cat_value, 
                start_col
            )

    def _same_week_based_values(self, active_worksheet, starting_col_idx:int, 
                                starting_row_idx:int, starting_day:str):
        ws = active_worksheet
        iterable_row = ws[starting_row_idx]

        starting_week_day = self.calendar_weekdays.index(starting_day)
        day_count = 1 + starting_week_day
        same_week_list = {}
        overall_list = []

        for cell in iterable_row[starting_col_idx-1:]:
            if cell.value is None:
                break

            if day_count > len(self.input_workweek):
                overall_list.append(same_week_list)
                same_week_list = {}
                day_count = 1
            
            same_week_list[cell.column] = cell.value
            day_count += 1

        if same_week_list:
            overall_list.append(same_week_list)

        return overall_list, starting_row_idx

    def _merge_same_value_cells(self, active_worksheet, columns_list:list, 
                                values_list:list, starting_row_idx:int) -> None:
        ws = active_worksheet

        if not columns_list:
            print("No columns to merge.")
            return

        last_value = values_list[0]
        overall_list = []
        same_val_list = []

        for col, val in zip(columns_list, values_list):
            if val == last_value:
                same_val_list.append(col)
            else:
                overall_list.append(same_val_list)
                same_val_list = [col]
                last_value = val

        if same_val_list:
            overall_list.append(same_val_list)

        for group in overall_list:
            if len(group) > 1:
                first_col = group[0]
                last_col = group[-1]
                ws.merge_cells(
                    start_row=starting_row_idx,
                    start_column=first_col,
                    end_row=starting_row_idx,
                    end_column=last_col
                )

    def _apply_color_format(self, active_worksheet, start_col:int, start_row:int) -> None:
        ws = active_worksheet
    
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row, 
            self.duration_processed, 
            "00C0C0C0", 
            "00F2F2F2"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 1, 
            self.duration_processed, 
            "00C0C0C0", 
            "00D9D9D9"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 2, 
            self.duration_processed, 
            "00C0C0C0", 
            "00FFFFFF"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 3, 
            self.duration_processed, 
            "00C0C0C0", 
            "00FFFFFF"
        )
    
        self._apply_vertical_sectioning(
            ws, 
            start_col, 
            start_row + 1, 
            self.duration_processed
        )

        print("Workbook styled and saved successfully.")

    def _paint_schedule_row(self, active_worksheet, start_col:str, start_row:int, 
                            duration:int, border_color:str, fill_color:str) -> None:
        ws = active_worksheet
        for row in ws.iter_rows(min_row=start_row, 
                                 max_row=start_row, 
                                 min_col=start_col, 
                                 max_col=start_col + duration - 1):
            for cell in row:  
                cell.border = Border(
                    top=Side(border_style="thin", color=border_color), 
                    left=Side(border_style="thin", color=border_color),  
                    right=Side(border_style="thin", color=border_color),  
                    bottom=Side(border_style="thin", color=border_color)  
                )
                cell.font = Font(
                    name='Century Gothic', 
                    size=12, 
                    bold=False, 
                    color="00000000"
                )  
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    def _apply_vertical_sectioning(self, active_worksheet, start_col:str, 
                                   start_row:int, duration:int) -> None:
        ws = active_worksheet
        last_header = ws[get_column_letter(start_col)+str(start_row)].value

        for col in range(start_col, start_col + duration):
            col_header = ws.cell(row=start_row, column=col).value

            if col_header and col_header != last_header:
                for row in range(start_row + 1, ws.max_row + 1):  
                    cell = ws.cell(row=row, column=col)

                    current_border = cell.border if cell.border else Border()
                    cell.border = Border(
                        top=current_border.top, 
                        left=Side(border_style="dashed"),  
                        right=current_border.right,  
                        bottom=current_border.bottom  
                    )

                last_header = col_header

    def _apply_post_styling(self, active_worksheet) -> None:
        ws = active_worksheet
        first_row = ws[self.wbs_start_row]
        
        if first_row is None:
            print("Error: The first row is empty.")
            return
        
        for cell in first_row:  
            cell.font = Font(
                name='Century Gothic', 
                size=12, 
                bold=False, 
                color="FFFFFF"
            )  
            cell.fill = PatternFill(start_color="00800080", end_color="00800080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")  
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)  
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.rows:    
            if isinstance(row[0].value, int):
                parent_row = row
                for cell in parent_row:
                    cell.font = Font(
                        name='Century Gothic', 
                        size=12, 
                        bold=False, 
                        color="00333333"
                    )  
                    cell.fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(bottom=thin)

        print(f"Post-styling successfully completed")

    def _apply_horizontal_sectioning(self, active_worksheet, section_header:str, 
                               border_style:str, start_col:int) -> None:
        ws = active_worksheet
        header_idx = self._find_column_idx(ws, section_header, self.wbs_start_row)

        section_list = self._list_cell_values(ws, header_idx)
        last_valid_section = section_list[0]

        for row_idx, row in enumerate(ws.iter_rows(min_col=start_col, 
                                                   max_col=ws.max_column, 
                                                   min_row=self.wbs_start_row + 1, 
                                                   max_row=ws.max_row)):
            current_section = section_list[row_idx]

            if current_section and current_section != last_valid_section:
                if current_section is None:
                    continue
                else:
                    self._style_row_border(row, border_style)
                    last_valid_section = current_section

        print("Post sectioning applied successfully.")

    def _list_cell_values(self, active_worksheet, col_idx:int) -> list:
        ws = active_worksheet
        col_list = []

        if col_idx is not None:
            for row in ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                    min_col=col_idx, 
                                    max_col=col_idx, 
                                    max_row=ws.max_row):
                for cell in row:
                    col_list.append(cell.value)
        
        return col_list 

    def _style_row_border(self, row, border_style:str) -> None:
        if border_style == "no_border":
            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=current_border.top, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border
        else:
            new_top_border = Side(border_style=border_style, color='000000')

            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=new_top_border, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border


if __name__ == "__main__":
    PostProcessingFramework.main(False)
