import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
""" from .setup import Setup """

# Imported Helper - As Package 
from modules.setup import Setup

class ScheduleFramework():
    def __init__(self, input_file_path, input_file_basename, input_file_extension, project_worksheet_name,
                 project_table, project_ordered_dict, project_phase_order, project_lead_struct, 
                 project_start_date, project_finish_date, input_start_row=1, input_start_col=""):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.worksheet_name = project_worksheet_name
        self.table = project_table
        self.ordered_dict = project_ordered_dict
        self.phase_order = project_phase_order
        self.lead_struct = project_lead_struct
        self.start_date = project_start_date
        self.finish_date = project_finish_date
        self.start_row = int(input_start_row)
        self.start_col = str(input_start_col)
        
        #Module Attributes
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.dark_default_hex_font = "00000000"
        self.light_default_hex_font = "00FFFFFF"
        self.default_hex_fill_color = "00FFFF00"

        #Structures
        self.wbs_final_categories = {
            "phase": "thick",
            "location": "no_border", 
            "area": "dashed",
        }
        self.allowed_headers = {
            "hyperlinks": ["activity_code", "activity_name", "wbs_code"]
        }
        self.calendar_weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

        #Instance Results
        self.initial_project_type:str = None
        self.final_project_type:any = None
        self.final_project_dict:dict = None

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, input_file_extension=None, 
             project_worksheet_name=None, project_table=None, project_ordered_dict=None, project_phase_order=None, 
             project_lead_struct=None, project_start_date=None, project_finish_date=None):
        if auto:
            project = ScheduleFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension, 
                project_worksheet_name,
                project_table, 
                project_ordered_dict, 
                project_phase_order, 
                project_lead_struct, 
                project_start_date, 
                project_finish_date,
            )
        else:
            project = ScheduleFramework.generate_ins()
            project_details = Setup.main(False)

        if project:
            active_workbook, active_worksheet = project.return_excel_workspace(project.worksheet_name)

            if active_workbook and active_worksheet:
                project.create_schedule(
                    active_workbook, 
                    active_worksheet, 
                    project.table, 
                    project.ordered_dict
                )
            else:
                print("Error. Could not open Excel file as Workbook & Worksheet")

        return project

    @staticmethod
    def generate_ins():
        input_excel_file = input("Please enter the path to the Excel file or directory: ")
        input_excel_path, input_excel_basename = ScheduleFramework.file_verification(
            input_excel_file, 'e', 'r')
        input_worksheet_name = input("Please enter the name for the new or existing worksheet: ")
        input_start_row = input("Please enter the starting row to write the schedule: ")
        input_start_col = input("Please enter the starting column to write the schedule: ")
        input_json_file = input("Please enter the path to the Json file or directory: ")
        input_json_path, input_json_basename = ScheduleFramework.file_verification(
            input_json_file, 'j', 'r')
        input_start_date = input("Please enter the start date for the schedule (format: dd-MMM-yyyy): ")
        input_end_date = input("Please enter the end date for the schedule (format: dd-MMM-yyyy): ")

        ins = ScheduleFramework(input_excel_path, input_excel_basename, input_json_path, 
                                input_json_basename, input_worksheet_name, input_start_row, 
                                input_start_col, input_start_date, input_end_date)

        return ins
    
    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, 
                          project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                          project_lead_struct, project_start_date, project_finish_date):
        ins = ScheduleFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension, 
            project_worksheet_name,
            project_table, 
            project_ordered_dict, 
            project_phase_order, 
            project_lead_struct, 
            project_start_date, 
            project_finish_date,
        )

        return ins

    @staticmethod
    def file_verification(input_file_path, file_type, mode, input_json_title=None):
        if input_json_title and os.path.isdir(input_file_path):
            file_basename = f"processed_{ScheduleFramework.normalize_string(input_json_title)}.json"
            path, basename = ScheduleFramework.handle_file(input_file_path, file_basename, file_type)
        else:
            if os.path.isdir(input_file_path):
                file_path, file_basename = ScheduleFramework.handle_dir(input_file_path, mode)
                if mode != 'c':
                    path, basename = ScheduleFramework.handle_file(file_path, file_basename, file_type)
                else:
                    path = file_path
                    basename = file_basename
            elif os.path.isfile(input_file_path):
                file_path = os.path.dirname(input_file_path)
                file_basename = os.path.basename(input_file_path)
                path, basename = ScheduleFramework.handle_file(file_path, file_basename, file_type)

        return path, basename
    
    @staticmethod
    def handle_dir(input_path, mode):
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_path)
            selection = ScheduleFramework.display_directory_files(dir_list)
            base_name = dir_list[selection]
            print(f'File selected: {base_name}\n')
        elif mode == 'c':
            base_name = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return input_path, base_name

    @staticmethod
    def handle_file(file_path, file_basename, file_type):
        file = os.path.join(file_path, file_basename)

        if (file_type == 'e' and ScheduleFramework.is_xlsx(file)) or \
           (file_type == 'j' and ScheduleFramework.is_json(file)):
            return os.path.dirname(file), os.path.basename(file)
        
        print("Error: Please verify that the directory and file exist and that the file is of type .xlsx or .json")
        return -1

    @staticmethod
    def display_directory_files(list):
        selection_idx = 0
        if len(list)==0:
            print('Error. No files found')
            return -1
        
        if len(list)>1:
            print(f'-- {len(list)} files found:')
            idx = 0
            for file in list:
                idx += 1
                print(f'{idx}. {file}')

            selection_idx = input('\nPlease enter the index number to select the one to process: ') 
        else:
            print(f'Single file found: {list[0]}')
            print('Will go ahead and process')

        return int(selection_idx) - 1

    @staticmethod
    def is_json(file_name):
        if file_name.endswith(".json"):
            return True
        else:
            print("Error: Selected file is not a JSON file")
            return False

    @staticmethod
    def is_xlsx(file_name):
        if file_name.endswith('.xlsx'):
            return True
        else:
            print('Error. Selected file is not an Excel')
            return False

    @staticmethod
    def hex2rgb(hex_color):
        trimmed_hex = hex_color.lstrip('#')
        calc_rgb = tuple(int(trimmed_hex[i:i+2], 16) for i in (0, 2, 4))

        return calc_rgb

    @staticmethod
    def calculateLuminance(R, G, B):
        lum = 0.2126*(R/255.0)**2.2 + 0.7152*(G/255.0)**2.2 + 0.0722*(B/255.0)**2.2

        return lum

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    def return_excel_workspace(self, worksheet_name):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    print(f"New worksheet '{worksheet_name}' created.")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws = ScheduleFramework.display_directory_files(ws_list)
                    worksheet = workbook.worksheets[selected_ws]
                    return workbook, worksheet
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def create_schedule(self, active_workbook, active_worksheet, proc_table, 
                        json_dict:list) -> dict:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        if self.start_col == "" or self.start_col is None:
            self.start_col = get_column_letter(self._find_column_idx(active_worksheet, 'finish') + 1)
        
        custom_ordered_dict = {val["entry"]: val for val in json_dict}

        self.generate_schedule_frame(active_worksheet, self.start_date, self.finish_date)
        reworked_custom_ordered_dict = self.apply_schedule_cfa_style(
            active_worksheet, custom_ordered_dict, proc_table
        )
        
        active_workbook.save(filename=file)
        print("CFA Schedule successfully created")
        active_workbook.close()

        return reworked_custom_ordered_dict

    def _find_column_idx(self, active_ws, column_header:str):
        ws = active_ws
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.replace(" ", "_").lower()

        for row in ws.iter_rows(min_row=self.wbs_start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def generate_schedule_frame(self, active_worksheet, start_date:str, end_date:str) -> None:
        ws = active_worksheet

        start_datetime_obj = datetime.strptime(start_date, '%d-%b-%Y')
        end_datetime_obj = datetime.strptime(end_date, '%d-%b-%Y')
        duration = (end_datetime_obj - start_datetime_obj).days + 1  
        
        self._fill_schedule_row(
            ws, start_datetime_obj, duration, self.start_row, self.start_col, '%Y'
        )
        self._fill_schedule_row(
            ws, start_datetime_obj, duration, self.start_row+1, self.start_col, '%b'
        )
        self._fill_schedule_row(
            ws, start_datetime_obj, duration, self.start_row+2, self.start_col, '%d'
        )
        self._fill_schedule_row(
            ws, start_datetime_obj, duration, self.start_row+3, self.start_col, "", True
        )

        print("Schedule frame generated successfully.")

    def _fill_schedule_row(self, active_worksheet, start_datetime_obj:datetime, duration:int, 
                           start_row:int, start_col:str, format:str, week_days:bool=False) -> None:
        if week_days:
            for day in range(duration):
                date = start_datetime_obj + timedelta(days=day)
                active_worksheet.cell(
                    row=start_row, 
                    column=column_index_from_string(start_col) + day, 
                    value=self.calendar_weekdays[date.weekday()]
                )
        else:
            for day in range(duration):
                date = start_datetime_obj + timedelta(days=day)
                active_worksheet.cell(
                    row=start_row, 
                    column=column_index_from_string(start_col) + day, 
                    value=date.strftime(format)
                )

    def apply_schedule_gantt_style(self, active_worksheet, custom_ordered_dict:dict, proc_table):
        ws = active_worksheet

        start_ovr_date = datetime.strptime(self.start_date, "%d-%b-%Y")
        final_ovr_date = datetime.strptime(self.finish_date, "%d-%b-%Y")

        for idx, value in enumerate(proc_table.index.get_level_values("entry")):
            item = custom_ordered_dict[value]
            initial_date = datetime.strptime(item["start"], "%d-%b-%Y")
            final_date = datetime.strptime(item["finish"], "%d-%b-%Y")

            if not (start_ovr_date <= initial_date <= final_ovr_date and start_ovr_date <= final_date <= final_ovr_date):
                continue

            start_search_col = column_index_from_string(self.start_col) + (initial_date - start_ovr_date).days
            finish_search_col = column_index_from_string(self.start_col) + (final_date - start_ovr_date).days

            count = 0
            for col in range(start_search_col, finish_search_col + 1):
                cell = ws.cell(row=self.wbs_start_row + idx + 1, column=col)
                self._style_cell(cell, item)
                if count < 1:
                    self._add_comment(cell, item)

                count += 1

        print("Workbook filled successfully.")

    def apply_schedule_cfa_style(self, active_worksheet, custom_ordered_dict:dict, 
                                proc_table) -> dict:
        ws = active_worksheet
        schedule_setup = {
            "start_ovr_date": datetime.strptime(self.start_date, "%d-%b-%Y"),
            "final_ovr_date": datetime.strptime(self.finish_date, "%d-%b-%Y"),
            "starting_point": self.wbs_start_row + 1,
            "ref_lead": 0,
            "occupied_rows": {},
            "completed_tasks":  set(),
        }


        self._paint_structured_schedule(
            ws, custom_ordered_dict, proc_table, schedule_setup
        )

        print("Workbook filled successfully.")
        return custom_ordered_dict

    def _paint_structured_schedule(self, active_worksheet, custom_ordered_dict:dict, proc_table, 
                                   schedule_setup:dict) -> None:
        ws = active_worksheet
        start_ovr_date = schedule_setup.get("start_ovr_date")
        final_ovr_date = schedule_setup.get("final_ovr_date")
        starting_point = schedule_setup.get("starting_point")
        ref_lead = schedule_setup.get("ref_lead")
        occupied_rows = schedule_setup.get("occupied_rows")
        completed_tasks = schedule_setup.get("completed_tasks")

        for idx, value in enumerate(proc_table.index.get_level_values("entry")):
            item = custom_ordered_dict[value]
            current_lead = self._generate_compound_category_name(item)
            initial_date = datetime.strptime(item["start"], "%d-%b-%Y")
            final_date = datetime.strptime(item["finish"], "%d-%b-%Y")

            if not (start_ovr_date <= initial_date <= final_ovr_date and start_ovr_date <= final_date <= final_ovr_date):
                continue

            start_search_col = column_index_from_string(self.start_col) + (initial_date - start_ovr_date).days
            finish_search_col = column_index_from_string(self.start_col) + (final_date - start_ovr_date).days

            if current_lead != ref_lead:
                starting_point = max(starting_point, self.wbs_start_row + idx + 1)
                ref_lead = current_lead
                occupied_rows = {}

            target_row = starting_point
            while any(get_column_letter(col) in occupied_rows.get(target_row, set()) for col in range(start_search_col, finish_search_col + 1)):
                target_row += 1
                if target_row not in occupied_rows:
                    occupied_rows[target_row] = set()

            for column in range(start_search_col, finish_search_col + 1):
                col_letter = get_column_letter(column)
                occupied_rows.setdefault(target_row, set()).add(col_letter)

            count = 0
            original_sequence = []
            for col in range(start_search_col, finish_search_col + 1):
                cell = ws.cell(row=target_row, column=col)

                if item.get("predecessor"):
                    self._style_cell(cell, item, True)
                else:
                    self._style_cell(cell, item)

                if count == 0:
                    self._add_comment(cell, item)
                    count += 1

                original_sequence.append(cell.coordinate)
                
            item["cell_sequence"] = {
                "original": original_sequence,
                "reworked": None
            }
            completed_tasks.add(value)

        missing_tasks = set(proc_table.index.get_level_values("entry")) - completed_tasks
        if missing_tasks:
            print(f"Warning: The following tasks were not painted: {missing_tasks}")

    def _generate_compound_category_name(self, item:dict) -> str:
        category_names = []

        for category in self.wbs_final_categories.keys():
            cat_name = item.get(category)
            if cat_name:
                category_names.append(cat_name)

        return "|".join(category_names)

    def _style_cell(self, cell, current_item:dict, paint_border:bool=False) -> None:
        try:
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if paint_border:
                cell.border = Border(
                    top=Side(border_style="double", color="00FF0000"), 
                    left=Side(border_style="double", color="00FF0000"), 
                    right=Side(border_style="double", color="00FF0000"), 
                    bottom=Side(border_style="double", color="00FF0000"))
            
            color = current_item.get("color")
            if color:
                cell.fill = PatternFill(start_color=re.sub('#', "00", current_item.get("color")), 
                                    end_color=re.sub('#', "00", current_item.get("color")), 
                                    fill_type='solid')
            else:
                print(f"Color hex not found for: {current_item['entry']}")
                cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                    end_color=self.default_hex_fill_color, 
                                    fill_type='solid')
            
            rgb = ScheduleFramework.hex2rgb(color)
            lum_rgb = ScheduleFramework.calculateLuminance(rgb[0], rgb[1], rgb[2])
            if lum_rgb > 0.5:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.dark_default_hex_font)
            else:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.light_default_hex_font)

            cell.value = current_item.get("activity_code") if current_item.get("activity_code") else "NaN"
            
        except Exception as e:
            print(f"Error. While painting cell: {e}")

    def _add_comment(self, cell, current_item:dict) -> None:
        msg = f"""
            Activity Name: {current_item.get('activity_name', "NaN")}
            Phase: {current_item.get("phase", "NaN")} 
            Location: {current_item.get("location", "NaN")}
            Trade: {current_item.get("trade", "NaN")} 

            Start Date: {current_item.get("start", "NaN")}
            End Date: {current_item.get("finish", "NaN")}
        """

        comment = Comment(msg, "AMP")
        comment.width = 300
        comment.height = 200

        cell.comment = comment


if __name__ == "__main__":
    project = ScheduleFramework.main(False)
