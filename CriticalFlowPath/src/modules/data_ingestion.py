import os
import re
import json
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime, date
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class DataIngestion:
    def __init__(self, input_file_path, input_file_basename, input_file_extension, output_file_dir, 
                 input_project_code, input_project_title, input_project_subtitle, input_project_client, 
                 input_project_dates_created, xlsx_start_row=1, xlsx_start_col='A'):
        #Inherited attributes
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.output_path = output_file_dir
        self.output_json_basename = DataIngestion.normalize_string(input_project_title)
        self.project_code = input_project_code
        self.project_title = input_project_title
        self.project_subtitle = input_project_subtitle
        self.project_client = input_project_client
        self.file_created_at = input_project_dates_created
        self.xlsx_start_row = xlsx_start_row
        self.xlsx_start_col = xlsx_start_col

        #Module attributes
        self.ws_name = None

        #Structures
        self.project_content_headers = {
            "entry": None, 
            "phase": ["phase"],
            "location": ["location"], 
            "area": ["area"],
            "trade": ["trade"], 
            "color": ["color"],
            "parent_id": None, 
            "activity_code": ["activity_code", "code", "task_code", "act_code"],
            "activity_id": None,
            "activity_name": ["activity_name", "act_name"], 
            "activity_status": ["activity_status", "status", "task_status"], 
            "activity_ins": None, 
            "start": ["start", "start_date", "start_dates"], 
            "finish": ["finish", "finish_date", "finish_dates", "end", "end_date"], 
            "total_float": ["total_float"],
            "activity_successor_id": ["successor"],
            "activity_predecessor_id": ["predecessor"],
        }
        self.json_struct_categories = ["phase", "location", "area", "trade", "activity_code"]

    @staticmethod
    def main(auto=True, process_continuity=None, input_file_path=None, input_file_basename=None, 
             input_file_extension=None, output_file_dir=None, input_project_code=None, input_project_title=None,
             input_project_subtitle=None, input_project_client=None, input_project_dates_created=None):
        if auto:
            project = DataIngestion.auto_generate_ins(
                process_continuity, input_file_path, input_file_basename, input_file_extension, 
                output_file_dir, input_project_code, input_project_title, input_project_subtitle, 
                input_project_client, input_project_dates_created
            )
        else:
            project = DataIngestion.generate_ins()
 
        module_data = {
            "logs": {
                "start": project.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            },
            "content": {}
        }

        if project.input_extension == "xlsx":
            if project.ws_name is None:
                worksheet = input("Please enter the name for the new or existing worksheet: ")
            else:
                worksheet = project.ws_name

            reworked_json, nested_json = project.handle_xlsx(worksheet)
            module_data["content"] = nested_json
        elif project.input_extension == "xml":
            excel_path, excel_basename = DataIngestion.return_valid_file(project.output_file_dir)
            reworked_json = project.handle_xml()
            module_data["content"] = project.create_wbs_table_to_fill(reworked_json, excel_path, excel_basename)
        else:
            print("Error. Unsuported file type")
            module_data["logs"]["status"].append((
                DataIngestion.__name__, 
                dict(Error=f"Unsuported input file extension: {project.input_extension}")
            ))

        module_data["logs"]["finish"] = project.return_valid_date()
        module_data["logs"]["run-time"] = DataIngestion.calculate_time_duration(
            module_data.get("start"), 
            module_data.get("finish")
        )

        return module_data

    @staticmethod
    def generate_ins():
        input_file = DataIngestion.return_valid_file(
            input("Please enter the path to the file or directory: ").strip()
        )
        input_file_path = input_file.get("path")
        input_file_basename = input_file.get("basename")
        input_file_extension = input_file.get("extension")
        output_file_dir = DataIngestion.return_valid_path(
            "Please enter the directory to save the new module results: "
        )
        input_project_code = input("Enter Project Code: ").strip()
        input_project_title = input("Enter Project Title: ").strip()
        input_project_subtitle = input("Enter Project Subtitle: ").strip()
        input_project_client = input("Enter Project Client: ").strip()
        project_dates_created = DataIngestion.return_valid_date()

        ins =  DataIngestion(input_file_path, input_file_basename, input_file_extension, 
                             output_file_dir, input_project_code, input_project_title, 
                             input_project_subtitle, input_project_client, project_dates_created)

        return ins
        
    @staticmethod
    def auto_generate_ins(process_continuity, input_file_path, input_file_basename, input_file_extension,
                          output_file_dir, input_project_code, input_project_title, input_project_subtitle, 
                          input_project_client, input_project_dates_created):
        
        if process_continuity == 'y':
            file_path = input_file_path 
            file_basename = input_file_basename
            file_extension = input_file_extension
            output_dir = output_file_dir
            project_code = input_project_code
            project_title = input_project_title
            project_subtitle = input_project_subtitle
            project_client = input_project_client
            project_dates_created = input_project_dates_created

            ins = DataIngestion(file_path, file_basename, file_extension, output_dir,
                                project_code, project_title, project_subtitle, project_client, 
                                project_dates_created)
            
            return ins
    
    @staticmethod
    def return_valid_file(input_file_dir:str) -> dict|int:        
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = DataIngestion._handle_dir(input_file_dir)
        elif os.path.isfile(input_file_dir):
            file_dict = DataIngestion._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str="r") -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = DataIngestion._display_directory_files(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = os.path.dirname(input_file_dir), 
            basename = os.path.basename(input_file_basename).split(".")[0],
            extension = os.path.basename(input_file_basename).split(".")[-1],
        )
    
    @staticmethod
    def _display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def _handle_file(input_file_dir:str):
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in ["xlsx", "xml"]):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return -1

    @staticmethod
    def return_valid_path(prompt_message:str) -> (str|None):
        while(True):   
            value = input(prompt_message).strip()
            try:
                if os.path.isdir(value):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, format)
            finish_time = datetime.strptime(finish_date, format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    def handle_xlsx(self, ws_name:str):
        _, ws = self._return_excel_workspace(ws_name)
        file_headers = self._xlsx_return_header(ws)

        json_header = self._json_fill_header(file_headers)
        json_obj = self._json_fill_body(ws, json_header)
        reworked_json = self._fill_missing_dates(json_obj)
        earliest_start, latest_finish = self._project_dates(reworked_json)
        reworked_json["start"] = earliest_start
        reworked_json["finish"] = latest_finish

        # WIP
        #json_obj_with_locations = self.identify_location(reworked_json)

        nested_json = self._build_nested_dic(reworked_json)

        return reworked_json, nested_json
    
    def _return_excel_workspace(self, worksheet_name:str):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        print(file)
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{self.ws_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = DataIngestion._display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def _xlsx_return_header(self, active_worksheet) -> list:
        ws = active_worksheet

        start_col = column_index_from_string(self.xlsx_start_col)
        header_list = []

        for row in ws.iter_rows(min_row=self.xlsx_start_row, max_row=self.xlsx_start_row, 
                                min_col=start_col, max_col= ws.max_column):
            for cell in row:
                if cell.value is not None:
                    normalized_str = self.normalize_string(cell.value)
                    header_tuple = (cell.coordinate, self._verify_header(normalized_str))
                    header_list.append(header_tuple)

        return header_list
    
    def _verify_header(self, entry_str:str) -> str:
        flat_allowed_headers = {
            value: key for key, values in self.project_content_headers.items() if values is not None for value in values
        }

        if entry_str in flat_allowed_headers:
            result = flat_allowed_headers[entry_str]
        else:
            print(
                f"Warning. Header '{entry_str}' not found in declared dictionary"
            )
            result = entry_str
        
        return result

    def _json_fill_header(self, file_headers:list) -> dict:
        json_obj = {
            "start": None,
            "finish": None,
            "header": {key: None for key, _ in self.project_content_headers.items()},
            "body": [],
        }

        for coordinates, data in file_headers:
            header_dict = json_obj["header"]

            header_dict[data] = coordinates

        return json_obj

    def _json_fill_body(self, active_worksheet, json_obj:dict) -> dict:
        ws = active_worksheet
        body_dict = json_obj["body"]
        header_dict = json_obj["header"]

        normalized_header = sorted(
            {k: v for k, v in header_dict.items() if v is not None}.items(), 
            key=lambda item: item[1]
        )

        header_coordinates_list = [dict_tuple[1] for dict_tuple in normalized_header]
        header_key_list = [dict_tuple[0] for dict_tuple in normalized_header]

        first_header_col_letter, first_header_row = self._get_column_coordinates(
            header_dict, normalized_header[0][0]
        )
        last_header_col_letter, _ = self._get_column_coordinates(header_dict, normalized_header[-1][0])
        first_header_col = column_index_from_string(first_header_col_letter)
        last_header_col = column_index_from_string(last_header_col_letter)

        entry_counter = 1
        for row in ws.iter_rows(min_row=first_header_row + 1, max_row=ws.max_row, 
                                min_col=first_header_col, max_col=last_header_col):
            json_activity = {key: None for key in header_dict.keys()}
            json_activity["entry"] = entry_counter
            for cell in row:
                parent_header = next((val for val in header_coordinates_list if get_column_letter(cell.column) in val), None)
                if cell.value:        
                    position = header_coordinates_list.index(parent_header)
                    key = header_key_list[position]
                    json_activity[key] = self._normalize_data_value(key, cell.value)
                else:
                    position = header_coordinates_list.index(parent_header)
                    key = header_key_list[position]

                    if key == "area":
                        json_activity[key] = "N/A"
                    else:
                        json_activity[key] = ""

            body_dict.append(json_activity)
            entry_counter += 1

        return json_obj
    
    def _get_column_coordinates(self, header_dict:dict, header_key:str):
        coordinates = header_dict[header_key]

        row_match = re.search(r'\d+', coordinates)
        column_match = re.search(r'[a-zA-Z]+', coordinates)

        if not row_match or not column_match:
            raise ValueError(f"Invalid coordinates format: {coordinates}")

        row = int(row_match.group(0))
        column = column_match.group(0)

        return column, row 

    def _normalize_data_value(self, header_column, cell_value):
        if isinstance(cell_value, str):
            if header_column == 'start' or header_column == 'finish':
                result = self._format_date_string(cell_value)
            else:
                result = cell_value.strip()

            return result
        elif isinstance(cell_value, date):
            return cell_value.strftime('%d-%b-%Y')
        elif isinstance(cell_value, int) or isinstance(cell_value, float):
            if cell_value == 0:
                cell_value = "0"
            return str(cell_value)
        else:
            return None
        
    def _format_date_string(self, date_string:str, 
                           output_format="%d-%b-%Y") -> str|None:
        try:
            special_chars = re.compile('[@_!#$%^&*()<>?\|}{~:]')

            if re.search(special_chars, date_string):
                cleaned_date_string = re.sub(special_chars, '', date_string)
            else:
                cleaned_date_string = re.sub("[a-zA-Z]$", '', date_string)
            

            parsed_date = parser.parse(cleaned_date_string.strip())
            formatted_date = parsed_date.strftime(output_format)

            return formatted_date
        except ValueError as e:
            print(f"Error parsing date: {e}")
            return None

    def _fill_missing_dates(self, json_obj:dict):
        body_dict = json_obj["body"]

        for item in body_dict:
            start = item.get("start")
            finish = item.get("finish")

            if start == "" and finish == "":
                print(f"Both 'start' and 'finish' are missing for entry: {item['entry']}")
            elif start == "":
                item["start"] = finish
            elif finish == "":
                item["finish"] = start
        
        json_obj["body"] = body_dict

        return json_obj
    
    def _project_dates(self, json_obj:dict):
        body_dict = json_obj["body"]
        
        start_list = [date["start"] for date in body_dict]
        finish_list = [date["finish"] for date in body_dict]

        earliest_start = self._bubble_sort_dates(start_list)[0]        
        latest_finish = self._bubble_sort_dates(finish_list)[-1]        

        return earliest_start, latest_finish

    def _bubble_sort_dates(self, unsorted_list:list) -> list:
        n = len(unsorted_list)

        for i in range(n):
            for j in range(n-i-1):
                value_j = unsorted_list[j]
                value_j1 = unsorted_list[j+1]

                try:
                    if isinstance(value_j, str) and isinstance(value_j1, str):
                        value_j = datetime.strptime(value_j, "%d-%b-%Y").date()
                        value_j1 = datetime.strptime(value_j1, "%d-%b-%Y").date()

                    elif isinstance(value_j1, datetime):
                        value_j = value_j.date()
                        value_j1 = value_j1.date()

                except Exception as e:
                    print(f"Error: {e}")
                else:
                    if value_j > value_j1:
                        unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]
        
        return unsorted_list

    def _build_nested_dic(self, json_obj:dict) -> dict:
        header_dict = json_obj["header"]
        body_dict = json_obj["body"]

        nested_dict = {
            "header": header_dict,
            "body": {},
        }

        def dict_management(obj, key_list, recursive_dict):
            if not key_list:
                return recursive_dict

            current_key = key_list.pop(0)
            obj_key = obj.get(current_key)

            if obj_key is not None and obj_key != "":
                if len(key_list) < 1:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = []
                    recursive_dict[obj_key].append(obj)
                else:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = {}
                    dict_management(obj, key_list, recursive_dict[obj_key])
            else:
                if "M_DATA" not in recursive_dict:
                    recursive_dict["M_DATA"] = []
                recursive_dict["M_DATA"].append(obj)

            return recursive_dict


        for obj in body_dict:
            keys = [category for category in self.json_struct_categories if obj.get(category) is not None]
            nested_dict["body"] = dict_management(
                obj, keys, nested_dict["body"]
            )
        
        return nested_dict

    def write_json(self, json_obj:dict, processed_json:bool=False) -> None:
        if processed_json:
            basename = f"processed_{self.output_json_basename}.json"
            file = os.path.join(self.output_path, basename)
        else:
            basename = f"{self.output_json_basename}.json"
            file = os.path.join(self.output_path, basename)

        with open(file, 'w') as file_writer:
            json.dump(json_obj, file_writer)

        print(f"JSON data successfully created and saved to {basename}.")

    def handle_xml(self):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        tree = ET.parse(file)
        root = tree.getroot()
        ns = {'ms': 'http://schemas.microsoft.com/project'}

        xml_tasks = []
        for task in root.findall("./ms:Tasks/ms:Task", ns):
            xml_tasks.append(task)

        json_dict_list = []
        if xml_tasks:
            for task in xml_tasks:
                task_data = {child.tag.split('}')[-1]: child.text.strip() for child in task}
                json_dict_list.append(task_data)
        else:
            print("No <Task> elements found")

        return json_dict_list   
        
    def create_wbs_table_to_fill(self, data:list, excel_path:str, 
                                 excel_basename:str):
        proc_table = self._generate_wbs_to_fill(data)
        self._write_data_to_excel(proc_table, excel_path, excel_basename)

        return proc_table

    def _generate_wbs_to_fill(self, json_obj:dict):
        df = pd.DataFrame(json_obj) 
        df.fillna("", inplace=True)

        return df

    def _write_data_to_excel(self, proc_table, excel_path:str, 
                            excel_basename:str) -> None:
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            excel_basename = self.input_basename.split('.')[0] + '.xlsx'
            file = os.path.join(excel_path, excel_basename)
            ws_name = "CFA - Fill Table"

            if not os.path.exists(excel_path):
                os.mkdir(excel_path)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='w') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name = ws_name, 
                        startrow = self.xlsx_start_row - 1, 
                        startcol = column_index_from_string(self.xlsx_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {ws_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")
    
    def identify_phase(self, json_obj):
        pass

    def identify_location(self, json_obj):
        body_dict = json_obj["body"]
        activity_names = [activity["activity_name"] for activity in body_dict]

        location_list = ["colo", "cell", "level", "zone", "area"]
        conventional_task_chars = re.compile("(?<=\W)[-_<>/|}{~:](?=\W)")
        prepositions_locations = re.compile("( for )|( in )|( at )|( on )|( inside )")

        for idx, name in enumerate(activity_names):
            normalized_name = name.lower().strip()
            has_special_chars_match = re.search(conventional_task_chars, normalized_name)
            has_prep_match = re.search(prepositions_locations, normalized_name)

            if has_special_chars_match:
                has_special_chars = has_special_chars_match.group(0)
                location_found = False
                components_found = normalized_name.split(has_special_chars)
                for component in components_found:
                    for location in location_list:
                        if location in component:
                            #print(component)
                            zone = component
                            self.check_for_subzone(zone, idx, body_dict)
                            location_found = True
                            break
                    
                    if location_found:
                        break
            elif has_prep_match:
                has_prep = has_prep_match.group(0)
                component = normalized_name.split(has_prep)[-1]
                for location in location_list:
                    if location in component:
                        zone = component
                        #print(component)
                        self.check_for_subzone(zone, idx, body_dict)
                        break
            else:
                body_dict[idx]["location"] = ""
    
        return json_obj

    def check_for_subzone(self, zone_name, element_idx, json_obj_body):
        location_keywords = ["colo", "cell", "level", "zone", "area"]
        conventional_subzone_chars = re.compile("[-<>/|}{~:\s]")
        normalized_name = zone_name.lower().strip()

        matched_locations = [location for location in location_keywords if location in normalized_name]

        if len(matched_locations) > 1:
            for location_type in matched_locations:
                match_style_one = re.search(f'{location_type}\s\d+', normalized_name)
                match_style_two = re.search(f'{location_type}\d+', normalized_name)

                if match_style_one:
                    formatted_match = re.sub(conventional_subzone_chars, '_', match_style_one.group(0)).upper()
                    if not json_obj_body[element_idx].get("location"):
                        json_obj_body[element_idx]["location"] = formatted_match
                    else:
                        json_obj_body[element_idx]["sub_location"] = formatted_match
                elif match_style_two:
                    formatted_match = re.sub(conventional_subzone_chars, '_', match_style_two.group(0)).upper()
                    if not json_obj_body[element_idx].get("location"):
                        json_obj_body[element_idx]["location"] = formatted_match
                    else:
                        json_obj_body[element_idx]["sub_location"] = formatted_match

        else:
            formatted_match = re.sub(conventional_subzone_chars, '', normalized_name).upper()
            json_obj_body[element_idx]["location"] = formatted_match
            json_obj_body[element_idx]["sub_location"] = ""


if __name__ == "__main__":
    project, project_dict = DataIngestion.main(False)

    if project:
        project.write_json(project_dict, True)
