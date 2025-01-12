import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
#from utils.data_frame_setup import DataFrameSetup

# Imported Helper - As Package 
from modules.utils.data_frame_setup import DataFrameSetup

class ExcelPostProcessing():
    def __init__(self, input_excel_path, input_excel_basename, input_worksheet_name, input_start_date,
                 input_end_date, input_json_path, input_json_basename, input_json_title):
        self.excel_path = input_excel_path
        self.excel_basename = input_excel_basename
        self.ws_name = input_worksheet_name
        self.start_date = input_start_date
        self.end_date = input_end_date
        self.json_path = input_json_path
        self.json_basename = input_json_basename
        self.json_title = input_json_title
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'

        #Structures
        self.wbs_final_categories = {
            "phase": "thick",
            "location": "dashed", 
            "area": "dashed",
        }

    @staticmethod
    def main(auto=True, input_excel_file=None, input_worksheet_name=None, input_start_date=None, 
             input_end_date=None, input_json_file=None, input_json_title=None):
        if auto:
            project = ExcelPostProcessing.auto_generate_ins(
                input_excel_file, input_worksheet_name, input_start_date, input_end_date, 
                input_json_file, input_json_title
            )
            project_details = DataFrameSetup.main(True, None, input_json_file, input_json_title)
        else:
            project = ExcelPostProcessing.generate_ins()
            project_details = DataFrameSetup.main(False)

        active_workbook, active_worksheet = project.return_excel_workspace(project.ws_name)
        proc_table = project_details.get("proc_table")
        custom_ordered_dict = project_details.get("custom_ordered_dict")
        lead_schedule_struct = project_details.get("lead_schedule_struct")
        json_struct_categories = project_details.get("json_struct_categories")

        if active_workbook and active_worksheet:
            project.update_schedule_frame(
                active_workbook, 
                active_worksheet, 
                custom_ordered_dict, 
                proc_table, 
                lead_schedule_struct,
            )
            project.update_wbs_table(lead_schedule_struct, json_struct_categories)

    @staticmethod
    def generate_ins():
        input_file_path = input("Please enter the path to the Excel file or directory: ")
        input_json_file = input("Please enter the path to the Json file or directory: ")
        input_start_date = input("Please enter the start date of the project (format: dd-MMM-yyyy): ")
        input_end_date = input("Please enter the end date of the project (format: dd-MMM-yyyy): ")

        input_excel_path, input_excel_basename = ExcelPostProcessing.file_verification(
            input_file_path, 'e', 'u')
        input_json_path, input_json_basename = ExcelPostProcessing.file_verification(
            input_json_file, 'j', 'r')
        
        if input_excel_path == -1:
            return None  

        input_worksheet_name = input("Please enter the name to create a worksheet: ")

        return ExcelPostProcessing(input_excel_path, input_excel_basename, input_worksheet_name, input_start_date,
                                   input_end_date, input_json_path, input_json_basename, None)

    @staticmethod
    def auto_generate_ins(input_excel_file, input_worksheet_name, input_start_date, 
                          input_end_date, input_json_file, input_json_title):
        input_excel_path, input_excel_basename = ExcelPostProcessing.file_verification(
            input_excel_file, 'e', 'u')
        
        input_json_path, input_json_basename = ExcelPostProcessing.file_verification(
            input_json_file, 'j', 'r')

        if input_excel_path == -1:
            return None  

        return ExcelPostProcessing(input_excel_path, input_excel_basename, input_worksheet_name, input_start_date, 
                                   input_end_date, input_json_path, input_json_basename, input_json_title)

    @staticmethod
    def file_verification(input_file_path, file_type, mode, input_json_title=None):
        if input_json_title and os.path.isdir(input_file_path):
            file_basename = f"processed_{ExcelPostProcessing.normalize_entry(input_json_title)}.json"
            path, basename = ExcelPostProcessing.handle_file(input_file_path, file_basename, file_type)
        else:
            if os.path.isdir(input_file_path):
                file_path, file_basename = ExcelPostProcessing.handle_dir(input_file_path, mode)
                if mode != 'c':
                    path, basename = ExcelPostProcessing.handle_file(file_path, file_basename, file_type)
                else:
                    path = file_path
                    basename = file_basename
            elif os.path.isfile(input_file_path):
                file_path = os.path.dirname(input_file_path)
                file_basename = os.path.basename(input_file_path)
                path, basename = ExcelPostProcessing.handle_file(file_path, file_basename, file_type)

        return path, basename
    
    @staticmethod
    def handle_dir(input_path, mode):
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_path)
            selection = ExcelPostProcessing.display_directory_files(dir_list)
            base_name = dir_list[selection]
            print(f'File selected: {base_name}\n')
        elif mode == 'c':
            base_name = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return input_path, base_name

    @staticmethod
    def handle_file(file_path, file_basename, file_type):
        file = os.path.join(file_path, file_basename)

        if (file_type == 'e' and ExcelPostProcessing.is_xlsx(file)) or \
           (file_type == 'j' and ExcelPostProcessing.is_json(file)):
            return os.path.dirname(file), os.path.basename(file)
        
        print("Error: Please verify that the directory and file exist and that the file is of type .xlsx or .json")
        return -1

    @staticmethod
    def display_directory_files(list):
        selection_idx = 0
        if len(list)==0:
            print("Error. No files found")
            return -1
        
        if len(list)>1:
            print(f"-- {len(list)} files found:")
            idx = 0
            for file in list:
                idx += 1
                print(f"{idx}. {file}")

            selection_idx = input("\nPlease enter the index number to select the one to process: ") 
        else:
            print(f"Single file found: {list[0]}")
            print("Will go ahead and process")

        return int(selection_idx) - 1

    @staticmethod
    def is_json(file_name):
        if file_name.endswith(".json"):
            return True
        else:
            print("Error: Selected file is not a JSON file")
            return False

    @staticmethod
    def is_xlsx(file_name):
        if file_name.endswith('.xlsx'):
            return True
        else:
            print('Error. Selected file is not an Excel')
            return False

    @staticmethod
    def normalize_entry(entry_str):
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower())
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    def return_excel_workspace(self, worksheet_name):
        file = os.path.join(self.excel_path, self.excel_basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{self.ws_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = ExcelPostProcessing.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def update_schedule_frame(self, active_workbook, active_worksheet, custom_ordered_dict:dict, 
                              proc_table, lead_schedule_struct:str) -> None:
        file = os.path.join(self.excel_path, self.excel_basename)
        overlap_results = self._overlapping_dates(custom_ordered_dict, lead_schedule_struct)
        entry_dict_available = {}
        
        ref_point = list(proc_table.index.get_level_values(lead_schedule_struct))[0]
        count_ins = 0

        for location in proc_table.index.get_level_values(lead_schedule_struct):
            if location == ref_point:
                count_ins += 1
            else:
                entry_dict_available.setdefault(ref_point, []).append(count_ins)
                ref_point = location
                count_ins = 1

        entry_dict_available.setdefault(ref_point, []).append(count_ins)
        self._process_file(active_worksheet, lead_schedule_struct, overlap_results, entry_dict_available)

        active_workbook.save(filename=file)
        print("Schedule Frame successfully updated")
        active_workbook.close()

    def _overlapping_dates(self, custom_ordered_dict:dict, 
                          lead_schedule_struct:str, alloted_space:int=2) -> dict:
        lead_based_lists = []
        nested_list = []
        ref_lead = self._generate_compound_category_name(custom_ordered_dict[0])

        for item in custom_ordered_dict:
            comp_lead_cat_title = self._generate_compound_category_name(item)

            if comp_lead_cat_title == ref_lead:
                nested_list.append(item)
            else:
                lead_based_lists.append(nested_list)
                nested_list = [item]
                ref_lead = comp_lead_cat_title

        if nested_list:
            lead_based_lists.append(nested_list)

        sorted_entries_list = []
        for location_list in lead_based_lists:
            sorted_list = self._bubble_sort_entries_by_dates(location_list)
            sorted_entries_list.append(sorted_list)

        overlap_results = self._calculate_overlap(sorted_entries_list, lead_schedule_struct, alloted_space)

        return overlap_results

    def _generate_compound_category_name(self, item: dict) -> str:
        category_names = []

        for category in self.wbs_final_categories.keys():
            cat_name = item.get(category)
            if cat_name:
                category_names.append(cat_name)

        return "|".join(category_names)

    def _bubble_sort_entries_by_dates(self, unsorted_list:list) -> list:
        n = len(unsorted_list)

        for i in range(n):
            for j in range(0, n-i-1):
                date_typed_start = datetime.strptime(unsorted_list[j]["start"], "%d-%b-%Y")
                date_typed_start_n1 = datetime.strptime(unsorted_list[j+1]["start"], "%d-%b-%Y")

                if date_typed_start > date_typed_start_n1:
                    unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]

                elif date_typed_start == date_typed_start_n1:
                    date_typed_end = datetime.strptime(unsorted_list[j]["finish"], "%d-%b-%Y")
                    date_typed_end_n1 = datetime.strptime(unsorted_list[j+1]["finish"], "%d-%b-%Y")

                    if date_typed_end > date_typed_end_n1:
                        unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]
        
        sorted_list = unsorted_list
        return sorted_list

    def _calculate_overlap(self, location_based_lists:list, 
                          lead_schedule_struct:str, alloted_space:int) -> dict:
        overlap_results = []

        for location_list in location_based_lists:
            active_overlaps = []
            max_overlap = 0
            ref_point = location_list[0][lead_schedule_struct]

            for activity in location_list:
                start = datetime.strptime(activity["start"], "%d-%b-%Y")
                finish = datetime.strptime(activity["finish"], "%d-%b-%Y")

                # Remove activities that have ended
                active_overlaps = [
                    (active_start, active_finish) 
                    for active_start, active_finish in active_overlaps 
                    if finish >= active_start and start <= active_finish
                ]

                # Add the current activity to active overlaps
                active_overlaps.append((start, finish))

                # Update the maximum overlap count
                max_overlap = max(max_overlap, len(active_overlaps))

            overlap_results.append((ref_point, max_overlap))

        max_rows_dict = {}
        for location, max_row in overlap_results:
            if location in max_rows_dict:
                max_rows_dict[location].append(max_row + alloted_space)
            else:
                max_rows_dict[location] = [max_row + alloted_space]

        return max_rows_dict
    
    def _process_file(self, active_worksheet, lead_schedule_struct:str, 
                     overlap_results:dict, con_available_ins:dict) -> None:
        ws = active_worksheet

        print(f"Processing Excel file: {self.excel_basename}")
        
        wbs_start_col = column_index_from_string(self.wbs_start_col)
        start_col_idx = self._find_column_idx(ws, lead_schedule_struct, self.wbs_start_row) + 1
        end_col_idx = self._return_first_column_idx(ws, 1, 1)

        self._unmerge_columns(ws, wbs_start_col, end_col_idx)
        self._delete_excess_rows(ws, lead_schedule_struct, overlap_results, con_available_ins)
        self._delete_columns(ws, start_col_idx, end_col_idx)
        self._insert_columns(ws, start_col_idx, end_col_idx - start_col_idx)
    
    def _find_column_idx(self, active_ws, column_header, start_row):
        ws = active_ws
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.replace(" ", "_").lower()

        for row in ws.iter_rows(min_row=start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def _return_first_column_idx(self, active_worksheet, start_row_idx:int, start_col_idx:int):
        ws = active_worksheet

        for row in ws.iter_rows(min_row=start_row_idx, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    return cell.column
        
        print(f"Error. No valid value encountered in specified row [{start_row_idx}, {start_row_idx}]")
        return -1

    def _unmerge_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        for merged_cell in list(ws.merged_cells.ranges):
            if (merged_cell.bounds[0] >= start_column_idx and 
                merged_cell.bounds[2] <= end_column_idx):
                ws.unmerge_cells(str(merged_cell))
        
        start_col_letter = get_column_letter(start_column_idx)
        end_col_letter = get_column_letter(end_column_idx)

        print(f"Column unmerging applied successfully: [{start_col_letter}, {end_col_letter}]")

    def _delete_excess_rows(self, active_worksheet, column_header:str, 
                            overlap_results:dict, con_available_ins:dict) -> None:
        ws = active_worksheet
        header_col = self._find_column_idx(active_worksheet, column_header, self.wbs_start_row)
        all_current_locations = list(overlap_results.keys())

        rows_to_delete = []
        init_count = 0
        max_rows = 0
        for row in ws.iter_rows(min_row=self.wbs_start_row + 1,
                                        max_row=ws.max_row,
                                        min_col=header_col,
                                        max_col=header_col):
            for cell in row:
                if cell.value is not None:  
                    init_count = 0
                    if cell.value in all_current_locations:
                        if len(overlap_results[cell.value]) > 1:
                            if overlap_results[cell.value][0] > con_available_ins[cell.value][0]:
                                max_rows = con_available_ins[cell.value][0]
                                con_available_ins[cell.value].pop(0)
                            else:
                                max_rows = overlap_results[cell.value][0]
                                overlap_results[cell.value].pop(0)
                        else:
                            if overlap_results[cell.value][0] > con_available_ins[cell.value][0]:
                                max_rows = con_available_ins[cell.value][0]
                            else:
                                max_rows = overlap_results[cell.value][0]

                if init_count > max_rows:
                    rows_to_delete.append(cell.row)
                
            init_count += 1

        for row in sorted(set(rows_to_delete), reverse=True):
            ws.delete_rows(row)
        
        print("Excess rows removed successfully.")

    def _delete_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        amount_to_delete = end_column_idx - start_column_idx
        ws.delete_cols(start_column_idx, amount_to_delete)

        start_col_letter = get_column_letter(start_column_idx)
        end_col_letter = get_column_letter(end_column_idx)

        print(f"Columns deleted successfully: [{start_col_letter}, {end_col_letter}]")

    def _insert_columns(self, active_worksheet, start_column_idx:int, num_cols:int) -> None:
        ws = active_worksheet

        for _ in range(num_cols):
            ws.insert_cols(start_column_idx)

        print(f"Columns ({num_cols}) added successfully")

    def update_wbs_table(self, lead_schedule_struct:str, json_struct_categories:list) -> None:
        file = os.path.join(self.excel_path, self.excel_basename)

        wb = load_workbook(file)
        ws = wb[self.ws_name]

        column_idxs = {
            item:self._find_column_idx(ws, item, self.wbs_start_row) for item in json_struct_categories 
            if item in self.wbs_final_categories.keys()
        }
        point_col_idx = self._find_column_idx(ws, lead_schedule_struct, self.wbs_start_row)

        start_col_idx = point_col_idx + 1
        end_col_idx = self._return_first_column_idx(ws, 1, 1)

        self._delete_columns(ws, start_col_idx, end_col_idx)

        for _, value in column_idxs.items():
            self._merge_until_different_value(ws, value, self.wbs_start_row)

        self._style_file(ws, start_col_idx)

        wb.save(filename=file)
        print("WBS Table successfully updated")
        wb.close()

    def _merge_until_different_value(self, active_worksheet, starting_col_idx:int, 
                                     starting_row_idx:int) -> None:
        ws = active_worksheet

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        current_merge_range = []
        merge_ranges = []

        for row_idx in range(starting_row_idx, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=starting_col_idx)

            if cell.value is not None:
                if current_merge_range:
                    merge_ranges.append(current_merge_range)
                current_merge_range = [cell]
            else:
                current_merge_range.append(cell)

        if current_merge_range:
            merge_ranges.append(current_merge_range)

        for merge_range in merge_ranges:
            if len(merge_range) > 1:
                start_row = merge_range[0].row
                end_row = merge_range[-1].row

                ws.merge_cells(
                    start_row=start_row,
                    start_column=starting_col_idx,
                    end_row=end_row,
                    end_column=starting_col_idx
                )

                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=starting_col_idx)
                    cell.border = thin_border

    def _style_file(self, active_worksheet, start_col:int, start_row:int=1) -> None:
        ws = active_worksheet

        year_list, year_row = self._same_cell_values(ws, start_col, start_row)
        for year in year_list:
            self._merge_same_value_cells(ws, year, year_row)  

        month_list, month_row = self._same_cell_values(ws, start_col, start_row + 1)
        for month in month_list:
            self._merge_same_value_cells(ws, month, month_row)  

        self._style_worksheet(ws, self.start_date, self.end_date, start_col, start_row)
        self._apply_post_styling(ws)

        for key, value in reversed(self.wbs_final_categories.items()):
            self._apply_post_sectioning(ws, key, value, start_col)

    def _same_cell_values(self, active_worksheet, starting_col_idx:int, starting_row_idx:int):
        ws = active_worksheet

        iterable_row = ws[starting_row_idx]
        last_value = iterable_row[starting_col_idx].value  
        same_cell_list = []  
        overall_list = []  

        for cell in iterable_row[starting_col_idx - 1:]:  
            if cell.value is None:
                break  
            elif cell.value == last_value:  
                same_cell_list.append(cell.column)  
            else:
                if same_cell_list:  
                    overall_list.append(same_cell_list)
                same_cell_list = [cell.column]  
                last_value = cell.value  

        if same_cell_list:
            overall_list.append(same_cell_list)

        return overall_list, starting_row_idx  

    def _merge_same_value_cells(self, active_worksheet, column_indices:list, 
                                starting_row_idx:int) -> None:
        ws = active_worksheet
        
        if not column_indices:
            print("No columns to merge.")
            return

        first_col = column_indices[0]
        last_col = column_indices[-1]

        ws.merge_cells(start_row=starting_row_idx, start_column=first_col, 
                              end_row=starting_row_idx, end_column=last_col)

    def _style_worksheet(self, active_worksheet, start_date, end_date, start_col, start_row):
        ws = active_worksheet
    
        start_datetime_obj = datetime.strptime(start_date, '%d-%b-%Y')
        end_datetime_obj = datetime.strptime(end_date, '%d-%b-%Y')
        duration = (end_datetime_obj - start_datetime_obj).days + 1  
    
        for cell in ws.iter_rows(min_row=start_row, max_row=start_row, 
                                        min_col=start_col, max_col=start_col + duration - 1):
            for cell in cell:  
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="FFFFFF")  
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00800080', end_color='00800080', fill_type='solid')
    
        for cell in ws.iter_rows(min_row=start_row + 1, max_row=start_row + 1, 
                                        min_col=start_col, max_col=start_col + duration - 1):
            for cell in cell:  
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="00333333")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00CC99FF', end_color='00CC99FF', fill_type='solid')
    
        for cell in ws.iter_rows(min_row=start_row + 2, max_row=start_row + 2, 
                                        min_col=start_col, max_col=start_col + duration - 1):
            for cell in cell:  
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="00000000")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')
        
        for row in ws.iter_rows(min_row=start_row + 3, min_col=start_col):  
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
        for col in range(start_col, start_col + duration):
            last_month_cell = ws.cell(row=start_row + 1, column=col)  
            if last_month_cell.value is not None:  
                for row in range(start_row + 1, ws.max_row + 1):  
                    cell = ws.cell(row=row, column=col)

                    current_border = cell.border if cell.border else Border()
                    cell.border = Border(
                        top=current_border.top, 
                        left=Side(border_style="dashed"),  
                        right=current_border.right,  
                        bottom=current_border.bottom  
                    )
    
        print("Workbook styled and saved successfully.")

    def _apply_post_styling(self, active_worksheet):
        ws = active_worksheet
        first_row = ws[self.wbs_start_row]
        
        if first_row is None:
            print("Error: The first row is empty.")
            return
        
        for cell in first_row:  
            cell.font = Font(name='Century Gothic', size=12, bold=True, color="FFFFFF")  
            cell.fill = PatternFill(start_color="00800080", end_color="00800080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")  
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)  
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.rows:    
            if isinstance(row[0].value, int):
                parent_row = row
                for cell in parent_row:
                    cell.font = Font(name='Century Gothic', size=12, bold=True, color="00333333")  
                    cell.fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(bottom=thin)

        print(f"Post-styling successfully completed")

    def _apply_post_sectioning(self, active_worksheet, section_header:str, 
                               border_style:str, start_col:int) -> None:
        ws = active_worksheet
        header_idx = self._find_column_idx(ws, section_header, self.wbs_start_row)

        section_list = self._list_cell_values(ws, header_idx)
        last_valid_section = section_list[0]

        for row_idx, row in enumerate(ws.iter_rows(min_col=start_col, 
                                                   max_col=ws.max_column, 
                                                   min_row=self.wbs_start_row + 1, 
                                                   max_row=ws.max_row)):
            current_section = section_list[row_idx]

            if current_section and current_section != last_valid_section:
                self._style_row_border(row, border_style)
                last_valid_section = current_section

        print("Post sectioning applied successfully.")

    def _list_cell_values(self, active_worksheet, col_idx):
        ws = active_worksheet
        col_list = []

        if col_idx is not None:
            for row in ws.iter_rows(min_row=self.wbs_start_row + 1, min_col=col_idx, 
                                    max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    col_list.append(cell.value)
        
        return col_list 

    def _style_row_border(self, row, border_style:str) -> None:
        if border_style == "no_border":
            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=current_border.top, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border
        else:
            new_top_border = Side(border_style=border_style, color='000000')

            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=new_top_border, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border


if __name__ == "__main__":
    ExcelPostProcessing.main(False)
