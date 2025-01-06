import os
import re
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

class DataRelationship:
    def __init__(self, input_xlsx_path, input_xlsx_basename, input_worksheet_name, input_json_path, 
                 input_json_basename, input_file_id, input_file_code, input_file_title, 
                 input_file_subtitle, input_file_issue_date, input_file_created_at, input_file_updated_at):
        self.xlsx_path = input_xlsx_path
        self.xlsx_basename = input_xlsx_basename
        self.ws_name = input_worksheet_name
        self.json_path = input_json_path
        self.json_path = input_json_path
        self.json_basename = input_json_basename
        self.output_json_path = input_json_path
        self.output_json_basename = DataRelationship.normalize_string(input_file_title)
        self.file_id = input_file_id
        self.file_code = input_file_code
        self.file_title = input_file_title
        self.file_subtitle = input_file_subtitle
        self.file_issue_date = input_file_issue_date
        self.file_created_at = input_file_created_at
        self.file_updated_at = input_file_updated_at
        self.json_categories = ["phase", "location", "area", "trade", "activity_code"]
        self.allowed_headers = {
            "activity_code": ["code", "task_code"],
            "activity_status": ["status", "task_status"],
            "hyperlinks": ["activity_code", "activity_name", "wbs_code"],
            "relationship_type": ["SS", "SF", "FS", "FF"]
        }

        #Instance Results
        self.initial_project_type:str = None
        self.final_project_type:any = None
        self.final_project_dict:dict = None

    @staticmethod
    def main(auto=True, process_continuity=None, input_file=None, input_worksheet_name=None,
             input_json_file=None, input_json_title=None, project_ins=None):
        if auto:
            project = DataRelationship.auto_generate_ins(process_continuity, input_file, input_worksheet_name,
                                                      input_json_file, input_json_title)
        else:
            project = DataRelationship.generate_ins()

        if project:
            active_workbook, active_worksheet = project.return_excel_workspace(project.ws_name)

            if active_workbook and active_worksheet:
                temp_custom_ordered_dict = project.read_json_dict()
                project.create_hyperlinks(active_worksheet, temp_custom_ordered_dict)
            else:
               print("Error. Could not open Excel file as Workbook & Worksheet") 
               
    @staticmethod
    def generate_ins():
        input_xlsx_file = input("Please enter the path to read the Excel file or directory: ")
        input_xlsx_path, input_xlsx_basename = DataRelationship.file_verification(
            input_xlsx_file, 'excel', 'r')
        input_worksheet_name = None
        input_json_file = input("Please enter the path to read the Json file or directory: ")
        input_json_path, input_json_basename = DataRelationship.file_verification(
            input_json_file, 'json', 'r')
        input_file_id = len([file for file in os.listdir(input_json_path) if file.endswith('json')])
        input_file_code = input("Enter Document CODE: ").strip()
        input_file_title = input("Enter Document Title: ").strip()
        input_file_subtitle = input("Enter Document Subtitle: ").strip()
        input_file_issue_date = DataRelationship.return_valid_date()
        input_file_created_at = datetime.strftime(datetime.now(), "%d/%m/%Y %H:%M:%S")
        input_file_updated_at = input_file_created_at

        ins = DataRelationship(input_xlsx_path, input_xlsx_basename, input_worksheet_name, input_json_path, 
                               input_json_basename, input_file_id, input_file_code, input_file_title, 
                               input_file_subtitle, input_file_issue_date, input_file_created_at, input_file_updated_at)
        
        return ins

    @staticmethod
    def auto_generate_ins(process_continuity, input_file, input_worksheet_name, 
                          input_json_file, input_json_title):
        
        if process_continuity == 'y':
            input_file_path, input_file_basename = DataRelationship.file_verification(
                input_file, 'excel', 'r')
            input_json_path, _ = DataRelationship.file_verification(
                    input_json_file, 'json', 'c')
            input_file_id = len([file for file in os.listdir(input_json_path) if file.endswith('json')])
            input_file_code = f"{input_json_title.strip()}_{input_file_id}"
            input_file_title = input_json_title.strip()
            input_file_subtitle = input_json_title.strip()
            input_file_issue_date = DataRelationship.return_valid_date()
            input_file_created_at = datetime.strftime(datetime.now(), "%d/%m/%Y %H:%M:%S")
            input_file_updated_at = input_file_created_at

            ins = DataRelationship(input_file_path, input_file_basename, input_worksheet_name, input_json_path, 
                                input_file_id, input_file_code, input_file_title, input_file_subtitle, 
                                input_file_issue_date, input_file_created_at, input_file_updated_at)
            
            return ins

    @staticmethod
    def is_type_file(input_basename):
        file_types = {
            "xlsx": ['e', 'excel', 'xlsx'],
            "csv": ['c', 'csv'],
            "json": ['j', 'json'],
            "xml": ['x', 'xml']
        }

        for file_type, extensions in file_types.items():
            if any(input_basename.endswith(ext) for ext in extensions):
                return file_type

        print("Unknown file type.")
        return None
    
    @staticmethod
    def file_verification(input_file_path, file_type, mode):
        if os.path.isdir(input_file_path):
            file_path, file_basename = DataRelationship.handle_dir(input_file_path, mode)
            if mode != 'c':
                path, basename = DataRelationship.handle_file(file_path, file_basename, file_type)
            else:
                path = file_path
                basename = file_basename
        elif os.path.isfile(input_file_path):
            file_path = os.path.dirname(input_file_path)
            file_basename = os.path.basename(input_file_path)
            path, basename = DataRelationship.handle_file(file_path, file_basename, file_type)
        else:
            print("Warning. File type not workable")
            file_path = None
            file_basename = None

        return path, basename
    
    @staticmethod
    def handle_dir(input_path, mode):
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_path)
            selection = DataRelationship.display_directory_files(dir_list)
            base_name = dir_list[selection]
            print(f'File selected: {base_name}\n')
        elif mode == 'c':
            base_name = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return input_path, base_name

    @staticmethod
    def handle_file(file_path, file_basename, file_type):
        file = os.path.join(file_path, file_basename)

        valid_file_types = {
            "csv": "c",
            "excel": DataRelationship.is_xlsx(file),
            "json": DataRelationship.is_json(file),
            "pdf": "p",
        }

        if (valid_file_types[file_type]):
            return os.path.dirname(file), os.path.basename(file)
        
        print("Error: Please verify that the directory and file exist and that the file is of type .xlsx or .json")
        return -1

    @staticmethod
    def return_valid_path(message):
        while(True):   
            value = input(message).strip()
            try:
                if os.path.isdir(value):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")

    @staticmethod
    def return_valid_date():
        while(True):   
            value = input("Enter Issue Date (format: dd-MMM-yyyy): ").strip()
            try:
                if datetime.strptime(value, "%d-%b-%Y"):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")

    @staticmethod
    def display_directory_files(list):
        selection_idx = 0
        if len(list)==0:
            print("Error. No files found")
            return -1
        
        if len(list)>1:
            print(f"-- {len(list)} files found:")
            idx = 0
            for file in list:
                idx += 1
                print(f"{idx}. {file}")

            selection_idx = input("\nPlease enter the index number to select the one to process: ") 
        else:
            print(f"Single file found: {list[0]}")
            print("Will go ahead and process")

        return int(selection_idx) - 1

    @staticmethod
    def is_json(file_name):
        if file_name.endswith('.json'):
            return True
        else:
            print('Error: Selected file is not a JSON file')
            return False

    @staticmethod
    def is_xlsx(file_name):
        if file_name.endswith('.xlsx'):
            return True
        else:
            print('Error. Selected file is not an Excel file')
            return False

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    @staticmethod
    def document_project(project_ins, project_final:dict) -> None:
        project_ins.initial_project_type = project_final.get("init_file_type")
        project_ins.final_project_type = project_final.get("final_file_type")
        project_ins.final_project_dict = project_final.get("final_project_dict")

    def return_excel_workspace(self, worksheet_name):
        file = os.path.join(self.xlsx_path, self.xlsx_basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{self.ws_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = DataRelationship.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def read_json_dict(self):
        j_file = os.path.join(self.json_path, self.json_basename)

        with open(j_file, 'r') as json_file:
            data = json.load(json_file)

        return data

    def create_hyperlinks(self, active_workbook, active_worksheet, custom_ordered_dict:dict) -> None:
        file = os.path.join(self.xlsx_path, self.xlsx_basename)

        hyperlink_entry_list = {key:value for key, value in custom_ordered_dict.items() 
                                if value.get("predecessor") is not None and len(value.get("predecessor")) > 1}
        self.hyperlink_cells(active_worksheet, custom_ordered_dict, hyperlink_entry_list)

        active_workbook.save(filename=file)
        print("CFA Schedule successfully linked")
        active_workbook.close()

    def hyperlink_cells(self, active_worksheet, custom_ordered_dict:dict, hyperlink_entry_list:list):
        ws = active_worksheet
        entry_dict = hyperlink_entry_list
        all_rltnshp_list = []
        embedded_list = []

        for _, value in entry_dict.items():
            embedded_list.append(self._verify_relationships(value))
        
        ovr_count = 1
        for relationship_list in embedded_list:
            for item in relationship_list:
                item["ovr_relationship_id"] = ovr_count
                ovr_count += 1

        for relationship_list in embedded_list:
            all_rltnshp_list += relationship_list

        link_dict = {item["ovr_relationship_id"]:item for item in all_rltnshp_list}
        ref_link_dict = self._generate_relationships(custom_ordered_dict, link_dict)

        # == WIP ==
        #nearest_links_dict = self._determine_nearest_links(ref_link_dict)

        self._hyperlink_items(ws, ref_link_dict)

    def _verify_relationships(self, entry_dict:dict, relationship_count:int=1) -> list:
        rltshps = []
        current_pred_list = entry_dict.get("predecessor").split(',')
        current_succ_list = entry_dict.get("successor").split(',')
        current_type_list = entry_dict.get("relationship_type").split(',')
        
        if len(current_pred_list) == len(current_succ_list):
            for idx, pred in enumerate(current_pred_list):
                rltshps.append(
                    self._generate_relationship_dict(
                        entry_dict, 
                        relationship_count, 
                        pred, 
                        current_succ_list[idx], 
                        current_type_list[idx]
                    )
                )
                relationship_count += 1

        elif len(current_pred_list) > len(current_succ_list):
            print(f"Warning: Relationship defined in entry '{entry_dict['entry']}' is missing values.")
            available_links = len(current_succ_list)
            new_pred_list = current_pred_list[:available_links]

            for idx, pred in enumerate(new_pred_list):
                rltshps.append(
                    self._generate_relationship_dict(
                        entry_dict, 
                        relationship_count, 
                        pred, 
                        current_succ_list[idx], 
                        current_type_list[idx]
                    )
                )
                relationship_count += 1

        else:
            print(f"Warning: Relationship defined in entry '{entry_dict['entry']}' is missing values.")
            available_links = len(current_pred_list)
            new_succ_list = current_succ_list[:available_links]

            for idx, pred in enumerate(new_succ_list):
                rltshps.append(
                    self._generate_relationship_dict(
                        entry_dict, 
                        relationship_count, 
                        pred, 
                        current_succ_list[idx], 
                        current_type_list[idx]
                    )
                )
                relationship_count += 1

        return rltshps

    def _generate_relationship_dict(self, entry_dict:dict, count:int, pred:str, succ:str, rel_type:str) -> dict:
        try:
            return {         
                "ref_entry": entry_dict["entry"],
                "relationship_id": count,
                "predecessor": pred.strip(),
                "successor": succ.strip(),
                "type": rel_type.strip()
            }
        except IndexError:
            print(f"Warning: Relationship type not defined in entry '{entry_dict['entry']}'. Falling back to default configuration.")
            return {         
                "ref_entry": entry_dict["entry"],
                "relationship_id": count,
                "predecessor": pred.strip(),
                "successor": succ.strip(),
                "type": "FS"
            }
        
    def _generate_relationships(self, custom_ordered_dict:dict, hyperlink_dict:dict) -> dict:
        for _, value in hyperlink_dict.items():
            current_pred = value.get("predecessor")
            current_succ = value.get("successor")
            
            ref_entry_pred, ref_pred_sequence = self._get_reference_entry(
                custom_ordered_dict, current_pred
            )
            ref_entry_succ, ref_succ_sequence = self._get_reference_entry(
                custom_ordered_dict, current_succ
            )

            if ref_entry_pred not in value:
                value["ref_entry_predecessor"] = ref_entry_pred
                value["predecessor_cell_sequence"] = ref_pred_sequence

            if ref_entry_succ not in value:
                value["ref_entry_successor"] = ref_entry_succ
                value["successor_cell_sequence"] = ref_succ_sequence

        return hyperlink_dict
    
    def _get_reference_entry(self, custom_ordered_dict:dict, entry_str:str, category:str="hyperlinks"):
        header = None
        ref_entry = None
        ref_cell_sequence = []

        for category in self.allowed_headers[category]:
            category_list = [value.get(category) for _, value in custom_ordered_dict.items()]

            if entry_str in category_list:
                header = category
                break

        for _, value in custom_ordered_dict.items():
            if value.get(header) == entry_str:
                ref_entry = value.get("entry")
                ref_cell_sequence = value.get("cell_sequence")

        return ref_entry, ref_cell_sequence

    def _determine_nearest_links(self, hyperlinks_dict:dict) -> dict:
        nearest_links = {}

        for _, value in hyperlinks_dict.items():
            current_entry = value.get("ref_entry")
            entry_pred = value.get("ref_entry_pred")
            entry_succ = value.get("ref_entry_succ")

            if current_entry not in nearest_links:
                nearest_links[current_entry] = {"nearest_link": None, "delta": float("inf")}

            if entry_pred and entry_succ:
                for linked_entry in [entry_pred, entry_succ]:
                    if linked_entry != current_entry:
                        delta = abs(current_entry - linked_entry)
                        if delta < nearest_links[current_entry]["delta"]:
                            nearest_links[current_entry]["nearest_link"] = linked_entry
                            nearest_links[current_entry]["delta"] = delta

        #print("Nearest Links:", nearest_links)
        return nearest_links

    def _hyperlink_items(self, active_worksheet, link_dict:dict, category:str="relationship_type") -> None:
        ws = active_worksheet

        for _, value in link_dict.items():
            curr_relationship_type = value.get("type")
            
            if curr_relationship_type in self.allowed_headers[category]:
                ref_entry_predecessor = value.get("ref_entry_predecessor")
                ref_entry_successor = value.get("ref_entry_successor")

                if curr_relationship_type == "SS":
                    if ref_entry_predecessor and ref_entry_successor:
                        start_cell_pred = value.get("predecessor_cell_sequence")[0]
                        start_cell_succ = value.get("successor_cell_sequence")[0]
                        self._add_hyperlink(ws, start_cell_pred, start_cell_succ)
                elif curr_relationship_type == "SF":
                    if ref_entry_predecessor and ref_entry_successor:
                        start_cell_pred = value.get("predecessor_cell_sequence")[0]
                        start_cell_succ = value.get("successor_cell_sequence")[-1]
                        self._add_hyperlink(ws, start_cell_pred, start_cell_succ)
                elif curr_relationship_type == "FS":
                    if ref_entry_predecessor and ref_entry_successor:
                        start_cell_pred = value.get("predecessor_cell_sequence")[-1]
                        start_cell_succ = value.get("successor_cell_sequence")[0]
                        self._add_hyperlink(ws, start_cell_pred, start_cell_succ)
                else:
                    if ref_entry_predecessor and ref_entry_successor:
                        start_cell_pred = value.get("predecessor_cell_sequence")[-1]
                        start_cell_succ = value.get("successor_cell_sequence")[-1]
                        self._add_hyperlink(ws, start_cell_pred, start_cell_succ)
        
    def _add_hyperlink(self, active_worksheet, predecessor_cell: str, successor_cell: str) -> None:
        ws = active_worksheet

        pred_cell = ws[predecessor_cell]
        succ_cell = ws[successor_cell]
        pred_comment = pred_cell.comment
        succ_comment = succ_cell.comment

        wb_name = self.xlsx_basename
        ws_title = ws.title

        succ_hprl_title = f"Link to {successor_cell}"
        succ_hyperlink_msg = f'=HYPERLINK("[{wb_name}]\'{ws_title}\'!{successor_cell}", "{succ_hprl_title}")'
        
        pred_hprl_title = f"Link to {predecessor_cell}"
        pred_hyperlink_msg = f'=HYPERLINK("[{wb_name}]\'{ws_title}\'!{predecessor_cell}", "{pred_hprl_title}")'

        if pred_comment:
            if "=== Successors ===" not in pred_comment.text:
                pred_comment.text += f"\n=== Successors ===\n"
            pred_comment.text += f"{succ_hyperlink_msg}\n"
        else:
            pred_cell.comment = Comment(f"=== Successors ===\n{succ_hyperlink_msg}\n", "AMP")

        if succ_comment:
            if "=== Predecessors ===" not in succ_comment.text:
                succ_comment.text += f"\n=== Predecessors ===\n"
            succ_comment.text += f"{pred_hyperlink_msg}\n"
        else:
            succ_cell.comment = Comment(f"=== Predecessors ===\n{pred_hyperlink_msg}\n", "AMP")


if __name__ == "__main__":
    DataRelationship.main(False)
