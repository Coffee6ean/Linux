import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

# Imported Helper - As Module
#from utils.data_frame_setup import DataFrameSetup

# Imported Helper - As Package 
from modules.utils.data_frame_setup import DataFrameSetup

class WbsFramework:
    def __init__(self, input_process_cont, input_excel_path, input_excel_basename, 
                input_worksheet_name, input_json_path, input_json_basename):
        self.process_cont = input_process_cont
        self.excel_path = input_excel_path
        self.excel_basename = input_excel_basename
        self.ws_name = input_worksheet_name
        self.json_path = input_json_path
        self.json_basename = input_json_basename
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.default_hex_font_color = "00FFFFFF"
        self.default_hex_fill_color = "00FFFF00"
    
    @staticmethod
    def main(auto=True, process_continuity=None, input_excel_file=None, 
             input_worksheet_name=None, input_json_file=None, input_json_title=None):
        if auto:
            project = WbsFramework.auto_generate_ins(process_continuity, input_excel_file, 
                                                     input_worksheet_name, input_json_file,
                                                     input_json_title)
            project_details = DataFrameSetup.main(True, process_continuity, input_json_file, input_json_title)
        else:
            project = WbsFramework.generate_ins()
            project_details = DataFrameSetup.main(False)

        proc_table = project_details.get("proc_table")
        custom_ordered_dict = project_details.get("custom_ordered_dict")

        while True:
            color_list = [project.process_hex_val(item["color"]) for item in custom_ordered_dict]
            process_choice = input("Enter 'c' to create WBS Data Table, 'u' to update an existing WBS Table, or 'q' to quit: ").lower()

            if process_choice == 'c':
                project.create_wbs_table(proc_table)
                project.process_wbs_column('activity code', color_list)
                project.process_wbs_column('color', color_list)

                break
            elif process_choice == 'u':
                project.process_wbs_column('activity code', color_list)
                project.process_wbs_column('color', color_list)
                
                break
            elif process_choice == 'q':
                print("Exiting the program.")
                break
            else:
                print("Invalid input. Please enter 'c', 'u', or 'q'.")

    @staticmethod
    def generate_ins():
        input_process_cont = WbsFramework.ynq_user_interaction("Continue with the program? ")
        if input_process_cont == 'q':
            print("Exiting the program.")
            return -1 
        
        input_excel_file = input("Please enter the path to the Excel file or directory: ")
        input_worksheet_name = input("Please enter the name for the new or existing worksheet: ")
        input_json_file = input("Please enter the path to the Json file or directory: ")

        input_excel_path, input_excel_basename = WbsFramework.file_verification(
            input_excel_file, 'e', 'r')
        input_json_path, input_json_basename = WbsFramework.file_verification(
            input_json_file, 'j', 'r')

        ins = WbsFramework(input_process_cont, input_excel_path, input_excel_basename, 
                            input_worksheet_name, input_json_path, input_json_basename)
        
        return ins

    @staticmethod
    def auto_generate_ins(process_continuity, input_excel_file, input_worksheet_name, 
                          input_json_file, input_json_title):
        input_excel_path, input_excel_basename = WbsFramework.file_verification(
            input_excel_file, 'e', 'r')
        input_json_path, input_json_basename = WbsFramework.file_verification(
                input_json_file, 'j', 'r', input_json_title)

        ins = WbsFramework(process_continuity, input_excel_path, input_excel_basename, 
                            input_worksheet_name, input_json_path, input_json_basename)
        
        return ins

    @staticmethod
    def ynq_user_interaction(prompt_message):
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower()  
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def display_directory_files(list):
        selection_idx = 0
        if len(list)==0:
            print("Error. No files found")
            return -1
        
        if len(list)>1:
            print(f"-- {len(list)} files found:")
            idx = 0
            for file in list:
                idx += 1
                print(f"{idx}. {file}")

            selection_idx = input("\nPlease enter the index number to select the one to process: ") 
        else:
            print(f"Single file found: {list[0]}")
            print("Will go ahead and process")

        return int(selection_idx) - 1

    @staticmethod
    def is_json(file_name):
        if file_name.endswith(".json"):
            return True
        else:
            print("Error: Selected file is not a JSON file")
            return False

    @staticmethod
    def is_xlsx(file_name):
        if file_name.endswith(".xlsx"):
            return True
        else:
            print("Error. Selected file is not an Excel")
            return False

    @staticmethod
    def clear_directory(directory_path):
        if os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            for file in file_list:
                file_path = os.path.join(directory_path, file)
                os.remove(file_path)

    @staticmethod
    def file_verification(input_file_path, file_type, mode, input_json_title=None):
        if input_json_title and os.path.isdir(input_file_path):
            file_basename = f"processed_{WbsFramework.normalize_entry(input_json_title)}.json"
            path, basename = WbsFramework.handle_file(input_file_path, file_basename, file_type)
        else:
            if os.path.isdir(input_file_path):
                file_path, file_basename = WbsFramework.handle_dir(input_file_path, mode)
                if mode != 'c':
                    path, basename = WbsFramework.handle_file(file_path, file_basename, file_type)
                else:
                    path = file_path
                    basename = file_basename
            elif os.path.isfile(input_file_path):
                file_path = os.path.dirname(input_file_path)
                file_basename = os.path.basename(input_file_path)
                path, basename = WbsFramework.handle_file(file_path, file_basename, file_type)

        return path, basename
    
    @staticmethod
    def handle_dir(input_path, mode):
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_path)
            selection = WbsFramework.display_directory_files(dir_list)
            base_name = dir_list[selection]
            print(f'File selected: {base_name}\n')
        elif mode == 'c':
            base_name = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return input_path, base_name

    @staticmethod
    def handle_file(file_path, file_basename, file_type):
        file = os.path.join(file_path, file_basename)

        if (file_type == 'e' and WbsFramework.is_xlsx(file)) or \
           (file_type == 'j' and WbsFramework.is_json(file)):
            return os.path.dirname(file), os.path.basename(file)
        
        print("Error: Please verify that the directory and file exist and that the file is of type .xlsx or .json")
        return -1
    
    @staticmethod
    def normalize_entry(entry_str):
        remove_bewteen_parenthesis = re.sub('(?<=\()(.*?)(?=\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower())
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    def process_wbs_column(self, col_header, color_list):
        file = os.path.join(self.excel_path, self.excel_basename)
        wb = load_workbook(file)
        ws = wb[self.ws_name]

        header_idx = self.find_column_idx(ws, col_header)
        self.fill_color_col(ws, header_idx, color_list)

        wb.save(file)

    def return_excel_workspace(self, worksheet_name):
        file = os.path.join(self.excel_path, self.excel_basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{self.ws_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = WbsFramework.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def create_wbs_table(self, proc_table):
        self.write_data_to_excel(proc_table)

    def write_data_to_excel(self, proc_table):
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            file = os.path.join(self.excel_path, self.excel_basename)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name=self.ws_name, 
                        startrow=self.wbs_start_row - 1, 
                        startcol=column_index_from_string(self.wbs_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {self.ws_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")

    def find_column_idx(self, active_ws, column_header):
        ws = active_ws
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.replace(" ", "_").lower()

        for row in ws.iter_rows(min_row=self.wbs_start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def process_hex_val(self, hex_val):
        return hex_val.replace('#', "00")

    def fill_color_col(self, active_ws, col_idx, col_list):
        ws = active_ws

        if not col_list:
            print("Error: Color list is empty.")
            return

        for idx, row in enumerate(ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                                max_row=ws.max_row, 
                                                min_col=col_idx,
                                                max_col=col_idx)):
            for cell in row:
                color = col_list[idx]
                try:
                    cell.fill = PatternFill(start_color=color, 
                                            end_color=color, 
                                            fill_type="solid")
                except Exception as e:
                    print(f"Color hex not found: {color}. Error: {e}")
                    cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                            end_color=self.default_hex_fill_color, 
                                            fill_type="solid")

        column_letter = get_column_letter(col_idx)
        print(f"Column ({column_letter}) styled successfully")


if __name__ == "__main__":
    WbsFramework.main(False)
