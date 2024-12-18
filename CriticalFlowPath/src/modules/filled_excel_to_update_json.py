import os
import json
from datetime import date
from typing import Union
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string 

class FilledExcelToUpdateJson():
    def __init__(self, input_process_cont, input_excel_path, input_excel_basename, 
                input_worksheet_name, input_json_path, input_json_basename, 
                input_json_title):
        self.process_cont = input_process_cont
        self.excel_path = input_excel_path
        self.excel_basename = input_excel_basename
        self.ws_name = input_worksheet_name
        self.json_path = input_json_path
        self.json_basename = input_json_basename
        self.json_title = input_json_title
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.final_data_col = 13
        self.json_categories = ["phase", "location", "trade", "activity_code"]
    
    @staticmethod
    def main(auto=True, process_continuity=None, input_excel_file=None, 
             input_worksheet_name=None, input_json_file=None, input_json_title=None):
        if auto:
            project = FilledExcelToUpdateJson.auto_generate_ins(process_continuity, input_excel_file, 
                                                                input_worksheet_name, input_json_file, 
                                                                input_json_title)
        else:
            project = FilledExcelToUpdateJson.generate_ins()
        
        if project:
            try:
                wb, ws = project.return_excel_workspace(project.ws_name)
                entry_frame = project.build_entry_frame(ws)

                if project.process_cont == 'n':
                    # Update Existing JSON
                    for category in project.json_categories:
                        project.update_json(project.extract_col_cells(ws, category), category)
                else:
                    # Create New JSON
                    project.create_json()
                    project.build_project_dic(ws, entry_frame)

                # Create Restructured JSON
                file = os.path.join(project.json_path, project.json_basename)
                with open(file, 'r') as f:
                    json_obj = json.load(f)

                json_dic = {
                    "project_metadata": json_obj["project_metadata"],
                    "project_content": []
                }

                json_dic["project_content"].append(project.build_nested_dic(entry_frame))
                project.write_json(json_dic)

                print("JSON file(s) updated successfully!")
            except Exception as e:
                print(f"An error occurred: {e}")
                print("Please check the file path and try again.")
        else:
            print("Exiting the program.")
        
    @staticmethod
    def generate_ins():
        input_process_cont = FilledExcelToUpdateJson.ynq_user_interaction("Continue with the program? ")
        if input_process_cont == 'q':
            print("Exiting the program.")
            return -1 
        
        is_new = FilledExcelToUpdateJson.yn_user_interaction("Is this a completly new project? ")
        input_excel_file = input("Please enter the path to the Excel file or directory: ")
        input_worksheet_name = input("Please enter the name for the new or existing worksheet: ")
        input_excel_path, input_excel_basename = FilledExcelToUpdateJson.file_verification(
            input_excel_file, 'e', 'r')
        input_json_file = input("Please enter the directory to save the new JSON file: ")

        if is_new == 'y':
            input_json_path, input_json_basename = FilledExcelToUpdateJson.file_verification(
                input_json_file, 'j', 'c')
        elif is_new == 'n':
            input_json_path, input_json_basename = FilledExcelToUpdateJson.file_verification(
                input_json_file, 'j', 'u')

        ins = FilledExcelToUpdateJson(input_process_cont, input_excel_path, input_excel_basename, 
                                      input_worksheet_name, input_json_path, input_json_basename)
        
        return ins

    @staticmethod
    def auto_generate_ins(process_continuity, input_excel_file, input_worksheet_name,
                           input_json_file, input_json_title):
        input_excel_path, input_excel_basename = FilledExcelToUpdateJson.file_verification(
            input_excel_file, 'e', 'r')
        
        if process_continuity == 'y':
            input_json_path, input_json_basename = FilledExcelToUpdateJson.file_verification(
                input_json_file, 'j', 'c')
        elif process_continuity == 'n':
            input_json_path, input_json_basename = FilledExcelToUpdateJson.file_verification(
                input_json_file, 'j', 'u')

        ins = FilledExcelToUpdateJson(process_continuity, input_excel_path, input_excel_basename, 
                                      input_worksheet_name, input_json_path, input_json_basename,
                                      input_json_title)
        
        return ins

    @staticmethod
    def ynq_user_interaction(prompt_message):
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower()  
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def yn_user_interaction(prompt_message):
        valid_responses = {'y', 'n'}  
        
        while True:
            user_input = input(prompt_message).lower()  
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No]\n")

    @staticmethod
    def display_directory_files(list):
        selection_idx = 0
        if len(list)==0:
            print('Error. No files found')
            return -1
        
        if len(list)>1:
            print(f'-- {len(list)} files found:')
            idx = 0
            for file in list:
                idx += 1
                print(f'{idx}. {file}')

            selection_idx = input('\nPlease enter the index number to select the one to process: ') 
        else:
            print(f'Single file found: {list[0]}')
            print('Will go ahead and process')

        return int(selection_idx) - 1

    @staticmethod
    def is_json(file_name):
        if file_name.endswith('.json'):
            return True
        else:
            print('Error: Selected file is not a JSON file')
            return False

    @staticmethod
    def is_xlsx(file_name):
        if file_name.endswith('.xlsx'):
            return True
        else:
            print('Error. Selected file is not an Excel')
            return False

    @staticmethod
    def clear_directory(directory_path):
        if os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            for file in file_list:
                file_path = os.path.join(directory_path, file)
                os.remove(file_path)

    @staticmethod
    def file_verification(input_file_path, file_type, mode):
        if os.path.isdir(input_file_path):
            file_path, file_basename = FilledExcelToUpdateJson.handle_dir(input_file_path, mode)
            if mode != 'c':
                path, basename = FilledExcelToUpdateJson.handle_file(file_path, file_basename, file_type)
            else:
                path = file_path
                basename = file_basename
        elif os.path.isfile(input_file_path):
            file_path = os.path.dirname(input_file_path)
            file_basename = os.path.basename(input_file_path)
            path, basename = FilledExcelToUpdateJson.handle_file(file_path, file_basename, file_type)

        return path, basename
    
    @staticmethod
    def handle_dir(input_path, mode):
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_path)
            selection = FilledExcelToUpdateJson.display_directory_files(dir_list)
            base_name = dir_list[selection]
            print(f'File selected: {base_name}\n')
        elif mode == 'c':
            base_name = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return input_path, base_name

    @staticmethod
    def handle_file(file_path, file_basename, file_type):
        file = os.path.join(file_path, file_basename)

        if (file_type == 'e' and FilledExcelToUpdateJson.is_xlsx(file)) or \
           (file_type == 'j' and FilledExcelToUpdateJson.is_json(file)):
            return os.path.dirname(file), os.path.basename(file)
        
        print("Error: Please verify that the directory and file exist and that the file is of type .xlsx or .json")
        return -1

    def return_excel_workspace(self, worksheet_name):
        file = os.path.join(self.excel_path, self.excel_basename)
        workbook = load_workbook(filename=file)
        try:
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{worksheet_name}' created.")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = FilledExcelToUpdateJson.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def update_json(self, info_dic, dic_attr_type):
        file = os.path.join(self.json_path, self.json_basename)

        with open(file, 'r') as json_file:
            data = json.load(json_file)

        json_body = data["project_content"]["body"]
        phase_mapping = {entry["row"]: entry["value"] for entry in info_dic["body"]["filled"]}

        for activity in json_body:
            activity[dic_attr_type] = None
            entry_row = activity.get('entry')

            if entry_row in phase_mapping:
                activity[dic_attr_type] = phase_mapping[entry_row]

            for sub_activity in activity.get('activities', []):
                sub_entry_row = sub_activity.get('entry')
                if sub_entry_row in phase_mapping:
                    sub_activity[dic_attr_type] = phase_mapping[sub_entry_row]

        with open(file, 'w') as json_file:
            json.dump(data, json_file, indent=4)

    def build_project_dic(self, active_ws, entry_frame):
        file = os.path.join(self.json_path, self.json_basename)
        ws = active_ws
        initial_col_idx = self.find_column_idx(ws, self.json_categories[0])
        final_col_idx = self.find_column_idx(ws, 'finish')
        
        entry_counter = 1
        json_dic = {
            "project_metadata": {},
            "project_content": {
                "header": list(entry_frame.keys()),
                "body": []
            }
        }
        
        header_list = list(entry_frame.keys())[1:]

        for row in ws.iter_rows(min_col=initial_col_idx, max_col=final_col_idx, 
                                min_row=self.wbs_start_row + 1, max_row=ws.max_row):
            entry_frame["entry"] = entry_counter
            null_val_counter = 0

            for header_counter, cell in enumerate(row):
                if cell.value is None or cell.value == "":
                    null_val_counter += 1

                if header_counter < len(header_list): 
                    normalized_value = self.normalize_data_value(cell.value)
                    if normalized_value == -1:
                        print(f"Error processing cell: {cell.coordinate}. Skipping entry.")
                        continue 

                    entry_frame[header_list[header_counter]] = normalized_value

            if null_val_counter == len(row):
                break
            
            json_dic["project_content"]["body"].append(entry_frame.copy())
            entry_counter += 1

        with open(file, 'w') as json_file:
            json.dump(json_dic, json_file, indent=4)

    def build_entry_frame(self, active_ws):
        ws = active_ws
        
        try:
            initial_col_idx = self.find_column_idx(ws, self.json_categories[0])
            final_col_idx = self.find_column_idx(ws, 'finish')
        except Exception as e:
            print(f"Error finding columns: {e}")
            return None
        
        dic_frame = {
            "entry": None 
        }

        for row in ws.iter_rows(min_col=initial_col_idx, max_col=final_col_idx, 
                                min_row=self.wbs_start_row, max_row=self.wbs_start_row): 
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_key = cell.value.replace(" ", "_").lower()
                    if new_key not in dic_frame:
                        dic_frame[new_key] = None

        return dic_frame

    def normalize_data_value(self, date_val: Union[str, date]) -> str:
        if isinstance(date_val, str):
            return date_val.strip()
        elif isinstance(date_val, date):
            return date_val.strftime('%d-%b-%y')
        else:
            return -1

    def create_json(self):
        if self.json_title is not None:
            json_basename = self.json_title + ".json"
        else:
            json_basename = input("Please enter the name for the new JSON file (Base Style): ") + ".json"

        self.json_basename = json_basename
        file = os.path.join(self.json_path, self.json_basename)
        
        try:
            with open(file, 'w') as json_file:
                json_pro = {
                    "project_metadata": {},
                    "project_content": {}
                } 

                json.dump(json_pro, json_file)
            print(f"New JSON file created: {file}")
            
        except Exception as e:
            print(f"An error occurred while creating the JSON file: {e}")

    def build_nested_dic(self, entry_frame):
        file = os.path.join(self.json_path, self.json_basename)
        flat_data = self.flattend_json_data(file)
        nested_dict = {}
        header_list = list(entry_frame.keys())

        for obj in flat_data:
            keys = [obj.get(category) for category in self.json_categories]
            key_one, key_two, key_three, key_four = keys

            if key_one is not None and key_one != None:
                if key_one not in nested_dict:
                    nested_dict[key_one] = {}
                
                if key_two is not None:
                    if key_two not in nested_dict[key_one]:
                        nested_dict[key_one][key_two] = {}
                    
                    if key_three is not None:
                        if key_three not in nested_dict[key_one][key_two]:
                            nested_dict[key_one][key_two][key_three] = {}
                        
                        if key_four is not None:
                            if key_four not in nested_dict[key_one][key_two][key_three]:
                                nested_dict[key_one][key_two][key_three][key_four] = []

                            entry_data = {header: obj.get(header) for header in header_list}
                            nested_dict[key_one][key_two][key_three][key_four].append(entry_data)

        return nested_dict
    
    def write_json(self, json_dic):
        if self.json_title is not None:
            json_basename = f"processed_{self.json_basename}"
        else:
            json_basename = input("Please enter the name for the new JSON file (Processed Style): ") + ".json"

        file = open(os.path.join(self.json_path, json_basename), 'w')

        json.dump(json_dic, file, indent=4)
            
        print(f"JSON data successfully created and saved to {json_basename}.")

    def flattend_json_data(self, file_path):
        flatten_list = []
        with open(file_path, 'r') as file:
            json_obj = json.load(file)

        for entry in json_obj["project_content"]["body"]:
            if "activities" in entry:
                if len(entry["activities"]) > 0:
                    flatten_list.extend(entry["activities"])
                else:
                    flatten_list.append(entry)
            else:
                flatten_list.append(entry)
        
        return flatten_list

    def flatten_col_typed_data(self, info_dic):
        json_dic = {
            "header": info_dic["header"],
            "filled": [],
            "empty": []
        }

        for activity in info_dic["body"]["filled"]:
            json_dic["filled"].append(activity["row"])

        for activity in info_dic["body"]["empty"]:
            json_dic["empty"].append(activity["row"])

        return json_dic

    def extract_col_cells(self, active_ws, column_header):
        ws = active_ws
        json_dic = {
            "header": column_header,
            "body": {
                "filled": [],
                "empty": []
            }
        }
        phase_col_idx = self.find_column_idx(ws, column_header)

        for row in ws.iter_rows(min_col=phase_col_idx, max_col=phase_col_idx, min_row=self.wbs_start_row + 1, max_row=ws.max_row):
            for cell in row:
                cell_info = {
                    "value": cell.value,
                    "row": cell.row - self.wbs_start_row
                }
                if cell.value is not None and cell.value != "" and isinstance(cell.value, str):
                    json_dic["body"]["filled"].append(cell_info)
                else:
                    json_dic["body"]["empty"].append(cell_info)

        return json_dic

    def find_column_idx(self, active_ws, column_header):
        ws = active_ws
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.replace(" ", "_").lower()

        for row in ws.iter_rows(min_row=self.wbs_start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def order_list(self, list):
        alphabtically_ordered_list = sorted(list)
        return alphabtically_ordered_list


if __name__ == '__main__':
    FilledExcelToUpdateJson.main(False)
