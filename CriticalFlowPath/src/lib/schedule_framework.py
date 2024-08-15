from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ScheduleFramework:
    def __init__(self, input_file_path, start_row, start_col):
        self.input_file_path = input_file_path
        self.start_row = int(start_row)
        self.start_col = int(start_col)

    def same_cell_values(self, starting_row_idx, starting_col_idx):
        # Load the workbook and select the relevant worksheet
        workbook = load_workbook(filename=self.input_file_path)
        worksheet = workbook['Gantt Schedule']

        # Access the specified row
        iterable_row = worksheet[starting_row_idx]  # Access the specified row
        last_value = iterable_row[starting_col_idx].value  # Store the value of the starting cell
        same_cell_list = []  # List to hold column indices of cells with the same values
        overall_list = []  # List to hold all groups of same-value cells

        # Iterate through the specified row starting from the starting column index
        for cell in iterable_row[starting_col_idx:]:  # Start from the specified column index
            if cell.value == last_value:  # Check if the current cell matches the last value
                same_cell_list.append(cell.column)  # Append the column index to the list
            else:
                if same_cell_list:  # If we have collected cells, append to overall_list
                    overall_list.append(same_cell_list)
                same_cell_list = [cell.column]  # Start a new list with the current cell's column
                last_value = cell.value  # Update last_value to the current cell's value

        # Append the last collected same-value cells if any
        if same_cell_list:
            overall_list.append(same_cell_list)

        workbook.save(self.input_file_path)
        workbook.close()
        return overall_list, starting_row_idx  # Return the overall list of same-value cells & row index

    def merge_same_value_cells(self, column_indices, starting_row_idx):
        # Load the workbook and select the relevant worksheet
        workbook = load_workbook(filename=self.input_file_path)
        worksheet = workbook['Gantt Schedule']

        # Ensure there are column indices to merge
        if not column_indices:
            print("No columns to merge.")
            return

        # Get the first and last columns to merge
        first_col = column_indices[0]
        last_col = column_indices[-1]

        # Merge the cells in the specified row
        worksheet.merge_cells(start_row=starting_row_idx, start_column=first_col, end_row=starting_row_idx, end_column=last_col)

        # Save the workbook after merging cells
        workbook.save(self.input_file_path)
        print(f"Merged cells from column {first_col} to {last_col} in row {starting_row_idx}.")
        workbook.close()

    def generate_body_schedule(self, start_date, end_date):
        # Load the workbook and create a new worksheet for the Gantt chart
        workbook = load_workbook(filename=self.input_file_path)
        worksheet = workbook.create_sheet('Gantt Schedule')

        # Convert string dates to datetime objects
        start_datetime_obj = datetime.strptime(start_date, '%d-%b-%y')
        end_datetime_obj = datetime.strptime(end_date, '%d-%b-%y')

        # Calculate the duration in days
        duration = (end_datetime_obj - start_datetime_obj).days + 1  # Include the end date

        # Fill the days row
        for day in range(duration):
            date = start_datetime_obj + timedelta(days=day)
            worksheet.cell(row=self.start_row + 2, column=self.start_col + day, value=date.strftime('%d'))

        # Fill the months row
        for day in range(duration):
            date = start_datetime_obj + timedelta(days=day)
            worksheet.cell(row=self.start_row + 1, column=self.start_col + day, value=date.strftime('%b'))

        # Fill the years row
        for day in range(duration):
            date = start_datetime_obj + timedelta(days=day)
            worksheet.cell(row=self.start_row, column=self.start_col + day, value=date.strftime('%y'))

        # Save the workbook after generating the Gantt chart
        workbook.save(self.input_file_path)
        print("Gantt chart generated successfully.")
        workbook.close()
    
    def style_worksheet(self):
        workbook = load_workbook(filename=self.input_file_path)
        worksheet = workbook['Gantt Schedule']

        # Define the border style
        thin_border = Border(left=Side(style='thin'))

        # Apply the font style, center alignment, and fill color to the header row (first row)
        for cell in worksheet.iter_rows(min_row=self.start_row, max_row=self.start_row, min_col=self.start_col):
            for cell in cell:  # Iterate through the cells in the header row
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="FFFFFF")  # Bold and white font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00800080', end_color='00800080', fill_type='solid')

        # Apply the font style, center alignment, and fill color to the second row (months)
        for cell in worksheet.iter_rows(min_row=self.start_row + 1, max_row=self.start_row + 1, min_col=self.start_col):
            for cell in cell:  # Iterate through the cells in the months row
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="00333333")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00CC99FF', end_color='00CC99FF', fill_type='solid')

        # Apply the font style, center alignment, and fill color to the third row (days)
        for cell in worksheet.iter_rows(min_row=self.start_row + 2, max_row=self.start_row + 2, min_col=self.start_col):
            for cell in cell:  # Iterate through the cells in the days row
                cell.font = Font(name='Century Gothic', size=12, bold=True, color="00000000")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')

        # Apply styles to other rows and accentuate the left border
        for row in worksheet.iter_rows(min_row=self.start_row + 3, min_col=self.start_col):  # Starting from the fourth row
            for cell in row:
                # Center align all cells
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Accent the left border for the last cell of each month
        for col in range(self.start_col, worksheet.max_column + 1):
            last_month_cell = worksheet.cell(row=self.start_row + 1, column=col)  # Get the month cell
            if last_month_cell.value is not None:  # Only apply border if there is a value
                # Apply border to all rows in this column starting from the third row
                for row in range(self.start_row + 2, worksheet.max_row + 1):  # Starting from the third row
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border

        workbook.save(self.input_file_path)
        print("Workbook styled and saved successfully.")
        workbook.close()

if __name__ == "__main__":
    input_file_path = input("Please enter the folder path to the Excel file: ")
    input_start_row = input("Please enter the starting row to write the schedule: ")
    input_start_col = input("Please enter the starting column to write the schedule: ")
    gantt_schedule = ScheduleFramework(input_file_path, input_start_row, input_start_col)
    gantt_schedule.generate_body_schedule('17-Feb-23', '16-Jun-25')

    # Merge year cells
    year_list, year_row = gantt_schedule.same_cell_values(gantt_schedule.start_row, gantt_schedule.start_col)
    for year in year_list:
        gantt_schedule.merge_same_value_cells(year, year_row)  # Only pass year 

    # Merge month cells
    month_list, month_row = gantt_schedule.same_cell_values(gantt_schedule.start_row + 1, gantt_schedule.start_col)
    for month in month_list:
        gantt_schedule.merge_same_value_cells(month, month_row)  # Only pass year

    gantt_schedule.style_worksheet()
