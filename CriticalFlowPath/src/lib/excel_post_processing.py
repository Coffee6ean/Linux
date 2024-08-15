from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ExcelPostProcessing:
    def __init__(self, input_file_path):
        self.input_file_path = input_file_path
        self.trade_list_validation = [
            'Carpenter',
            'Electrician',
            'Plumber',
            'Ironworker',
            'Sheet Metal Worker',
            'Pipefitter',
            'Laborer',
            'N/A',
            'Operating Engineer',
            'Bricklayer',
            'Cement Mason',
            'Roofer',
            'Glazier',
            'Painter',
            'Drywall Finisher',
            'Insulation Worker',
            'Elevator Constructor',
            'Millwright'
        ]
        self.phase_list_validation = [
            'Civil',
            'Comissioning',
            'Elevator',
            'Exteriors/Skin',
            'Foundations',
            'Framing Structure',
            'Garage',
            'Interior Finishes',
            'Interior Rough Ins',
            'Landscaping Decks',
            'N/A',
            'Roof',
            'Site Prep',
            'Site Utilities',
            'Structure',
            'Transformer Vault'
        ]
        self.scope_of_work_list_validation = [
            'Construction',
            'Concrete Works' # Foundations, slabs, etc
            'Demolition',  # Removing existing structures
            'Electrical',
            'Finishes',
            'Framing',
            'HVAC',  # Heating, Ventilation, and Air Conditioning
            'Insulation',  # Insulation installation
            'Landscaping',  # Outdoor work
            'N/A',
            'Painting',  # Interior and exterior painting
            'Plumbing',
            'Paving',  # Road and pathway work
            'Roofing',  # Work related to roofs
            'Site Work',
        ]
    
    def create_schedule_body(self):
        """
        Creates a Gantt chart in the Excel workbook based on activity start and finish dates.

        This function loads the Excel workbook, reads the activity data from the 'Project Content'
        sheet, and fills in the Gantt chart cells based on the duration of each activity.

        Assumptions:
        - Start dates are located in the 5th column (index 4).
        - Finish dates are located in the 6th column (index 5).
        - The Gantt chart starts from the 12th column (column L).
        """
        # Load the workbook and select the relevant worksheet
        workbook = load_workbook(filename=self.input_file_path)
        worksheet = workbook['Project Content']

        # Find the last column with data
        last_column = worksheet.max_column

        start_date_col = None
        end_date_col = None

        # Iterate over the columns to find the start and finish date columns
        for col in worksheet.iter_cols(min_col=1, max_col=last_column):
            if any('start' in cell.value for cell in col if cell.value):
                start_date_col = col[0].column
            elif any('finish' in cell.value for cell in col if cell.value):
                end_date_col = col[0].column
                break

        # Check if start and finish date columns were found
        if start_date_col is None or end_date_col is None:
            print("Error: Could not find start or finish date columns.")
            return

        # Iterate through the rows to process activities
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=last_column):
            # Extract start and finish dates
            activity_start = row[start_date_col - 1].value # Start date in the specified column
            activity_finish = row[end_date_col - 1].value  # Finish date in the specified column

            # Convert date strings to datetime objects
            if isinstance(activity_start, str):
                activity_start = datetime.strptime(activity_start, '%d-%b-%y')
            if isinstance(activity_finish, str):
                activity_finish = datetime.strptime(activity_finish, '%d-%b-%y')

            # Check for missing dates
            if activity_start is None or activity_finish is None:
                print(f"Skipping row {row[0].row} due to missing start or finish date.")
                continue  # Skip this row if dates are missing

            # Calculate the start and end column indices for the Gantt chart
            start_col = 12  # Gantt chart starts from column 12 (L)
            # Calculate the end column index based on the duration
            end_col = start_col + (activity_finish - activity_start).days  # Calculate end column index

            # Define the fill color for the Gantt chart cells
            fill_color = PatternFill(start_color="00CC99FF", end_color="00CC99FF", fill_type="solid")  # Light blue fill

            # Fill the cells in the Gantt chart
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row[0].row, column=col).fill = fill_color  # Fill the cell with color

        # Save the workbook after updating the Gantt chart
        workbook.save(self.input_file_path)
        print("Gantt chart created and saved successfully.")
        workbook.close()

    def validate_columns(self, header):
        """
        Applies data validation to columns in the 'Project Content' worksheet based on the specified header.

        Args:
            header (str): The header name to match columns for data validation.

        This function adds a drop-down list for valid entries in the specified columns.
        """
        # Load the workbook and select the relevant worksheet
        workbook = load_workbook(filename=self.input_file_path)
        ws = workbook['Project Content']

        # Access the first row
        first_row = ws['1']  # This gets the first row

        # Ensure that the first row is not empty
        if first_row is None:
            print("Error: The first row is empty.")
            return

        # Find the validation list based on the header
        if header == 'scope_of_work':
            validation_list = self.scope_of_work_list_validation
        elif header == 'phase':
            validation_list = self.phase_list_validation
        elif header == 'trade':
            validation_list = self.trade_list_validation
        else:
            print(f"Error: Invalid header '{header}'. Supported headers are 'scope_of_work', 'phase', and 'trade'.")
            return

        # Apply data validation to the columns that match the passed header
        header_columns = [col for col in first_row if col.value and header.lower() in str(col.value).lower()]

        if not header_columns:
            print(f"No columns found matching header: '{header}'")
            return

        for col in header_columns:
            col_index = col.column  # Get the index of the column (1-based)

            # Create a DataValidation object with the specified list
            dv = DataValidation(
                type='list',
                formula1=f'"{",".join(validation_list)}"',
                allow_blank=True
            )
            dv.error = 'Your entry is not in the list'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = 'Please select from the list'
            dv.promptTitle = 'List Selection'

            # Add data validation to the worksheet
            ws.add_data_validation(dv)

            # Define the range for data validation (from row 2 to the last row)
            dv.add(f'{get_column_letter(col_index)}2:{get_column_letter(col_index)}{ws.max_row}')

        # Save the workbook after applying data validation
        workbook.save(self.input_file_path)
        print(f"Data validation applied to columns matching '{header}' successfully.")
        workbook.close()

    def apply_post_processing(self):
        """
        Applies styling and data validation to the Excel workbook.
        """
        workbook = load_workbook(filename=self.input_file_path)
        ws = workbook['Project Content']

        # Access the first row
        first_row = ws['1']  # This gets the first row

        # Ensure that the first row is not empty
        if first_row is None:
            print("Error: The first row is empty.")
            return

        # Style the header row
        for cell in ws['1']:  # Assuming the first row contains headers
            cell.font = Font(name='Century Gothic', size=12, bold=True, color="FFFFFF")  # Bold and white font
            cell.fill = PatternFill(start_color="00800080", end_color="00800080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")  # Center alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.rows:    
            if row[0].value == 'N/A':
                parent_row = row
                for cell in parent_row:
                    cell.font = Font(name='Century Gothic', size=12, bold=True, color="00333333")  # Bold and white font
                    cell.fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(bottom=thin)

        workbook.save(self.input_file_path)
        print(f"Post-processing completed. Saved to {self.input_file_path}")
        workbook.close()


if __name__ == "__main__":
    post_processing = ExcelPostProcessing('/home/coffee_6ean/Linux/CriticalFlowPath/results/excel/output.xlsx')
    post_processing.validate_columns('scope_of_work')
    post_processing.validate_columns('phase')
    post_processing.validate_columns('trade')
    post_processing.apply_post_processing()
    post_processing.create_schedule_body()
