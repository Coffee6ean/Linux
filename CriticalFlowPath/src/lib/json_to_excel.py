import os
import json
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from pdf_to_jason import print_result

# Global variables
output_file_name = None
output_folder = None
output_file_path = None
list_validation = [
    'Carpenter',
    'Electrician',
    'Plumber',
    'Ironworker',
    'Sheet Metal Worker',
    'Pipefitter',
    'Laborer',
    'N/A',
    'Operating Engineer',
    'Bricklayer',
    'Cement Mason',
    'Roofer',
    'Glazier',
    'Painter',
    'Drywall Finisher',
    'Insulation Worker',
    'Elevator Constructor',
    'Millwright'
]

class JsonObj: 
    """
    A class for converting JSON data to an Excel file.
    """
    def __init__(self):
        pass
    
    def create_workbook(self):
        global output_file_name, output_folder, output_file_path  # Declare global variables

        # Prompt the user for the output folder
        output_folder = input("Please enter the folder path to save the Excel file: ")

        # Check if the specified folder exists
        if not os.path.exists(output_folder):
            print("Error: The specified folder does not exist.")
            return
        else:
            # If folder does exist, create new Workbook & Worksheet from openpyxl
            output_file_name = input("Please input file name: ")
            output_file_path = os.path.join(output_folder, f"{output_file_name}.xlsx")

            # Create a new workbook and select the active worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Project Content"  # Set a title for the worksheet

            return workbook, worksheet  # Return the workbook and worksheet for further use

    def retrieve_json_file(self):
        """
        Retrieves the JSON data from the 'results' folder.

        Returns:
            list: A list containing the JSON data.
        """
        json_input_path = os.path.join(os.getcwd(), 'results')
        if os.path.exists(json_input_path):
            print(f"Looking for JSON files in: {json_input_path}")
            json_data = []
            for json_input_file in os.listdir(json_input_path):
                if json_input_file.endswith("json"):
                    json_file_path = os.path.join(json_input_path, json_input_file)
                    with open(json_file_path, 'r') as f:
                        json_data.append(json.load(f))
            if not json_data:
                print('Error. Folder has no current JSON files')
        else:
            print('Error. Path not found in system')
        return json_data

    def json_results_to_data_frame(self, excel_file):
        """
        Writes the JSON data to an Excel file and applies data validation to 'trade' columns.

        Args:
            excel_file (str): The path to the Excel file where the data will be written.

        This function retrieves JSON data, processes it, and writes it to an Excel file.
        It creates two sheets: one for metadata and another for project content.
        Data validation is applied to columns containing the substring 'trade'.
        """
        # Retrieve JSON data from the specified source
        json_data = self.retrieve_json_file()

        # Check if any JSON data was retrieved
        if json_data:
            # Assuming we only need the first JSON object for processing
            data = json_data[0]  # If there are multiple JSON objects, handle accordingly

            # Prepare to write to Excel using the openpyxl engine
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # Write metadata to a separate sheet
                metadata = data['project_metadata']
                metadata_df = pd.DataFrame(metadata.items(), columns=['Field', 'Value'])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

                # Extract body from the JSON data
                body = data['project_content']['body']

                # Prepare to flatten the body with activities into a list
                flattened_data = []

                # Iterate through each entry in the body
                for entry in body:
                    # Create a dictionary for the parent activity row
                    parent_data = {
                        'id': entry['id'],
                        'name': entry['name'],
                        'trade': '',  # Placeholder for trade, to be filled later
                        'location': '',  # Placeholder for location
                        'duration': entry['duration'],
                        'start': entry['start'],
                        'finish': entry['finish'],
                        'total_float': entry['total_float'],
                        'activity_id': None,
                        'activity_name': None,
                        'activity_duration': None,
                        'activity_start': None,
                        'activity_finish': None,
                        'activity_total_float': None
                    }
                    flattened_data.append(parent_data)

                    # If there are activities associated with this entry, flatten them
                    if entry['activities']:
                        for activity in entry['activities']:
                            # Create a dictionary for each child activity row
                            child_data = {
                                'id': None,  # Parent ID is not needed here
                                'name': None,  # Parent name is not needed here
                                'duration': None,  # Parent duration is not needed here
                                'start': None,  # Parent start is not needed here
                                'finish': None,  # Parent finish is not needed here
                                'total_float': None,  # Parent total_float is not needed here
                                'activity_id': activity['id'],
                                'activity_name': activity['name'],
                                'activity_trade': '',  # Placeholder for activity trade
                                'activity_location': '',  # Placeholder for activity location
                                'activity_duration': activity['duration'],
                                'activity_start': activity['start'],
                                'activity_finish': activity['finish'],
                                'activity_total_float': activity['total_float']
                            }
                            flattened_data.append(child_data)

                # Convert the flattened data list to a DataFrame
                flattened_df = pd.DataFrame(flattened_data)

                # Write the flattened body to a new sheet in the Excel file
                flattened_df.to_excel(writer, sheet_name='Project Content', index=False)

                # Apply data validation to 'trade' columns
                trade_columns = [col for col in flattened_df.columns if 'trade' in col.lower()]
                for col in trade_columns:
                    col_index = flattened_df.columns.get_loc(col)  # Get the index of the column
                    # Create a data validation object for the 'trade' column
                    dv = DataValidation(type='list', formula1=f'"{",".join(list_validation)}"', allow_blank=True)
                    dv.error = 'Your entry is not in the list'
                    dv.errorTitle = 'Invalid Entry'
                    dv.prompt = 'Please select from the list'
                    dv.promptTitle = 'List Selection'
                    writer.sheets['Project Content'].add_data_validation(dv)
                    # Apply the validation to the specified column range
                    dv.add(f'{get_column_letter(col_index + 1)}2:{get_column_letter(col_index + 1)}{flattened_df.shape[0] + 1}')

            print(f"Successfully converted JSON to Excel and saved to {excel_file}")
        else:
            print("Error: No data to convert.")


if __name__ == "__main__":
    json_obj_instance = JsonObj()  # Create an instance of the JsonObj class
    wb, ws = json_obj_instance.create_workbook()  # Call the method to create the workbook

    if wb and ws:
        json_obj_instance.json_results_to_data_frame(output_file_path)  # Write JSON data to Excel
        print(f"Workbook created and saved at: {output_file_path}")
