import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

import re
import json
from pypdf import PdfReader
from pdf2image import convert_from_path
import pytesseract
from PIL import Image, ImageFilter, ImageEnhance
from datetime import datetime

class PdfDataProcessing():
    def __init__(self, input_file_path, input_file_basename, input_file_extension, 
                 input_worksheet_name, output_file_dir):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.output_file = output_file_dir
        self.worksheet_name = input_worksheet_name

        #Module Attributes
        self.xlsx_start_col = 'A'
        self.xlsx_start_row = 1

        #Structures
        self.to_fill_headers = ["phase", "area", "zone", "trade", "color", "activity_code"]
        self.filled_headers = ["wbs_code", "activity_name", "start", "finish"]
    
    @staticmethod
    def main(auto):
        if auto:
            project = PdfDataProcessing.auto_generate_ins()
        else:
            project = PdfDataProcessing.generate_ins()

        if project:
            output = PdfDataProcessing.return_valid_file(project.output_file, 'c')
            json_dict = PdfDataProcessing.read_data_from_json(
                project.input_path, 
                project.input_basename, 
                project.input_extension
            )
            
            processable_dict = project.retrieve_valid_data(json_dict)
            project.create_wbs_table_to_fill(processable_dict, output["path"], output["basename"])

    @staticmethod
    def generate_ins():
        input_file = PdfDataProcessing.return_valid_file(
            input("Please enter the path to the file or directory: ").strip()
        )
        output_file_dir = PdfDataProcessing.return_valid_path(
            "Please enter the directory to save the new module results: "
        )

        ins = PdfDataProcessing(
            input_file.get("path"), 
            input_file.get("basename"), 
            input_file.get("extension"),
            "Sheet1",
            output_file_dir
        )

        return ins
    
    @staticmethod
    def auto_generate_ins():
        pass

    @staticmethod
    def return_valid_file(input_file_dir:str, mode:str='r') -> dict|int:        
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = PdfDataProcessing._handle_dir(input_file_dir, mode)
        elif os.path.isfile(input_file_dir):
            file_dict = PdfDataProcessing._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str='r') -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = PdfDataProcessing._display_directory_files(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = input_file_dir, 
            basename = os.path.basename(input_file_basename).split(".")[0] if input_file_basename else "",
            extension = os.path.basename(input_file_basename).split(".")[-1] if input_file_basename else "",
        )
    
    @staticmethod
    def _display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def _handle_file(input_file_dir:str):
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in ["json"]):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return -1

    @staticmethod
    def return_valid_path(prompt_message:str) -> (str|None):
        while(True):   
            value = input(prompt_message).strip()
            try:
                if os.path.isdir(value):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")
    
    @staticmethod
    def clear_directory(directory_path):
        if os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            for file in file_list:
                file_path = os.path.join(directory_path, file)
                os.remove(file_path)

    @staticmethod
    def read_data_from_json(file_path:str, file_basename:str, file_extension:str="json"):
        basename = f"{file_basename}.{file_extension}"
        file = os.path.join(file_path, basename)
        
        with open(file, 'r') as json_file:
            data = json.load(json_file)

        return data

    def file_verification(self):
        file_path = self.pdf_file_path

        if os.path.isdir(file_path):
            file_list = os.listdir(file_path)
            selection = PdfDataProcessing.display_directory_files(file_list)

            file_name = file_list[selection]

            print(f'File selected: {file_name}')
            file = os.path.join(file_path, file_name)
            if PdfDataProcessing.is_pdf(file):
                return file
            else:
                return -1
        elif os.path.isfile(file_path):
            if PdfDataProcessing.is_pdf(file_path):
                return file_path
            else:
                return -1
        else: 
            print('Error. Please verify the directory and file exist and that the file is of type .pdf')    
            return -1
        
    def read_pdf_file(self):
        pdf = self.file_verification()
        if PdfDataProcessing.is_pdf(pdf):
            reader = PdfReader(pdf)
            first_page = reader.pages[0]
            content_str = first_page.extract_text()
            print('#-------------#')
            print(content_str)

    def pdf_to_img(self, output_directory):
        pdf = self.file_verification()
        if pdf:
            images = convert_from_path(pdf)
            os.makedirs(output_directory, exist_ok=True)

            for i, image in enumerate(images):
                #processed_image = self.process_image(image)

                image_path = os.path.join(output_directory, f'page_{i + 1}.jpg')
                image.save(image_path, 'JPEG')

            print(f"Images saved to: {output_directory}")
        else:
            print("Error: Invalid PDF file.")
    
    def process_image(self, image):
        image = image.convert('L')
        image = image.filter(ImageFilter.MedianFilter(size=3))

        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2)
        threshold = 128
        image = image.point(lambda p: 255 if p > threshold else 0)

        return image

    def read_pdf_img(self, input_directory):
        img_path = os.path.join(input_directory, "page_1.jpg")
        image = Image.open(img_path)
        process_img = self.process_image(image)
        text = pytesseract.image_to_string(process_img)
        print('#-------------#')
        print(text)

    def retrieve_valid_data(self, json_dict:list):
        processable_content = []

        for item in json_dict:
            if item["body"].get("content"):
                processable_content.extend(item["body"].get("content"))

        return processable_content

    def create_wbs_table_to_fill(self, data:list, output_path:str, output_basename:str) -> None:
        if not output_basename:
            output_basename = input("Please enter the name to save the new .xlsx file: ")        

        proc_table = self._generate_wbs_to_fill(data)
        self._write_data_to_excel(proc_table, output_path, output_basename)

    def _generate_wbs_to_fill(self, json_dict:dict):
        df_table = pd.DataFrame(json_dict)
        df_table.fillna("N/A", inplace=True)

        reset_table = df_table.reset_index()
        fill_table = self._add_new_empty_columns(reset_table)

        column_header_list = fill_table.columns.tolist()

        temp_list_1 = self._check_column_order(
            column_header_list, 
            self._bring_column_to_front(column_header_list, "wbs_code")
        )
        temp_list_2 = self._check_column_order(
            temp_list_1, 
            self._send_column_to_back(temp_list_1, "finish")
        )

        fill_table = fill_table[temp_list_2]

        return fill_table

    def _add_new_empty_columns(self, proc_table):
        for item in self.to_fill_headers:
            proc_table[item] = ""

        return proc_table
    
    def _check_column_order(self, column_header_list:list, func) -> list:
        if column_header_list[-1] != "finish":
            ordered_header_list = func
        else:
            ordered_header_list = column_header_list

        return ordered_header_list

    def _bring_column_to_front(self, column_list:list, target_column:str) -> list:
        for col in column_list:
            if col == target_column:
                target_idx = column_list.index(target_column)
                target = column_list.pop(target_idx)
                ordered_list = [target] + column_list
                break
        
        return ordered_list
    
    def _send_column_to_back(self, column_list:list, target_column:str) -> list:
        for idx, col in enumerate(column_list):
            if col == target_column:
                temp = column_list[-1]
                column_list[-1] = col
                column_list[idx] = temp
                break
        
        return column_list

    def _write_data_to_excel(self, proc_table, excel_path:str, 
                            excel_basename:str) -> None:
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            basename = excel_basename + '.xlsx'
            file = os.path.join(excel_path, basename)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='w') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name=self.worksheet_name, 
                        startrow=self.xlsx_start_row - 1, 
                        startcol=column_index_from_string(self.xlsx_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {self.worksheet_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")


if __name__ == "__main__":
    PdfDataProcessing.main(False)
        