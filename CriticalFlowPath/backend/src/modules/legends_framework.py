import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from backend.config.paths import RSLTS_DIR

class LegendsFramework():
    def __init__(self, input_file_path, input_file_basename, input_file_extension, 
                 input_worksheet_name, project_table, input_start_row=1, input_start_col='A'):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.worksheet_name = input_worksheet_name
        self.table = project_table

        #Module Attributes
        self.start_row = int(input_start_row)
        self.start_col = str(input_start_col)
        self.default_hex_fill_color = "00FFFF00"
        self.dark_default_hex_font = "00000000"
        self.light_default_hex_font = "00FFFFFF"

        #Module Results
        self.module_data = {
            "logs": {
                "start": LegendsFramework.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            }
        }

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, 
             input_file_extension=None, input_worksheet_name=None, project_table=None):
        if auto:
            project = LegendsFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_worksheet_name, 
                project_table,
            )
        else:
            project = LegendsFramework.generate_ins()
        
        if project:
            active_workbook, active_worksheet = project.return_excel_workspace(auto, project.worksheet_name)

            if active_workbook and active_worksheet:
                project.generate_legends_table(
                    active_workbook, 
                    active_worksheet, 
                    project.table.reset_index()
                )
            else:
                project.module_data["logs"]["status"].append(dict(
                    Error= f"{LegendsFramework.__name__}| Could not open Excel file as Workbook & Worksheet"
                ))
        else:
            project.module_data["logs"]["status"].append(dict(
                Error= f"{LegendsFramework.__name__}| Module's instance was not generated correctly"
            ))

        project.module_data["logs"]["finish"] = LegendsFramework.return_valid_date()
        project.module_data["logs"]["run-time"] = LegendsFramework.calculate_time_duration(
            project.module_data["logs"].get("start"), 
            project.module_data["logs"].get("finish")
        )
        project.module_data["logs"]["status"].append(dict(
            Info=f"{LegendsFramework.__name__}| Module ran successfully"
        ))

        return project.module_data

    @staticmethod
    def generate_ins():
        LegendsFramework.ynq_user_interaction(
            "Run as Module as stand-alone? "
        )

        setup_cls = setup.Setup.main()
        
        ins = LegendsFramework(
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            "CFA - Legends",
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("table"),
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, 
                          input_worksheet_name, project_table):
        ins = LegendsFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_worksheet_name, 
            project_table,
        )
        
        return ins

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    @staticmethod
    def ynq_user_interaction(prompt_message:str) -> str:
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def hex2rgb(hex_color:str) -> tuple:
        trimmed_hex = hex_color.lstrip('#')
        calc_rgb = tuple(int(trimmed_hex[i:i+2], 16) for i in (0, 2, 4))

        return calc_rgb

    @staticmethod
    def calculateLuminance(R:float, G:float, B:float):
        lum = 0.2126*(R/255.0)**2.2 + 0.7152*(G/255.0)**2.2 + 0.0722*(B/255.0)**2.2

        return lum

    @staticmethod
    def write_data_to_json(file_title:str, json_dict:dict):
        file = os.path.join(RSLTS_DIR, file_title)
        
        with open(file, 'w') as writer:
            json.dump(json_dict, writer)

    def return_excel_workspace(self, auto:bool, worksheet_name:str):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            if auto:
                workbook = load_workbook(filename=file)
                worksheet = workbook.create_sheet(worksheet_name)
                self.ws_name = worksheet_name
                print(f"New worksheet '{worksheet_name}' created.\n")

                return workbook, worksheet
            else:
                workbook = load_workbook(filename=file)
                worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{worksheet_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = LegendsFramework.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.\n")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def generate_legends_table(self, active_workbook, active_worksheet, og_table) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        wb = active_workbook
        ws = active_worksheet
        reset_table = og_table.reset_index()

        proc_table = pd.pivot_table(
            reset_table,
            index=["entry", "phase", "trade", "color"],
            values=["activity_name", "activity_code"],
            aggfunc="first",
            observed=True
        )

        column_width_dict = self._define_column_width(proc_table)

        phases_df = dict(tuple(proc_table.groupby("phase", observed=True)))

        try:
            self._print_phase_structure(ws, phases_df, column_width_dict)
            print("Legends tables generated successfully.")
            self.module_data["logs"]["status"].append(dict(
                Info= f"{LegendsFramework.__name__}| Legends tables generated successfully."
            ))
        except Exception as e:
            print("Failed to generate legends tables.")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{LegendsFramework.__name__}| Failed to generate legends tables."
            ))
        finally:
            wb.save(filename=file)

    def _define_column_width(self, proc_table) -> dict:
        phase_widths = {}

        for phase, group in proc_table.groupby(level="phase", observed=True):
            max_length_code = max(len(str(activity)) for activity in group["activity_code"].dropna())
            max_length_name = max(len(str(activity)) for activity in group["activity_name"].dropna())
            phase_widths[phase] = dict(
                code = max_length_code + 2,
                name = max_length_name + 2,
            )

        return phase_widths

    def _print_phase_structure(self, active_worksheet, df_dict:dict, column_width_dict:dict) -> None:
        ws = active_worksheet
        current_row = self.start_row
        current_col = column_index_from_string(self.start_col)

        for phase_key, phase_data in df_dict.items():
            self._style_table_header(
                ws,
                current_row,
                current_col,
                phase_key,
                phase_key,
                "00800080",
                phase_key,
                column_width_dict
            )

            self._print_trade_structure(
                ws,
                current_row + 2,
                current_col,
                phase_key,
                phase_data,
                column_width_dict
            )

            current_col += 3

    def _print_trade_structure(self, active_worksheet, current_row:int, current_col:int,
                          phase_key: str, phase_values, column_width_dict: dict) -> None:
        ws = active_worksheet
        trades_df = dict(tuple(phase_values.groupby("trade", observed=True)))

        for trade_key, trade_data in trades_df.items():
            reset_trade_table = trade_data.reset_index()

            for idx, row in reset_trade_table.iterrows():
                trade, color = row['trade'], row['color']
                activity_code, activity_name = row['activity_code'], row['activity_name']

                if idx == 0:
                    self._style_table_header(
                        ws,
                        current_row + idx,
                        current_col,
                        trade,
                        trade,
                        color,
                        phase_key,
                        column_width_dict
                    )

                self._style_table_structure(
                    ws,
                    current_row + idx + 1,
                    current_col,
                    activity_code,
                    activity_name,
                    color,
                    phase_key,
                    column_width_dict
                )

            current_row += len(reset_trade_table) + 2

    def _style_table_header(self, active_worksheet, current_row:int, current_col:int, 
                            label:str, value:str, color:str,
                            column_with_key:str, column_width_dict:dict) -> None:
        ws = active_worksheet
        if label == value:
            self._style_cell(
                ws,
                current_row,
                current_col,
                label,
                color,
                column_width_dict[column_with_key].get("name")
            )

            ws.merge_cells(
                start_row=current_row, 
                start_column=current_col, 
                end_row=current_row, 
                end_column=current_col + 1
            )

    def _style_table_structure(self, active_worksheet, current_row:int, current_col:int, 
                               label:str, value:str, color:str, 
                               column_with_key:str, column_width_dict:dict) -> None:
        ws = active_worksheet
        self._style_cell(
            ws, 
            current_row, 
            current_col, 
            label, 
            color,
            column_width_dict[column_with_key].get("code")
        )
        self._style_cell(
            ws, 
            current_row, 
            current_col + 1, 
            value, 
            "00FFFFFF",
            column_width_dict[column_with_key].get("name")
        )

    def _style_cell(self, active_worksheet, current_row:int, current_col:int, val:str, 
                    color_hex:str, cell_width:int=50, horz_alignment:str="center") -> None:
        ws = active_worksheet
        try:
            cell = ws.cell(row=current_row, 
                            column=current_col, 
                            value=val)
            cell.alignment = Alignment(horizontal=horz_alignment, 
                                        vertical="center")
            
            cell.border = Border(left=Side(border_style="thin", color="00000000"),
                                right=Side(border_style="thin", color="00000000"),
                                top=Side(border_style="thin", color="00000000"),
                                bottom=Side(border_style="thin", color="00000000"))

            cell.fill = PatternFill(start_color=self._process_hex_val(color_hex), 
                                    end_color=self._process_hex_val(color_hex), 
                                    fill_type="solid")
            
            rgb = LegendsFramework.hex2rgb(color_hex)
            lum_rgb = LegendsFramework.calculateLuminance(rgb[0], rgb[1], rgb[2])

            if lum_rgb > 0.5:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.dark_default_hex_font)
            else:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.light_default_hex_font)

            ws.column_dimensions[get_column_letter(cell.column)].width = cell_width
        except Exception as e:
            print(f"Error styling cell at row {current_row}, column {current_col}: {e}")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{LegendsFramework.__name__}| Error styling cell at row {current_row}, column {current_col}: {e}"
            ))

    def _process_hex_val(self, hex_val:str) -> str:
        if '#' in hex_val:
            hex_val = hex_val.replace('#', "00")

        if len(hex_val) != 8:
            print(f"Error: Invalid hex value '{hex_val}'. Hex values must be 6 characters long.\n")
            self.module_data["logs"]["status"].append(dict(
                Warning= f"{LegendsFramework.__name__}| Invalid '{hex_val}' value."
            ))
            return self.default_hex_fill_color

        return hex_val


if __name__ == "__main__":
    LegendsFramework.main(False)
