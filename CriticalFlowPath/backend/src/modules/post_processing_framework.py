import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from backend.config.paths import RSLTS_DIR

class PostProcessingFramework():
    def __init__(self, input_file_path, input_file_basename, input_file_extension, 
                 input_file_workweek,project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                 project_lead_struct, project_duration_processed, project_start_date, project_finish_date, time_scale):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.input_workweek = input_file_workweek
        self.worksheet_name = project_worksheet_name
        self.table = project_table
        self.ordered_dict = project_ordered_dict
        self.phase_order = project_phase_order
        self.lead_struct = project_lead_struct
        self.duration_processed = project_duration_processed
        self.start_date = project_start_date
        self.finish_date = project_finish_date
        self.time_scale = time_scale

        #Module Attributes
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.LIGHT_GRAY = "00F2F2F2"
        self.DARK_GRAY = "00C0C0C0"

        #Structures
        self.time_scale_options = ["d", "w"]
        self.json_struct_categories = [
            "phase", 
            "area", 
            "zone", 
            "subzone",
            "level",
            "trade", 
            "activity_code"
        ]
        self.wbs_final_categories = {
            "phase": "thin",
            "area": "dashDot", 
            "zone": "dashed",
            "subzone": "dotted",
            "level": "hair",
        }
        self.calendar_weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        #Module Results
        self.module_data = {
            "logs": {
                "start": PostProcessingFramework.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            }
        }

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, input_file_extension=None, 
             input_file_workweek=None, project_worksheet_name=None, project_table=None, project_ordered_dict=None, project_phase_order=None, 
             project_lead_struct=None, project_duration_processed=None, project_start_date=None, project_finish_date=None, time_scale=None):
        if auto:
            project = PostProcessingFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_file_workweek, 
                project_worksheet_name,
                project_table, 
                project_ordered_dict, 
                project_phase_order, 
                project_lead_struct,
                project_duration_processed,
                project_start_date, 
                project_finish_date,
                time_scale
            )
        else:
            project = PostProcessingFramework.generate_ins()


        if project:
            if not project.time_scale:
                print("Please choose the time scale to process the schedule: ")
                scale_idx = PostProcessingFramework.display_options(project.time_scale_options)
                project.time_scale = project.time_scale_options[scale_idx[0] - 1]

            active_workbook, active_worksheet = project.return_excel_workspace(project.worksheet_name)

            if active_workbook and active_worksheet:
                project.update_schedule_size(
                    active_workbook, 
                    active_worksheet, 
                    project.ordered_dict, 
                    project.lead_struct,
                    project.time_scale
                )

                project.update_schedule_style(
                    active_workbook, 
                    active_worksheet,
                    project.lead_struct, 
                    project.json_struct_categories
                )
            else:
                project.module_data["logs"]["status"].append(dict(
                    Error= f"{PostProcessingFramework.__name__}| Could not open Excel file as Workbook & Worksheet"
                ))
        else:
            project.module_data["logs"]["status"].append(dict(
                Error= f"{PostProcessingFramework.__name__}| Could not open Excel file as Workbook & Worksheet"
            ))

        project.module_data["logs"]["finish"] = PostProcessingFramework.return_valid_date()
        project.module_data["logs"]["run-time"] = PostProcessingFramework.calculate_time_duration(
            project.module_data["logs"].get("start"), 
            project.module_data["logs"].get("finish")
        )
        project.module_data["logs"]["status"].append(dict(
            Info=f"{PostProcessingFramework.__name__}| Module ran successfully"
        ))

        return project.module_data

    @staticmethod
    def generate_ins():
        PostProcessingFramework.ynq_user_interaction(
            "Run as Module as stand-alone? "
        )

        setup_cls = setup.Setup.main()

        ins = PostProcessingFramework(
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            setup_cls.obj["project"]["modules"]["MODULE_3"]["details"].get("workweek"),
            "CFA - WBS Table",
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("table"),
            setup_cls.obj["project"]["modules"]["MODULE_3"].get("content"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("custom_phase_order"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("lead_schedule_struct"),
            setup_cls.obj["project"]["modules"]["MODULE_3"]["details"]["calendar"]["processed"]["days"].get("total"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("start_date"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("finish_date"),
            None
        )

        return ins

    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, 
                          input_file_workweek, project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                          project_lead_struct, project_duration_processed, project_start_date, project_finish_date, time_scale):
        ins = PostProcessingFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_file_workweek,
            project_worksheet_name,
            project_table, 
            project_ordered_dict, 
            project_phase_order, 
            project_lead_struct,
            project_duration_processed,
            project_start_date, 
            project_finish_date,
            time_scale
        )

        return ins

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    @staticmethod
    def ynq_user_interaction(prompt_message:str) -> str:
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def binary_user_interaction(prompt_message:str) -> bool:
        valid_responses = {'y', 'n'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                if user_input == 'y':
                    return True 
                else:
                    return False 
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No]\n")

    @staticmethod
    def display_options(file_list:list) -> list:
        if not file_list:
            print('Error: No elements found.')
            return []

        print(f'-- {len(file_list)} elements found:')
        for idx, file in enumerate(file_list, start=1):
            print(f'{idx}. {file}')

        result = []
        selection_input = input('\nEnter index numbers (comma-separated) to select elements to process: ').split(',')

        for selection in selection_input:
            selection = selection.strip()
            if not selection.isdigit():
                print(f'Error: Invalid input "{selection}", skipping.')
                continue

            index = int(selection)
            if 1 <= index <= len(file_list):
                result.append(index)
            else:
                print(f'Error: "{index}" is out of range (1 to {len(file_list)}).')

        return result

    @staticmethod
    def date_range(start:datetime, finish:datetime) -> list:
        days = []
        current = start
        while current <= finish:
            days.append(datetime.strftime(current, '%d-%b-%Y'))
            current += timedelta(days=1)
        
        return days

    def return_excel_workspace(self, worksheet_name:str):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.worksheet_name = worksheet_name
                    print(f"New worksheet '{self.worksheet_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = PostProcessingFramework.display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.worksheet_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.worksheet_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def update_schedule_size(self, active_workbook, active_worksheet, custom_ordered_dict:dict, 
                             lead_schedule_struct:str, time_scale:str, alloted_space:int=2) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        overlap_results = self._calculate_overlapping_dates(
            custom_ordered_dict, 
            time_scale
        )
        available_results = self._calculate_available_instances(custom_ordered_dict)

        self._process_file(
            active_worksheet, 
            lead_schedule_struct, 
            overlap_results, 
            available_results,
            alloted_space
        )

        try:
            active_workbook.save(filename=file)
            print("CFA Schedule frame successfully updated.")
            self.module_data["logs"]["status"].append({
                "Info": f"{PostProcessingFramework.__name__}| CFA Schedule Frame successfully updated."
            })
        except Exception as e:
            print(f"Failed to update WBS Table: {e}")
            self.module_data["logs"]["status"].append({
                "Error": f"{PostProcessingFramework.__name__}| Failed to update CFA Frame: {e}"
            })
        finally:
            active_workbook.close()

    def _calculate_overlapping_dates(self, custom_ordered_dict:dict, time_scale:str='d') -> dict:
        if not isinstance(custom_ordered_dict, list) or not custom_ordered_dict:
            raise ValueError("_calculate_overlapping_dates received invalid or empty list")
    
        lead_based_lists = []
        nested_list = []
        ref_lead = self._generate_compound_category_name(custom_ordered_dict[0])

        for item in custom_ordered_dict:
            comp_lead_cat_title = self._generate_compound_category_name(item)

            if comp_lead_cat_title == ref_lead:
                nested_list.append(item)
            else:
                lead_based_lists.append(nested_list)
                nested_list = [item]
                ref_lead = comp_lead_cat_title

        if nested_list:
            lead_based_lists.append(nested_list)

        sorted_entries_list = []
        for location_list in lead_based_lists:
            sorted_list = self._bubble_sort_entries_by_dates(location_list)
            sorted_entries_list.append(sorted_list)
        
        overlap_results = self._get_overlap(
            sorted_entries_list,
            time_scale
        )

        return overlap_results

    def _generate_compound_category_name(self, item:dict) -> str:
        if not isinstance(item, dict):
            raise ValueError("_generate_compound_category_name expected a dict")
    
        category_names = []

        for category in self.wbs_final_categories.keys():
            cat_name = item.get(category)
            if cat_name:
                category_names.append(cat_name)

        return "|".join(category_names)

    def _bubble_sort_entries_by_dates(self, unsorted_list:list) -> list:
        if not unsorted_list:
            raise ValueError("_bubble_sort_entries_by_dates received empty list")

        n = len(unsorted_list)
        for i in range(n):
            for j in range(0, n - i - 1):
                try:
                    date_start_j = datetime.strptime(unsorted_list[j]["start"], "%d-%b-%Y")
                    date_start_j1 = datetime.strptime(unsorted_list[j + 1]["start"], "%d-%b-%Y")
                except Exception as e:
                    raise ValueError(f"Invalid 'start' date format in entry: {e}")

                if date_start_j > date_start_j1:
                    unsorted_list[j], unsorted_list[j + 1] = unsorted_list[j + 1], unsorted_list[j]
                elif date_start_j == date_start_j1:
                    try:
                        date_end_j = datetime.strptime(unsorted_list[j]["finish"], "%d-%b-%Y")
                        date_end_j1 = datetime.strptime(unsorted_list[j + 1]["finish"], "%d-%b-%Y")
                    except Exception as e:
                        raise ValueError(f"Invalid 'finish' date format in entry: {e}")

                    if date_end_j > date_end_j1:
                        unsorted_list[j], unsorted_list[j + 1] = unsorted_list[j + 1], unsorted_list[j]

        return unsorted_list
    
    def _get_overlap(self, location_based_lists: list, time_scale: str) -> dict:
        try:
            if time_scale == 'd':
                return self._day_based_overlap(location_based_lists)
            elif time_scale == 'w':
                return self._week_based_overlap(location_based_lists)
            else:
                raise ValueError(f"Invalid time scale: '{time_scale}'. Expected 'd' (day) or 'w' (week).")
        except Exception as e:
            err_msg = f"{self.__class__.__name__}| {str(e)}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{PostProcessingFramework.__name__}| {err_msg}"
            ))

            return {}

    def _day_based_overlap(self, location_based_lists:list) -> dict:
        category_results = {}

        for lead_list in location_based_lists:
            if not lead_list:
                continue

            try:
                ref_category = self._generate_compound_category_name(lead_list[0])
            except Exception as e:
                print(f"Failed to generate category name for {lead_list[0]}, error: {e}")
                continue

            active_overlaps = []
            current_max = 0

            for activity in lead_list:
                try:
                    start = datetime.strptime(activity["start"], "%d-%b-%Y")
                    finish = datetime.strptime(activity["finish"], "%d-%b-%Y")
                except Exception as e:
                    print(f"Invalid date in activity: {activity}, error: {e}")
                    continue

                active_overlaps = [
                    (active_start, active_finish)
                    for active_start, active_finish in active_overlaps
                    if finish >= active_start and start <= active_finish
                ]

                active_overlaps.append((start, finish))
                current_max = max(current_max, len(active_overlaps))

            category_results[ref_category] = max(category_results.get(ref_category, 0), current_max)

        return category_results
        
    def _week_based_overlap(self, location_based_lists:list) -> dict:
        category_results = {}

        for lead_list in location_based_lists:
            if not lead_list:
                continue

            try:
                ref_category = self._generate_compound_category_name(lead_list[0])
            except Exception as e:
                print(f"Failed to generate category name for lead_list[0]: {lead_list[0]}, error: {e}")
                continue

            active_weeks = {}
            current_max = 0

            for activity in lead_list:
                try:
                    start = datetime.strptime(activity["start"], "%d-%b-%Y")
                    finish = datetime.strptime(activity["finish"], "%d-%b-%Y")

                    week_start, week_finish = self._calculate_week(start, finish)
                    weeks = set(self._get_week_range(week_start, week_finish))

                    level = 0
                    while level in active_weeks and active_weeks[level].intersection(weeks):
                        level += 1

                    active_weeks.setdefault(level, set()).update(weeks)
                    current_max = max(current_max, level + 1)

                except Exception as e:
                    print(f"Invalid week overlap data for activity: {activity}, error: {e}")
                    continue

            category_results[ref_category] = max(category_results.get(ref_category, 0), current_max)

        return category_results

    def _calculate_week(self, start_date:datetime, finish_date:datetime) -> tuple:
        try:
            weekday_map = {day: i for i, day in enumerate(self.calendar_weekdays)}
            target_start = weekday_map[self.input_workweek[0]]
            target_end = weekday_map[self.input_workweek[-1]]

            start_offset = (target_start - start_date.weekday()) % 7
            week_start = start_date + timedelta(days=start_offset)

            end_offset = (target_end - finish_date.weekday()) % 7
            week_finish = finish_date + timedelta(days=end_offset)

            if week_start > finish_date:
                start_offset = (target_start - finish_date.weekday()) % 7
                week_start = finish_date + timedelta(days=start_offset)

            if week_finish < start_date:
                end_offset = (target_end - start_date.weekday()) % 7
                week_finish = start_date + timedelta(days=end_offset)

            return week_start, week_finish

        except Exception as e:
            print(f"[ERROR] Failed to calculate week with start_date={start_date} and finish_date={finish_date}: {e}.")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{PostProcessingFramework.__name__}| Failed to calculate week with start_date ({start_date}) and finish_date ({finish_date}): {e}."
            ))
            return start_date, finish_date  # safe fallback
    
    def _get_week_range(self, start_date:datetime|str, finish_date:datetime|str, obj_id:int=-1) -> list:
        try:
            if isinstance(start_date, str):
                start_date = datetime.strptime(start_date, '%d-%b-%Y')

            if isinstance(finish_date, str):
                finish_date = datetime.strptime(finish_date, '%d-%b-%Y')

            if start_date > finish_date:
                print(f"[Warning] start_date {start_date} is after finish_date {finish_date}. Returning empty range.")
                self.module_data["logs"]["status"].append(dict(
                    Warning= f"{PostProcessingFramework.__name__}| Failed to get week range for {obj_id}, start_date ({start_date}) is after finish_date ({finish_date})."
                ))
                return []

            step = timedelta(days=len(self.calendar_weekdays))
            count = ((finish_date - start_date).days // step.days) + 1
            return [start_date + i * step for i in range(count)]

        except Exception as e:
            print(f"[Error] _get_week_range with start_date={start_date}, finish_date={finish_date}: {e}")
            return []

    def _calculate_available_instances(self, custom_ordered_dict:dict) -> dict:
        entry_dict_available = {}

        for item in custom_ordered_dict:
            try:
                comp_lead_cat_title = self._generate_compound_category_name(item)
                entry_dict_available[comp_lead_cat_title] = entry_dict_available.get(comp_lead_cat_title, 0) + 1
            except Exception as e:
                print(f"[Error] Failed to generate compound name for item {item}: {e}.")
                self.module_data["logs"]["status"].append(dict(
                    Error= f"{PostProcessingFramework.__name__}| Failed to generate compound name for item {item}: {e}."
                ))
                continue  # skip bad entry, keep counting

        return entry_dict_available

    def _process_file(self, active_worksheet, lead_schedule_struct:str, 
                     overlap_results:dict, available_results:dict, alloted_space:int) -> None:
        ws = active_worksheet
        
        wbs_start_col = column_index_from_string(self.wbs_start_col)
        start_col_idx = self._find_column_idx(ws, lead_schedule_struct, self.wbs_start_row) + 1
        end_col_idx = self._find_first_nonempty_string_column_idx(ws, 1, 1)

        self._unmerge_columns(ws, wbs_start_col, end_col_idx)
        self._delete_excess_rows(
            ws, 
            overlap_results, 
            available_results,
            alloted_space
        )

        self._delete_columns(ws, start_col_idx, end_col_idx)
        self._insert_columns(ws, start_col_idx, end_col_idx - start_col_idx)
    
    def _find_column_idx(self, active_worksheet, column_header:str, start_row:int) -> int | None:
        try:
            ws = active_worksheet
            start_col_idx = column_index_from_string(self.wbs_start_col)
            normalized_header = column_header.strip().replace(" ", "_").lower()

            for row in ws.iter_rows(
                min_row=start_row,
                min_col=start_col_idx,
                max_col=ws.max_column
            ):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        normalized_cell_value = cell.value.strip().replace(" ", "_").lower()
                        if normalized_header == normalized_cell_value:
                            return cell.column

            prompt = f"Column header '{column_header}' not found. Use default column 'A'?"
            if PostProcessingFramework.binary_user_interaction(prompt):
                return 0
            return None

        except Exception as e:
            print(f"[ERROR] Failed to find column '{column_header}': {e}.")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{PostProcessingFramework.__name__}| Failed to find column '{column_header}': {e}."
            ))

            return None

    def _find_first_nonempty_string_column_idx(self, active_worksheet, start_row_idx:int, start_col_idx:int) -> int:
        try:
            ws = active_worksheet
            for row in ws.iter_rows(min_row=start_row_idx, min_col=start_col_idx, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        return cell.column

            raise ValueError(f"Failed to determine first valid column.")

        except Exception as e:
            print(f"[ERROR] {e}")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{PostProcessingFramework.__name__}| {e}."
            ))

            return -1

    def _unmerge_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        try:
            for merged_cell in list(ws.merged_cells.ranges):
                if (merged_cell.bounds[0] >= start_column_idx and 
                    merged_cell.bounds[2] <= end_column_idx):
                    ws.unmerge_cells(str(merged_cell))

            start_col_letter = get_column_letter(start_column_idx)
            end_col_letter = get_column_letter(end_column_idx)

            print(f"Column unmerging applied successfully: [{start_col_letter}, {end_col_letter}].")
            self.module_data["logs"]["status"].append({
                "Info": f"{PostProcessingFramework.__name__}| Column unmerging applied successfully: [{start_col_letter}, {end_col_letter}]."
            })
        except Exception as e:
            print(f"[ERROR] Failed during column unmerging: {e}")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{PostProcessingFramework.__name__}| Failed during column unmerging: {e}"
            ))
            
    def _delete_excess_rows(self, active_worksheet, overlap_results:dict, 
                                 available_results:dict, alloted_space:int) -> None:
        ws = active_worksheet
        lead_idx = self.json_struct_categories.index(self.lead_struct)
        grouping_levels = self.json_struct_categories[:lead_idx + 1]
        rows_to_delete = []
        starting_point = self.wbs_start_row + 1

        # Build rows_dict following the table group order
        rows_dict = {}
        current_row = starting_point
        grouped = self.table.groupby(level=grouping_levels, observed=True)

        for group_key, group_data in grouped:
            compound_key = "|".join(group_key)
            num_rows = available_results.get(compound_key, 0)
            for _ in range(num_rows):
                rows_dict[current_row] = compound_key
                current_row += 1

        # Now determine which rows to delete
        for group_key, group_data in grouped:
            compound_category_name = "|".join(group_key)
            allowed_rows = min(
                available_results.get(compound_category_name, 0),
                overlap_results.get(compound_category_name, 0) + alloted_space
            )

            # Get all worksheet rows associated with this group
            group_rows = [row for row, key in rows_dict.items() if key == compound_category_name]

            # If there are more rows than allowed, mark the excess for deletion
            if len(group_rows) > allowed_rows:
                excess = group_rows[allowed_rows:]  # keep the first 'allowed_rows', delete the rest
                rows_to_delete.extend(excess)

        # Sort in reverse to delete from bottom up
        try:
            rows_to_delete.sort(reverse=True)
            for row in sorted(set(rows_to_delete), reverse=True):
                ws.delete_rows(row)

            print("Excess rows removed successfully.")
            self.module_data["logs"]["status"].append({
                "Info": f"{PostProcessingFramework.__name__}| Excess rows removed successfully."
            })
        except Exception as e:
            self.module_data["logs"]["status"].append({
                "Error": f"{PostProcessingFramework.__name__}| Failed to delete excess rows: {e}"
            })

    def _delete_columns(self, active_worksheet, start_column_idx:int, end_column_idx:int) -> None:
        ws = active_worksheet

        amount_to_delete = end_column_idx - start_column_idx

        try:
            ws.delete_cols(start_column_idx, amount_to_delete)

            start_col_letter = get_column_letter(start_column_idx)
            end_col_letter = get_column_letter(end_column_idx)

            print(f"Columns deleted successfully: [{start_col_letter}, {end_col_letter}]")
            self.module_data["logs"]["status"].append({
                "Info": f"{PostProcessingFramework.__name__}| Columns deleted successfully: [{start_col_letter}, {end_col_letter}]"
            })
        except Exception as e:
            self.module_data["logs"]["status"].append({
                "Error": f"{PostProcessingFramework.__name__}| Failed to delete columns: {e}"
            })

    def _insert_columns(self, active_worksheet, start_column_idx:int, num_cols:int) -> None:
        ws = active_worksheet

        for _ in range(num_cols):
            ws.insert_cols(start_column_idx)

        print(f"Columns ({num_cols}) added successfully.")
        self.module_data["logs"]["status"].append({
            "Info": f"{PostProcessingFramework.__name__}| Columns ({num_cols}) added successfully."
        })

    def update_schedule_style(self, active_workbook, active_worksheet, 
                              lead_schedule_struct:str, json_struct_categories:list) -> None:
        wb = active_workbook
        ws = active_worksheet
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        point_col_idx = self._find_column_idx(
            ws, 
            lead_schedule_struct, 
            self.wbs_start_row
        )
        wbs_keys = list(self.wbs_final_categories.keys())
        lead_cat_columns = wbs_keys.index(lead_schedule_struct)
        relevant_keys = wbs_keys[:lead_cat_columns + 1]
        
        column_idxs = {
            item: self._find_column_idx(ws, item, self.wbs_start_row)
            for item in json_struct_categories
            if item in relevant_keys
        }
        start_col_idx = point_col_idx + 1
        end_col_idx = self._find_first_nonempty_string_column_idx(ws, 1, 1)

        self._delete_columns(ws, start_col_idx, end_col_idx)

        for _, value in column_idxs.items():
            self._merge_until_different_value(
                ws, value, self.wbs_start_row
            )

        self._style_file(ws, lead_schedule_struct, start_col_idx)
        
        try:
            wb.save(filename=file)
            print("WBS Table successfully updated.")
            self.module_data["logs"]["status"].append({
                "Info": f"{PostProcessingFramework.__name__}| WBS Table successfully updated."
            })
        except Exception as e:
            print(f"Failed to update WBS Table: {e}")
            self.module_data["logs"]["status"].append({
                "Error": f"{PostProcessingFramework.__name__}| Failed to update WBS Table: {e}"
            })
        finally:
            wb.close()

    def _merge_until_different_value(self, active_worksheet, starting_col_idx:int, 
                                     starting_row_idx:int) -> None:
        ws = active_worksheet

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        current_merge_range = []
        merge_ranges = []

        for row_idx in range(starting_row_idx, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=starting_col_idx)

            if cell.value is not None:
                if current_merge_range:
                    merge_ranges.append(current_merge_range)
                current_merge_range = [cell]
            else:
                current_merge_range.append(cell)

        if current_merge_range:
            merge_ranges.append(current_merge_range)

        for merge_range in merge_ranges:
            if len(merge_range) > 1:
                start_row = merge_range[0].row
                end_row = merge_range[-1].row

                ws.merge_cells(
                    start_row=start_row,
                    start_column=starting_col_idx,
                    end_row=end_row,
                    end_column=starting_col_idx
                )

                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=starting_col_idx)
                    cell.border = thin_border

    def _style_file(self, active_worksheet, lead_schedule_struct:str, 
                    start_col:int, start_row:int=1) -> None:
        ws = active_worksheet
        
        starting_day_of_the_week = ws[get_column_letter(start_col)+str(self.wbs_start_row)].value

        year_list, year_row = self._same_week_based_values(
            ws, 
            start_col, 
            start_row, 
            starting_day_of_the_week
        )
        for nested_list in year_list:
            self._merge_same_value_cells(
                ws, 
                list(nested_list.keys()), 
                list(nested_list.values()), 
                year_row
            ) 

        month_list, month_row = self._same_week_based_values(
            ws, 
            start_col, 
            start_row + 1, 
            starting_day_of_the_week
        )
        for nested_list in month_list:
            self._merge_same_value_cells(
                ws, 
                list(nested_list.keys()), 
                list(nested_list.values()), 
                month_row
            )  

        self._apply_color_format(
            ws,
            start_col, 
            start_row
        )
        self._apply_post_styling(ws)
        
        wbs_keys = list(self.wbs_final_categories.keys())
        lead_cat_columns = wbs_keys.index(lead_schedule_struct)
        relevant_keys = wbs_keys[:lead_cat_columns + 1]

        for categroy in reversed(relevant_keys):
            cat_value = self.wbs_final_categories.get(categroy)
            self._apply_horizontal_sectioning(
                ws, 
                categroy, 
                cat_value, 
                start_col
            )

    def _same_week_based_values(self, active_worksheet, starting_col_idx:int, 
                                starting_row_idx:int, starting_day:str) -> tuple[list, int]:
        ws = active_worksheet
        iterable_row = ws[starting_row_idx]

        starting_week_day = self.calendar_weekdays.index(starting_day)
        day_count = 1 + starting_week_day
        same_week_list = {}
        overall_list = []

        for cell in iterable_row[starting_col_idx-1:]:
            if cell.value is None:
                break

            if day_count > len(self.input_workweek):
                overall_list.append(same_week_list)
                same_week_list = {}
                day_count = 1
            
            same_week_list[cell.column] = cell.value
            day_count += 1

        if same_week_list:
            overall_list.append(same_week_list)

        return overall_list, starting_row_idx

    def _merge_same_value_cells(self, active_worksheet, columns_list:list, 
                                values_list:list, starting_row_idx:int) -> None:
        ws = active_worksheet

        if not columns_list:
            print("No columns to merge.")
            return

        last_value = values_list[0]
        overall_list = []
        same_val_list = []

        for col, val in zip(columns_list, values_list):
            if val == last_value:
                same_val_list.append(col)
            else:
                overall_list.append(same_val_list)
                same_val_list = [col]
                last_value = val

        if same_val_list:
            overall_list.append(same_val_list)

        for group in overall_list:
            if len(group) > 1:
                first_col = group[0]
                last_col = group[-1]
                ws.merge_cells(
                    start_row=starting_row_idx,
                    start_column=first_col,
                    end_row=starting_row_idx,
                    end_column=last_col
                )

    def _apply_color_format(self, active_worksheet, start_col:int, start_row:int) -> None:
        ws = active_worksheet

        if self.time_scale == 'w':
            self.duration_processed = len(self._get_week_range(self.start_date, self.finish_date))
    
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row, 
            self.duration_processed, 
            "00C0C0C0", 
            "00F2F2F2"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 1, 
            self.duration_processed, 
            "00C0C0C0", 
            "00D9D9D9"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 2, 
            self.duration_processed, 
            "00C0C0C0", 
            "00FFFFFF"
        )
        self._paint_schedule_row(
            ws, 
            start_col, 
            start_row + 3, 
            self.duration_processed, 
            "00C0C0C0", 
            "00FFFFFF"
        )
    
        self._apply_vertical_sectioning(
            ws, 
            start_col, 
            start_row + 1, 
            self.duration_processed
        )

        print("Workbook styled and saved successfully.")
        self.module_data["logs"]["status"].append({
            "Info": f"{PostProcessingFramework.__name__}| Workbook styled and saved successfully."
        })

    def _paint_schedule_row(self, active_worksheet, start_col:str, start_row:int, 
                            duration:int, border_color:str, fill_color:str) -> None:
        ws = active_worksheet
        for row in ws.iter_rows(min_row=start_row, 
                                 max_row=start_row, 
                                 min_col=start_col, 
                                 max_col=start_col + duration - 1):
            for cell in row:  
                cell.border = Border(
                    top=Side(border_style="thin", color=border_color), 
                    left=Side(border_style="thin", color=border_color),  
                    right=Side(border_style="thin", color=border_color),  
                    bottom=Side(border_style="thin", color=border_color)  
                )
                cell.font = Font(
                    name='Century Gothic', 
                    size=12, 
                    bold=False, 
                    color="00000000"
                )  
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    def _apply_vertical_sectioning(self, active_worksheet, start_col:str, 
                                   start_row:int, duration:int) -> None:
        ws = active_worksheet
        last_header = ws[get_column_letter(start_col)+str(start_row)].value

        for col in range(start_col, start_col + duration):
            col_header = ws.cell(row=start_row, column=col).value

            if col_header and col_header != last_header:
                for row in range(start_row + 1, ws.max_row + 1):  
                    cell = ws.cell(row=row, column=col)

                    current_border = cell.border if cell.border else Border()
                    cell.border = Border(
                        top=current_border.top, 
                        left=Side(border_style="dashed"),  
                        right=current_border.right,  
                        bottom=current_border.bottom  
                    )

                last_header = col_header

    def _apply_post_styling(self, active_worksheet) -> None:
        ws = active_worksheet
        first_row = ws[self.wbs_start_row]
        
        if first_row is None:
            print("Error: The first row is empty.")
            return
        
        for cell in first_row:  
            cell.font = Font(
                name='Century Gothic', 
                size=12, 
                bold=False, 
                color="FFFFFF"
            )  
            cell.fill = PatternFill(start_color="00800080", end_color="00800080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")  
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)  
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.rows:    
            if isinstance(row[0].value, int):
                parent_row = row
                for cell in parent_row:
                    cell.font = Font(
                        name='Century Gothic', 
                        size=12, 
                        bold=False, 
                        color="00333333"
                    )  
                    cell.fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(bottom=thin)

        print(f"Post-styling successfully completed.")
        self.module_data["logs"]["status"].append({
            "Info": f"{PostProcessingFramework.__name__}| Post-styling successfully completed."
        })
        
    def _apply_horizontal_sectioning(self, active_worksheet, section_header:str, 
                               border_style:str, start_col:int) -> None:
        ws = active_worksheet
        header_idx = self._find_column_idx(ws, section_header, self.wbs_start_row)

        section_list = self._list_cell_values(ws, header_idx)
        last_valid_section = section_list[0]

        for row_idx, row in enumerate(ws.iter_rows(min_col=start_col, 
                                                   max_col=ws.max_column, 
                                                   min_row=self.wbs_start_row + 1, 
                                                   max_row=ws.max_row)):
            current_section = section_list[row_idx]

            if current_section and current_section != last_valid_section:
                if current_section is None:
                    continue
                else:
                    self._style_row_border(row, border_style)
                    last_valid_section = current_section

        print("Post sectioning applied successfully.")
        self.module_data["logs"]["status"].append({
            "Info": f"{PostProcessingFramework.__name__}| Post sectioning applied successfully."
        })

    def _list_cell_values(self, active_worksheet, col_idx:int) -> list:
        ws = active_worksheet
        col_list = []

        if col_idx is not None:
            for row in ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                    min_col=col_idx, 
                                    max_col=col_idx, 
                                    max_row=ws.max_row):
                for cell in row:
                    col_list.append(cell.value)
        
        return col_list 

    def _style_row_border(self, row:list, border_style:str) -> None:
        if border_style == "no_border":
            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=current_border.top, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border
        else:
            new_top_border = Side(border_style=border_style, color='000000')

            for cell in row:
                current_border = cell.border if cell.border else Border()
                new_border = Border(
                    top=new_top_border, 
                    left=current_border.left,  
                    right=current_border.right,  
                    bottom=current_border.bottom  
                )

                cell.border = new_border


if __name__ == "__main__":
    PostProcessingFramework.main(False)
