import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from backend.config.paths import RSLTS_DIR

class WbsFramework:
    def __init__(self, is_framed, input_file_path, input_file_basename, input_file_extension, 
                 input_worksheet_name, project_table, project_ordered_dict):
        self.is_framed = is_framed
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.worksheet_name = input_worksheet_name
        self.table = project_table
        self.ordered_dict = project_ordered_dict

        #Module Attributes
        self.wbs_index_col = 1
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.default_hex_font_color = "00FFFFFF"
        self.default_hex_fill_color = "00FFFF00"

        #Structures
        self.project_content_headers = {
            "color": ["color"],
            "activity_code": ["activity_code", "activitycode", "code", "task_code", "act_code"],
            "wbs_code": ["wbs_code", "wbscode"],
            "activity_name": ["activity_name", "activityname", "act_name"], 
        }
    
    @staticmethod
    def main(auto=True, is_framed=False, input_file_path=None, input_file_basename=None, input_file_extension=None, 
             input_worksheet_name=None, project_table=None, project_ordered_dict=None):
        if auto:
            project = WbsFramework.auto_generate_ins(
                is_framed,
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_worksheet_name, 
                project_table, 
                project_ordered_dict,
            )
        else:
            project = WbsFramework.generate_ins()

        if project:
            color_list = [
                project.process_hex_val(item["color"]) for item in project.ordered_dict
            ]

            project.create_wbs_table(project.table)
            for header in project.project_content_headers.get("activity_code"):
                try:
                    project.process_wbs_column(header, color_list)
                except:
                    continue

            project.process_wbs_column('color', color_list)

    @staticmethod
    def generate_ins():
        WbsFramework.ynq_user_interaction(
            "Run module as stand-alone? "
        )

        setup_cls = setup.Setup.main()

        worksheet_name = input("Please enter the worksheet name: ")
        is_framed = WbsFramework.binary_user_interaction("Would you like the table to be framed [Y/N]? ")
        
        ins = WbsFramework(
            is_framed,
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            worksheet_name,
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("table"),
            setup_cls.obj["project"]["modules"]["MODULE_3"].get("content"),
        )

        return ins

    @staticmethod
    def auto_generate_ins(is_framed, input_file_path, input_file_basename, input_file_extension,
                          input_worksheet_name, project_table, project_ordered_dict):
        ins = WbsFramework(
            is_framed,
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_worksheet_name, 
            project_table,
            project_ordered_dict
        )
        
        return ins

    @staticmethod
    def ynq_user_interaction(prompt_message:str) -> str:
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def return_valid_file(input_file_dir:str, mode:str='r') -> dict|int:
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = WbsFramework._handle_dir(input_file_dir, mode)
        elif os.path.isfile(input_file_dir):
            file_dict = WbsFramework._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str='r') -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = WbsFramework._display_directory_files(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = input_file_dir, 
            basename = os.path.basename(input_file_basename).split(".")[0] if input_file_basename else "",
            extension = os.path.basename(input_file_basename).split(".")[-1] if input_file_basename else "",
        )
    
    @staticmethod
    def _display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def _handle_file(input_file_dir:str) -> dict:
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in ["json"]):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return {}

    @staticmethod
    def clear_directory(directory_path:str) -> None:
        if os.path.isdir(directory_path):
            file_list = os.listdir(directory_path)
            for file in file_list:
                file_path = os.path.join(directory_path, file)
                os.remove(file_path)

    @staticmethod
    def binary_user_interaction(prompt_message:str) -> bool:
        valid_responses = {'y', 'n'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                if user_input == 'y':
                    return True 
                else:
                    return False 
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No]\n")

    def process_hex_val(self, hex_val:str) -> str:
        return hex_val.replace('#', "00")

    def create_wbs_table(self, proc_table) -> None:
        if self.is_framed:
            wbs_table = proc_table
        else:
            wbs_table = proc_table.reset_index()
        
        self._write_data_to_excel(wbs_table, self.input_path, self.input_basename)
        
        if not self.is_framed:
            self._remove_index_column()
        
    def _write_data_to_excel(self, proc_table, excel_path:str, 
                            excel_basename:str) -> None:
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            basename = excel_basename + '.xlsx'
            file = os.path.join(excel_path, basename)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='a') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name=self.worksheet_name, 
                        startrow=self.wbs_start_row - 1, 
                        startcol=column_index_from_string(self.wbs_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {self.worksheet_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")

    def _remove_index_column(self) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            wb = load_workbook(filename=file)
            ws = wb[self.worksheet_name]

            ws.delete_cols(self.wbs_index_col)
            
            print("Successfully removed first column")
            wb.save(file)
        except:
            pass

    def process_wbs_column(self, col_header:str, color_list:list) -> None:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        wb = load_workbook(file)
        ws = wb[self.worksheet_name]

        header_idx = self._find_column_idx(ws, col_header)
        self._fill_color_col(ws, header_idx, color_list)

        wb.save(file)

    def _find_column_idx(self, active_worksheet, column_header:str) -> int|None:
        ws = active_worksheet
        start_col_idx = column_index_from_string(self.wbs_start_col)

        if ' ' in column_header:
            normalized_header = column_header.replace(" ", "_").lower()
        else:
            normalized_header = column_header.lower()

        for row in ws.iter_rows(min_row=self.wbs_start_row, min_col=start_col_idx, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.replace(" ", "_").lower()
                    if normalized_header in normalized_cell_value:
                        return cell.column

    def _fill_color_col(self, active_worksheet, col_idx:int, col_list:list) -> None:
        ws = active_worksheet

        if not col_list:
            print("Error: Color list is empty.")
            return

        for idx, row in enumerate(ws.iter_rows(min_row=self.wbs_start_row + 1, 
                                                max_row=ws.max_row, 
                                                min_col=col_idx,
                                                max_col=col_idx)):
            for cell in row:
                color = col_list[idx]
                try:
                    cell.fill = PatternFill(start_color=color, 
                                            end_color=color, 
                                            fill_type="solid")
                except Exception as e:
                    print(f"Color hex not found: {color}. Error: {e}")
                    cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                            end_color=self.default_hex_fill_color, 
                                            fill_type="solid")

        column_letter = get_column_letter(col_idx)
        print(f"Column ({column_letter}) styled successfully")


if __name__ == "__main__":
    WbsFramework.main(False)
