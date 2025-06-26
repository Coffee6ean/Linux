import os
import re
import json
import pandas as pd
from dateutil import parser
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime, date
from openpyxl.utils import get_column_letter, column_index_from_string

class DataIngestion:
    def __init__(self, input_file_path, input_file_basename, 
                 input_file_extension, input_file_roi, output_file_dir):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.output_path = output_file_dir
        self.xlsx_start_col = 'A'
        self.xlsx_start_row = 1

        #Module attributes
        self.ws_name = input_file_roi

        #Structures
        self.project_content_headers = {
            "entry": None, 
            "phase": ["phase", "phases"],
            "area": ["area", "areas"], 
            "zone": ["zone", "zones"],
            "subzone": ["subzone", "sub_zone"],
            "level": ["level", "levels"],
            "trade": ["trade", "trades"], 
            "color": ["color", "colour", "colors", "colours"],
            "parent_id": ["parent_id"], 
            "activity_code": ["activity_code", "activitycode", "code", "task_code", "act_code"],
            "wbs_code": ["wbs_code", "wbscode"],
            "activity_name": ["activity_name", "activityname", "act_name", "task_name"], 
            "activity_category": ["activity_category", "activitycategory", "category"], 
            "activity_status": ["activity_status", "activitystatus", "status", "task_status"], 
            "activity_ins": None, 
            "start": ["start", "start_date", "start_dates"], 
            "finish": ["finish", "finish_date", "finish_dates", "end", "end_date"], 
            "total_float": ["total_float", "totalfloat"],
            "activity_successor_id": ["successor", "successor"],
            "activity_predecessor_id": ["predecessor", "predecessors"],
        }
        self.json_struct_categories = ["phase", "area", "zone", "subzone", "level", "trade", "activity_code"]

        #Module Results
        self.module_data = {
            "details": {
                "workbook": None,
                "worksheet": None,
                "start_date": None,
                "finish_date": None,
                "entry_count": 0,
            },
            "logs": {
                "start": DataIngestion.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            },
            "content": {}
        }

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, 
             input_file_extension=None, input_file_roi=None, output_file_dir=None):
        if auto:
            project = DataIngestion.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_file_roi, 
                output_file_dir
            )
        else:
            project = DataIngestion.generate_ins()

        if project:
            if project.input_extension == "xlsx":
                if project.ws_name is None:
                    worksheet = input("Please enter the name for the new or existing worksheet: ")
                else:
                    worksheet = project.ws_name
                
                xlsx_results = project.handle_xlsx(worksheet, auto)
                basename = project.input_basename + '.' + project.input_extension 
                project.module_data["details"]["workbook"] = os.path.join(project.input_path, basename)
                project.module_data["details"]["worksheet"] = worksheet
                project.module_data["details"]["start_date"] = xlsx_results.get("earliest_start")
                project.module_data["details"]["finish_date"] = xlsx_results.get("latest_finish")
                project.module_data["details"]["entry_count"] = xlsx_results.get("entry_count")
                project.module_data["content"] = xlsx_results.get("nested_json")
            elif project.input_extension == "xml":
                excel_path, excel_basename = DataIngestion.return_valid_file(project.output_file_dir)
                processed_json = project.handle_xml()
                project.module_data["content"] = project._create_wbs_table_to_fill(processed_json, excel_path, excel_basename)
            else:
                print("Error. Unsuported file type")
                project.module_data["logs"]["status"].append(dict(
                    Error=f"{DataIngestion.__name__}| Unsuported input file extension: {project.input_extension}"
                ))
        else:
            project.module_data["logs"]["status"].append(dict(
                Error= f"{DataIngestion.__name__}| Module's instance was not generated correctly"
            ))
            
        project.module_data["logs"]["finish"] = DataIngestion.return_valid_date()
        project.module_data["logs"]["run-time"] = DataIngestion.calculate_time_duration(
            project.module_data["logs"].get("start"), 
            project.module_data["logs"].get("finish")
        )
        project.module_data["logs"]["status"].append(dict(
            Info=f"{DataIngestion.__name__}| Module ran successfully"
        ))

        return project.module_data

    @staticmethod
    def generate_ins():
        input_file = DataIngestion.return_valid_file(
            input("Please enter the path to the file or directory: ").strip()
        )
        output_file_dir = DataIngestion.return_valid_path(
            "Please enter the directory to save the new module results: "
        )

        ins = DataIngestion(
            input_file.get("path"), 
            input_file.get("basename"), 
            input_file.get("extension"), 
            output_file_dir
        )

        return ins
        
    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, 
                          input_file_extension, input_file_roi, output_file_dir):
        ins = DataIngestion(
            input_file_path, 
            input_file_basename, 
            input_file_extension, 
            input_file_roi,
            output_file_dir
        )
        
        return ins
    
    @staticmethod
    def return_valid_file(input_file_dir:str) -> dict|int:        
        if not os.path.exists(input_file_dir):
            raise FileNotFoundError("Error: Given directory or file does not exist in the system.")

        if os.path.isdir(input_file_dir):
            file_dict = DataIngestion._handle_dir(input_file_dir)
        elif os.path.isfile(input_file_dir):
            file_dict = DataIngestion._handle_file(input_file_dir)

        return file_dict
    
    @staticmethod
    def _handle_dir(input_file_dir:str, mode:str="r") -> dict|int:
        if mode in ['u', 'r', 'd']:
            dir_list = os.listdir(input_file_dir)
            selection = DataIngestion._display_directory_files(dir_list)
            input_file_basename = dir_list[selection]
            print(f'File selected: {input_file_basename}\n')
        elif mode == 'c':
            input_file_basename = None
        else:
            print("Error: Invalid mode specified.")
            return -1
        
        return dict(
            path = os.path.dirname(input_file_dir), 
            basename = os.path.basename(input_file_basename).split(".")[0],
            extension = os.path.basename(input_file_basename).split(".")[-1],
        )
    
    @staticmethod
    def _display_directory_files(file_list:list) -> int:
        selection_idx = -1  

        if len(file_list) == 0:
            print('Error. No files found')
            return -1
        
        print(f'-- {len(file_list)} files found:')
        for idx, file in enumerate(file_list, start=1):  
            print(f'{idx}. {file}')

        while True:
            try:
                selection_idx = int(input('\nPlease enter the index number to select the one to process: '))
                if 1 <= selection_idx <= len(file_list):  
                    return selection_idx - 1  
                else:
                    print(f'Error: Please enter a number between 1 and {len(file_list)}.')
            except ValueError:
                print('Error: Invalid input. Please enter a valid number.\n')

    @staticmethod
    def _handle_file(input_file_dir:str):
        input_file_extension = os.path.basename(input_file_dir).split(".")[-1]

        if (input_file_extension in ["xlsx", "xml"]):
            return dict(
                path = os.path.dirname(input_file_dir), 
                basename = os.path.basename(input_file_dir).split(".")[0],
                extension = input_file_extension,
            )

        print("Error: Please verify that the directory and file exist and that the file extension complies with class attributes")
        return -1

    @staticmethod
    def return_valid_path(prompt_message:str) -> (str|None):
        while(True):   
            value = input(prompt_message).strip()
            try:
                if os.path.isdir(value):
                    return value
            except Exception as e:
                print(f"Error. {e}\n")

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def normalize_string(entry_str:str) -> str:
        remove_bewteen_parenthesis = re.sub('(?<=\\()(.*?)(?=\\))', '', entry_str)
        special_chars = re.compile('[@_!#$%^&*()<>?/\\|}{~:]')
        remove_special_chars = re.sub(special_chars, '', remove_bewteen_parenthesis.lower()).strip()
        normalized_str = re.sub(' ', '_', remove_special_chars)

        return normalized_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    def handle_xlsx(self, ws_name:str, auto:bool) -> tuple[dict, dict]:
        try:
            if auto:
                _, ws = self._auto_return_excel_workspace(ws_name)
            else:
                _, ws = self._return_excel_workspace(ws_name)

            file_headers = self._xlsx_return_header(ws)
            json_header = self._json_fill_header(file_headers)
            json_obj, entry_counter = self._json_fill_body(ws, json_header)
            reworked_json = self._fill_missing_dates(json_obj)
            earliest_start, latest_finish = self._project_dates(reworked_json)
            nested_json = self._build_nested_dic(reworked_json)
            
            xlsx_results = {
                "nested_json": nested_json,
                "flattend_json": reworked_json,
                "entry_count": entry_counter,
                "earliest_start": earliest_start,
                "latest_finish": latest_finish,
            }
        except Exception as e:
            err_msg = f"Failed to handle XLSX processing: {e}."
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error= f"{DataIngestion.__name__}|{self.handle_xlsx.__name__}| Failed to handle XLSX processing: {e}."
            ))

            xlsx_results = {
                "nested_json": {},
                "flattend_json": [],
                "entry_count": 0,
                "earliest_start": None,
                "latest_finish": None,
            }
        finally:
            return xlsx_results
    
    def _auto_return_excel_workspace(self, worksheet_name:str) -> tuple:
        try:
            basename = self.input_basename + '.' + self.input_extension
            file = os.path.join(self.input_path, basename)

            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]

            return workbook, worksheet
            
        except Exception as e:
            err_msg = f"Failed to return Excel workspace: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._return_excel_workspace.__name__}| {err_msg}"
            ))
            return None, None

    def _return_excel_workspace(self, worksheet_name:str) -> tuple:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    self.ws_name = worksheet_name
                    print(f"New worksheet '{self.ws_name}' created.\n")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws_idx = DataIngestion._display_directory_files(ws_list)
                    
                    if selected_ws_idx >= 0:  
                        worksheet = workbook.worksheets[selected_ws_idx]
                        self.ws_name = ws_list[selected_ws_idx]
                        print(f"Worksheet selected: '{self.ws_name}'\n")
                        return workbook, worksheet
                    else:
                        print("Invalid selection. Returning without changes.\n")
                        return workbook, None
                        
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def _xlsx_return_header(self, active_worksheet) -> list:
        ws = active_worksheet

        start_col = column_index_from_string(self.xlsx_start_col)
        header_list = []

        for row in ws.iter_rows(min_row=self.xlsx_start_row, max_row=self.xlsx_start_row, 
                                min_col=start_col, max_col= ws.max_column):
            for cell in row:
                if cell.value is not None:
                    normalized_str = self.normalize_string(cell.value)
                    header_tuple = (cell.coordinate, self._verify_header(normalized_str))
                    header_list.append(header_tuple)

        return header_list
    
    def _verify_header(self, entry_str:str) -> str:
        flat_allowed_headers = {
            value: key for key, values in self.project_content_headers.items() if values is not None for value in values
        }

        if entry_str in flat_allowed_headers:
            result = flat_allowed_headers[entry_str]
        else:
            print(
                f"Warning. Header '{entry_str}' not found in declared dictionary"
            )
            result = entry_str
        
        return result

    def _json_fill_header(self, file_headers:list) -> dict:
        try:
            json_obj = {
                "start": None,
                "finish": None,
                "header": {key: None for key, _ in self.project_content_headers.items()},
                "body": [],
            }

            for coordinates, data in file_headers:
                header_dict = json_obj["header"]

                header_dict[data] = coordinates

            return json_obj
        
        except Exception as e:
            err_msg = f"Failed to build JSON header: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._json_fill_header.__name__}| {err_msg}"
            ))
            return {
                "start": None,
                "finish": None,
                "header": {},
                "body": [],
            }

    def _json_fill_body(self, active_worksheet, json_obj:dict) -> tuple[dict, int]:
        try:
            ws = active_worksheet
            body_dict = json_obj["body"]
            header_dict = json_obj["header"]

            normalized_header = sorted(
                {k: v for k, v in header_dict.items() if v is not None}.items(), 
                key=lambda item: item[1]
            )

            header_coordinates_list = [dict_tuple[1] for dict_tuple in normalized_header]
            header_key_list = [dict_tuple[0] for dict_tuple in normalized_header]

            first_header_col_letter, first_header_row = self._get_column_coordinates(
                header_dict, normalized_header[0][0]
            )
            last_header_col_letter, _ = self._get_column_coordinates(header_dict, normalized_header[-1][0])
            first_header_col = column_index_from_string(first_header_col_letter)
            last_header_col = column_index_from_string(last_header_col_letter)

            entry_counter = 1
            for row in ws.iter_rows(min_row=first_header_row + 1, max_row=ws.max_row, 
                                    min_col=first_header_col, max_col=last_header_col):
                json_activity = {key: None for key in header_dict.keys()}
                json_activity["entry"] = entry_counter
                for cell in row:
                    parent_header = next((val for val in header_coordinates_list if get_column_letter(cell.column) in val), None)
                    if cell.value:        
                        position = header_coordinates_list.index(parent_header)
                        key = header_key_list[position]
                        json_activity[key] = self._normalize_data_value(key, cell.value)
                    else:
                        position = header_coordinates_list.index(parent_header)
                        key = header_key_list[position]

                        if key == "area":
                            json_activity[key] = "N/A"
                        else:
                            json_activity[key] = ""

                body_dict.append(json_activity)
                entry_counter += 1

            return json_obj, entry_counter
        except Exception as e:
            err_msg = f"Failed to fill JSON body: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._json_fill_body.__name__}| {err_msg}"
            ))
            return json_obj, 0
    
    def _get_column_coordinates(self, header_dict:dict, header_key:str) -> tuple[str, int]:
        coordinates = header_dict[header_key]

        try:
            row_match = re.search(r'\d+', coordinates)
            column_match = re.search(r'[a-zA-Z]+', coordinates)

            if not row_match or not column_match:
                raise ValueError(f"Invalid coordinates format: {coordinates}")

            row = int(row_match.group(0))
            column = column_match.group(0)

            return column, row 
        except Exception as e:
            err_msg = f"Failed to extract coordinates for '{header_key}': {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._get_column_coordinates.__name__}| {err_msg}"
            ))
            return "", 0

    def _normalize_data_value(self, header_column:str, cell_value) -> str|None:
        if isinstance(cell_value, str):
            if header_column == 'start' or header_column == 'finish':
                result = self._format_date_string(cell_value)
            else:
                result = cell_value.strip()

            return result
        elif isinstance(cell_value, date):
            return cell_value.strftime('%d-%b-%Y')
        elif isinstance(cell_value, int) or isinstance(cell_value, float):
            if cell_value == 0:
                cell_value = "0"
            return str(cell_value)
        else:
            return None
        
    def _format_date_string(self, date_string:str, output_format="%d-%b-%Y") -> str|None:
        try:
            special_chars = re.compile('[@_!#$%^&*()<>?\\|}{~:]')

            if re.search(special_chars, date_string):
                cleaned_date_string = re.sub(special_chars, '', date_string)
            else:
                cleaned_date_string = re.sub("[a-zA-Z]$", '', date_string)
            

            parsed_date = parser.parse(cleaned_date_string.strip())
            formatted_date = parsed_date.strftime(output_format)

            return formatted_date
        
        except Exception as e:
            err_msg = f"Could not parse date string '{date_string}': {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._format_date_string.__name__}| {err_msg}"
            ))
            return None

    def _fill_missing_dates(self, json_obj:dict) -> dict:
        try:
            original_body = json_obj.get("body", [])
            cleaned_body = []

            for item in original_body:
                start = item.get("start")
                finish = item.get("finish")

                if not start and not finish:
                    msg = f"Skipping entry with no 'start' or 'finish': entry {item.get('entry')}"
                    print(f"[INFO] {msg}")
                    self.module_data["logs"]["status"].append(dict(
                        Info=f"{DataIngestion.__name__}|{self._fill_missing_dates.__name__}| {msg}"
                    ))

                    continue

                if not start:
                    item["start"] = finish
                if not finish:
                    item["finish"] = start

                cleaned_body.append(item)

            json_obj["body"] = cleaned_body

            return json_obj
        
        except Exception as e:
            err_msg = f"Failed to fill missing dates: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._fill_missing_dates.__name__}| {err_msg}"
            ))

            return json_obj
    
    def _project_dates(self, json_obj:dict) -> tuple:
        try:
            body_dict = json_obj["body"]
            
            start_list = [date["start"] for date in body_dict]
            finish_list = [date["finish"] for date in body_dict]

            earliest_start = self._bubble_sort_dates(start_list)[0]        
            latest_finish = self._bubble_sort_dates(finish_list)[-1]        

            return earliest_start, latest_finish
        
        except Exception as e:
            err_msg = f"Failed to compute project dates: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._project_dates.__name__}| {err_msg}"
            ))

            return None, None

    def _bubble_sort_dates(self, unsorted_list:list) -> list:
        try:
            n = len(unsorted_list)

            for i in range(n):
                for j in range(n-i-1):
                    value_j = unsorted_list[j]
                    value_j1 = unsorted_list[j+1]

                    try:
                        if isinstance(value_j, str) and isinstance(value_j1, str):
                            value_j = datetime.strptime(value_j, "%d-%b-%Y").date()
                            value_j1 = datetime.strptime(value_j1, "%d-%b-%Y").date()

                        elif isinstance(value_j1, datetime):
                            value_j = value_j.date()
                            value_j1 = value_j1.date()

                        if value_j > value_j1:
                            unsorted_list[j], unsorted_list[j+1] = unsorted_list[j+1], unsorted_list[j]

                    except Exception as e:
                        err_msg = f"Failed to parse date during sort: {e}"
                        print(f"[WARNING] {err_msg}")
                        self.module_data["logs"]["status"].append(dict(
                            Warning=f"{DataIngestion.__name__}|{self._bubble_sort_dates.__name__}| {err_msg}"
                        ))
            
            return unsorted_list
        
        except Exception as e:
            err_msg = f"Failed to sort date list: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._bubble_sort_dates.__name__}| {err_msg}"
            ))
            return unsorted_list

    def _build_nested_dic(self, json_obj:dict) -> dict:
        header_dict = json_obj["header"]
        body_dict = json_obj["body"]

        nested_dict = {
            "header": header_dict,
            "body": {},
        }

        def dict_management(json_obj:dict, key_list:list, recursive_dict:dict):
            if not key_list:
                return recursive_dict

            current_key = key_list.pop(0)
            obj_key = json_obj.get(current_key)

            if obj_key is not None and obj_key != "":
                if len(key_list) < 1:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = []
                    recursive_dict[obj_key].append(json_obj)
                else:
                    if obj_key not in recursive_dict:
                        recursive_dict[obj_key] = {}
                    dict_management(json_obj, key_list, recursive_dict[obj_key])
            else:
                if "M_DATA" not in recursive_dict:
                    recursive_dict["M_DATA"] = []
                recursive_dict["M_DATA"].append(json_obj)

            return recursive_dict

        try:
            for obj in body_dict:
                keys = [category for category in self.json_struct_categories if obj.get(category) is not None]
                nested_dict["body"] = dict_management(
                    obj, keys, nested_dict["body"]
                )
            
            return nested_dict
    
        except Exception as e:
            err_msg = f"Failed to build nested JSON structure: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self._build_nested_dic.__name__}| {err_msg}"
            ))

            return {"header": {}, "body": {}}

    def write_json(self, json_obj:dict, processed_json:bool=False) -> None:
        if processed_json:
            basename = f"processed_{self.output_json_basename}.json"
        else:
            basename = f"{self.output_json_basename}.json"
            
        file_dir = os.path.join(self.output_path, basename)

        try:
            with open(file_dir, 'w') as file_writer:
                json.dump(json_obj, file_writer)

            print(f"JSON data successfully created and saved to {basename}.")
        
        except Exception as e:
            err_msg = f"Failed to write JSON to file: {e}"
            print(f"[ERROR] {err_msg}")
            self.module_data["logs"]["status"].append(dict(
                Error=f"{DataIngestion.__name__}|{self.write_json.__name__}| {err_msg}"
            ))

    def handle_xml(self):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        tree = ET.parse(file)
        root = tree.getroot()
        ns = {'ms': 'http://schemas.microsoft.com/project'}

        xml_tasks = []
        for task in root.findall("./ms:Tasks/ms:Task", ns):
            xml_tasks.append(task)

        json_dict_list = []
        if xml_tasks:
            for task in xml_tasks:
                task_data = {child.tag.split('}')[-1]: child.text.strip() for child in task}
                json_dict_list.append(task_data)
        else:
            print("No <Task> elements found")

        return json_dict_list   
        
    def _create_wbs_table_to_fill(self, data:list, excel_path:str, 
                                 excel_basename:str):
        proc_table = self._generate_wbs_to_fill(data)
        self._write_data_to_excel(proc_table, excel_path, excel_basename)

        return proc_table

    def _generate_wbs_to_fill(self, json_obj:dict):
        df = pd.DataFrame(json_obj) 
        df.fillna("", inplace=True)

        return df

    def _write_data_to_excel(self, proc_table, excel_path:str, 
                            excel_basename:str) -> None:
        if proc_table.empty:    
            print("Error. DataFrame is empty\n")
        else:
            excel_basename = self.input_basename.split('.')[0] + '.xlsx'
            file = os.path.join(excel_path, excel_basename)
            ws_name = "CFA - Fill Table"

            if not os.path.exists(excel_path):
                os.mkdir(excel_path)

            try:
                with pd.ExcelWriter(file, engine="openpyxl", mode='w') as writer:
                    proc_table.to_excel(
                        writer, 
                        sheet_name = ws_name, 
                        startrow = self.xlsx_start_row - 1, 
                        startcol = column_index_from_string(self.xlsx_start_col) - 1
                    )
                
                print(f"Successfully converted JSON to Excel and saved to: {file}")
                print(f"Saved to sheet: {ws_name}\n")
            except Exception as e:
                print(f"An unexpected error occurred: {e}\n")


if __name__ == "__main__":
    module_data = DataIngestion.main(False)
