import os
import re
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Imported Helper - As Module
""" import setup """

import sys
sys.path.append("../")
from backend.config.paths import RSLTS_DIR

class ScheduleFramework():
    documentation_labels = ["ERROR", "INFO", "WARNING"]

    def __init__(self, input_file_path, input_file_basename, input_file_extension, input_file_workweek,
                 project_worksheet_name, project_table, project_ordered_dict, project_phase_order, project_lead_struct, 
                 project_start_date, project_finish_date, time_scale, input_start_row=1, input_start_col=""):
        self.input_path = input_file_path
        self.input_basename = input_file_basename
        self.input_extension = input_file_extension
        self.input_workweek = input_file_workweek
        self.worksheet_name = project_worksheet_name
        self.table = project_table
        self.ordered_dict = project_ordered_dict
        self.phase_order = project_phase_order
        self.lead_struct = project_lead_struct
        self.start_date = project_start_date
        self.finish_date = project_finish_date
        self.time_scale = time_scale
        self.start_row = int(input_start_row)
        self.start_col = str(input_start_col)
        
        #Module Attributes
        self.wbs_start_row = 4
        self.wbs_start_col = 'A'
        self.schedule_scale_based = []
        self.dark_default_hex_font = "00000000"
        self.light_default_hex_font = "00FFFFFF"
        self.default_hex_fill_color = "00FFFF00"

        #Structures
        self.time_scale_options = ["d", "w"]
        self.wbs_final_categories = {
            "phase": "thin",
            "area": "dashDot", 
            "zone": "dashed",
            "subzone": "dotted",
            "level": "hair",
        }
        self.allowed_headers = {
            "hyperlinks": ["activity_code", "activity_name", "wbs_code"]
        }
        self.calendar_weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        self.calendar_months = {
            1:"Jan",
            2:"Feb",
            3:"Mar",
            4:"Apr",
            5:"May",
            6:"Jun",
            7:"Jul",
            8:"Aug",
            9:"Sep",
            10:"Oct",
            11:"Nov",
            12:"Dec",
        }

        #Module Results
        self.module_data = {
            "logs": {
                "start": ScheduleFramework.return_valid_date(),
                "finish": None,
                "run-time": None,
                "status": [],
            }
        }

    @staticmethod
    def main(auto=True, input_file_path=None, input_file_basename=None, input_file_extension=None, input_file_workweek=None, 
             project_worksheet_name=None, project_table=None, project_ordered_dict=None, project_phase_order=None, 
             project_lead_struct=None, project_start_date=None, project_finish_date=None, time_scale=None):
        if auto:
            project = ScheduleFramework.auto_generate_ins(
                input_file_path, 
                input_file_basename, 
                input_file_extension,
                input_file_workweek, 
                project_worksheet_name,
                project_table, 
                project_ordered_dict, 
                project_phase_order, 
                project_lead_struct, 
                project_start_date, 
                project_finish_date,
                time_scale
            )
        else:
            project = ScheduleFramework.generate_ins()

        if project:
            if not project.worksheet_name:
                project.worksheet_name = "CFA - Schedule"

            if not project.time_scale:
                print("Please choose the time scale to process the schedule: ")
                scale_idx = ScheduleFramework.display_options(project.time_scale_options)
                project.time_scale = project.time_scale_options[scale_idx[0] - 1]

            active_workbook, active_worksheet = project.return_excel_workspace(project.worksheet_name)

            if active_workbook and active_worksheet:
                project.create_schedule(
                    active_workbook, 
                    active_worksheet, 
                    project.table, 
                    project.ordered_dict
                )
            else:
                project.module_data["logs"]["status"].append(dict(
                    Error= f"{ScheduleFramework.__name__}| Could not open Excel file as Workbook & Worksheet"
                ))
        else:
            project.module_data["logs"]["status"].append(dict(
                Error= f"{ScheduleFramework.__name__}| Module's instance was not generated correctly"
            ))

        project.module_data["logs"]["finish"] = ScheduleFramework.return_valid_date()
        project.module_data["logs"]["run-time"] = ScheduleFramework.calculate_time_duration(
            project.module_data["logs"].get("start"), 
            project.module_data["logs"].get("finish")
        )
        project.module_data["logs"]["status"].append(dict(
            Info=f"{ScheduleFramework.__name__}| Module ran successfully"
        ))

        return project.module_data

    @staticmethod
    def generate_ins():
        ScheduleFramework.ynq_user_interaction(
            "Run Module as stand-alone? "
        )

        setup_cls = setup.Setup.main()
        
        ins = ScheduleFramework(
            setup_cls.obj["input_file"].get("path"), 
            setup_cls.obj["input_file"].get("basename"),
            setup_cls.obj["input_file"].get("extension"),  
            setup_cls.obj["project"]["modules"]["MODULE_3"]["details"].get("workweek"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("worksheet"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("table"),
            setup_cls.obj["project"]["modules"]["MODULE_3"].get("content"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("custom_phase_order"),
            setup_cls.obj["project"]["modules"]["MODULE_2"]["content"].get("lead_schedule_struct"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("start_date"),
            setup_cls.obj["project"]["modules"]["MODULE_1"]["details"].get("finish_date"),
            None
        )

        return ins
    
    @staticmethod
    def auto_generate_ins(input_file_path, input_file_basename, input_file_extension, input_file_workweek, 
                          project_worksheet_name, project_table, project_ordered_dict, project_phase_order, 
                          project_lead_struct, project_start_date, project_finish_date, time_scale):
        ins = ScheduleFramework(
            input_file_path, 
            input_file_basename, 
            input_file_extension,
            input_file_workweek, 
            project_worksheet_name,
            project_table, 
            project_ordered_dict, 
            project_phase_order, 
            project_lead_struct, 
            project_start_date, 
            project_finish_date,
            time_scale
        )

        return ins

    @staticmethod
    def return_valid_date() -> str:
        now = datetime.now()
        date_str = now.strftime("%d-%b-%y %H:%M:%S")

        return date_str

    @staticmethod
    def calculate_time_duration(start_date:str, finish_date:str, 
                                date_format:str="%d-%b-%y %H:%M:%S") -> float|int:
        try:
            start_time = datetime.strptime(start_date, date_format)
            finish_time = datetime.strptime(finish_date, date_format)

            minutes_duration = (finish_time - start_time).total_seconds()

            return minutes_duration
        except (ValueError, TypeError) as e:
            print(f"Error calculating runtime: {e}")
            return -1

    @staticmethod
    def ynq_user_interaction(prompt_message:str) -> str:
        valid_responses = {'y', 'n', 'q'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                return user_input  
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No, 'Q/q' for Quit]\n")

    @staticmethod
    def binary_user_interaction(prompt_message:str) -> bool:
        valid_responses = {'y', 'n'}  
        
        while True:
            user_input = input(prompt_message).lower().strip()
            
            if user_input in valid_responses:
                if user_input == 'y':
                    return True 
                else:
                    return False 
            else:
                print("Error. Invalid input, please try again. ['Y/y' for Yes, 'N/n' for No]\n")

    @staticmethod
    def display_directory_files(file_list:list) -> int:
        selection_idx = 0
        if len(file_list)==0:
            print('Error. No files found')
            return -1
        
        if len(file_list)>1:
            print(f'-- {len(file_list)} files found:')
            idx = 0
            for file in file_list:
                idx += 1
                print(f'{idx}. {file}')

            selection_idx = input('\nPlease enter the index number to select the one to process: ') 
        else:
            print(f'Single file found: {file_list[0]}')
            print('Will go ahead and process')

        return int(selection_idx) - 1

    @staticmethod
    def display_options(file_list:list) -> list:
        if not file_list:
            print('Error: No elements found.')
            return []

        print(f'-- {len(file_list)} elements found:')
        for idx, file in enumerate(file_list, start=1):
            print(f'{idx}. {file}')

        result = []
        selection_input = input('\nEnter index numbers (comma-separated) to select elements to process: ').split(',')

        for selection in selection_input:
            selection = selection.strip()
            if not selection.isdigit():
                print(f'Error: Invalid input "{selection}", skipping.')
                continue

            index = int(selection)
            if 1 <= index <= len(file_list):
                result.append(index)
            else:
                print(f'Error: "{index}" is out of range (1 to {len(file_list)}).')

        return result

    @staticmethod
    def hex2rgb(hex_color:str) -> tuple:
        trimmed_hex = hex_color.lstrip('#')
        calc_rgb = tuple(int(trimmed_hex[i:i+2], 16) for i in (0, 2, 4))

        return calc_rgb

    @staticmethod
    def calculateLuminance(R:float, G:float, B:float) -> float:
        lum = 0.2126*(R/255.0)**2.2 + 0.7152*(G/255.0)**2.2 + 0.0722*(B/255.0)**2.2

        return lum

    @staticmethod
    def write_data_to_json(file_title:str, json_dict:dict) -> None:
        file = os.path.join(RSLTS_DIR, file_title)
        
        with open(file, 'w') as writer:
            json.dump(json_dict, writer)

    def return_excel_workspace(self, worksheet_name:str):
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)
        
        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[worksheet_name]
        except KeyError:
            print(f"Error: Worksheet '{worksheet_name}' does not exist.")
            
            while True:
                user_answer = input("Would you like to create a new worksheet under this name? (Y/N/Q): ").strip().lower()
                
                if user_answer == 'y':
                    worksheet = workbook.create_sheet(worksheet_name)
                    print(f"New worksheet '{worksheet_name}' created.")
                    break
                elif user_answer == 'n':
                    ws_list = workbook.sheetnames
                    selected_ws = ScheduleFramework.display_directory_files(ws_list)
                    worksheet = workbook.worksheets[selected_ws]
                    return workbook, worksheet
                elif user_answer == 'q':
                    print("Quitting without changes.")
                    return workbook, None
                else:
                    print("Invalid input. Please enter 'Y' for Yes, 'N' for No, or 'Q' to Quit.")
        
        return workbook, worksheet

    def create_schedule(self, active_workbook, active_worksheet, 
                        proc_table, json_dict:list) -> dict:
        basename = self.input_basename + '.' + self.input_extension
        file = os.path.join(self.input_path, basename)

        if self.start_col == "" or self.start_col is None:
            self.start_col = get_column_letter(
                self._find_column_idx(active_worksheet, "finish", self.wbs_start_row) + 1
            )
        
        custom_ordered_dict = {val["entry"]: val for val in json_dict}

        self.generate_schedule_frame(active_worksheet, self.start_date, self.finish_date)
        reworked_custom_ordered_dict = self.apply_schedule_cfa_style(
            active_worksheet, custom_ordered_dict, proc_table
        )
        
        active_workbook.save(filename=file)
        print("CFA Schedule successfully created")
        self.module_data["logs"]["status"].append(dict(
            Info= f"{ScheduleFramework.__name__}| CFA Schedule successfully created."
        ))
        active_workbook.close()

        return reworked_custom_ordered_dict

    def _find_column_idx(self, active_worksheet, column_header:str, start_row:int) -> int|None:
        ws = active_worksheet
        start_col_idx = column_index_from_string(self.wbs_start_col)
        normalized_header = column_header.strip().replace(" ", "_").lower()
        
        for row in ws.iter_rows(
            min_row=start_row,
            min_col=start_col_idx,
            max_col=active_worksheet.max_column
        ):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    normalized_cell_value = cell.value.strip().replace(" ", "_").lower()
                    if normalized_header == normalized_cell_value:
                        return cell.column
        
        if ScheduleFramework.binary_user_interaction(
            f"Column header '{column_header}' not found. Use default column 'A'?: "
        ):
            return 0
        return None

    def generate_schedule_frame(self, active_worksheet, start_date:str, finish_date:str) -> None:
        ws = active_worksheet

        start_datetime_obj = datetime.strptime(start_date, "%d-%b-%Y")
        finish_datetime_obj = datetime.strptime(finish_date, "%d-%b-%Y")
        duration = (finish_datetime_obj - start_datetime_obj).days + 1  
        
        self._fill_schedule_row(
            ws, start_date, duration, self.start_row, self.start_col, '%Y'
        )
        self._fill_schedule_row(
            ws, start_date, duration, self.start_row+1, self.start_col, '%b'
        )
        self._fill_schedule_row(
            ws, start_date, duration, self.start_row+2, self.start_col, '%d'
        )
        self._fill_schedule_row(
            ws, start_date, duration, self.start_row+3, self.start_col, "", True
        )

        self.module_data["logs"]["status"].append(dict(
            Info = f"{ScheduleFramework.__name__}| Schedule frame generated successfully."
        ))
        print("Schedule frame generated successfully.")

    def _fill_schedule_row(self, active_worksheet, start_date:str, duration:int, 
                         start_row:int, start_col:str, format:str, week_days:bool=False) -> None:
        schedule_workweek_based = []
        start_datetime = datetime.strptime(start_date, "%d-%b-%Y")

        for day in range(duration):
            date = start_datetime + timedelta(days=day)

            if self.calendar_weekdays[date.weekday()] in self.input_workweek:
                schedule_workweek_based.append(date)

        if self.time_scale == 'd':
            for idx, day in enumerate(schedule_workweek_based):
                cell_value = (
                    self.calendar_weekdays[day.weekday()]
                    if week_days 
                    else day.strftime(format)
                )
                active_worksheet.cell(
                    row=start_row, 
                    column=column_index_from_string(start_col) + idx, 
                    value=cell_value
                )
        else:
            self.schedule_scale_based = self._scale_schedule_frame(schedule_workweek_based, self.time_scale)
            if self.schedule_scale_based:
                for idx, dates in enumerate(self.schedule_scale_based):
                    cell_value = (
                        self.calendar_weekdays[dates[0].weekday()]
                        if week_days 
                        else dates[0].strftime(format)
                    )
                    active_worksheet.cell(
                        row=start_row, 
                        column=column_index_from_string(start_col) + idx, 
                        value=cell_value
                    )
            
    def _scale_schedule_frame(self, project_dates:list, time_scale:str) -> list:
        if time_scale not in self.time_scale_options:
            self.module_data["logs"]["status"].append(dict(
                Error= f"{ScheduleFramework.__name__}| Invalid time scale: {time_scale}"
            ))
            raise ValueError(f"Invalid time scale: {time_scale}. Use 'd', 'w', 'm', or 'y'")
        
        dates = []
        
        for date in project_dates:
            if time_scale == "w":
                dates.append(self._calculate_week(date, self.input_workweek))
            elif time_scale == "m":
                dates.append(ScheduleFramework._calculate_month(date, self.input_workweek))
        
        unique_dates = set(dates)

        return sorted(unique_dates)
    
    def _calculate_week(self, date:datetime, workweek:list) -> tuple[datetime, datetime]:
        start_of_week = workweek[0]
        end_of_week = workweek[-1]
        
        while self.calendar_weekdays[date.weekday()] != start_of_week:
            date = date - timedelta(days=1)
        start_of_week_date = date

        while self.calendar_weekdays[date.weekday()] != end_of_week:
            date = date + timedelta(days=1)
        end_of_week_date = date

        return (start_of_week_date, end_of_week_date)
    
    def _calculate_month(self, date:datetime) -> datetime:
        return date.month() 

    def apply_schedule_gantt_style(self, active_worksheet, custom_ordered_dict:dict, proc_table) -> None:
        ws = active_worksheet

        start_ovr_date = datetime.strptime(self.start_date, "%d-%b-%Y")
        final_ovr_date = datetime.strptime(self.finish_date, "%d-%b-%Y")

        for idx, value in enumerate(proc_table.index.get_level_values("entry")):
            item = custom_ordered_dict[value]
            initial_date = datetime.strptime(item["start"], "%d-%b-%Y")
            final_date = datetime.strptime(item["finish"], "%d-%b-%Y")

            if not (start_ovr_date <= initial_date <= final_ovr_date and start_ovr_date <= final_date <= final_ovr_date):
                continue

            start_search_col = column_index_from_string(self.start_col) + (initial_date - start_ovr_date).days
            finish_search_col = column_index_from_string(self.start_col) + (final_date - start_ovr_date).days

            count = 0
            for col in range(start_search_col, finish_search_col + 1):
                cell = ws.cell(row=self.wbs_start_row + idx + 1, column=col)
                self._style_cell(cell, item)
                if count < 1:
                    self._add_comment(cell, item)

                count += 1

    def apply_schedule_cfa_style(self, active_worksheet, custom_ordered_dict:dict, 
                                proc_table) -> dict:
        ws = active_worksheet
        schedule_setup = {
            "start_ovr_date": datetime.strptime(self.start_date, "%d-%b-%Y"),
            "final_ovr_date": datetime.strptime(self.finish_date, "%d-%b-%Y"),
            "starting_point_row": self.wbs_start_row + 1,
            "starting_point_col": self.start_col,
            "ref_lead": "",
            "occupied_rows": {},
            "completed_tasks": set(),
        }

        self._paint_structured_schedule(
            ws, custom_ordered_dict, proc_table, schedule_setup, self.time_scale
        )

    def _paint_structured_schedule(self, active_worksheet, custom_ordered_dict:dict, proc_table, 
                                 schedule_setup:dict, time_scale:str) -> None:
        if time_scale == 'd':
            self._day_based_schedule(active_worksheet, custom_ordered_dict, proc_table, schedule_setup)
        if time_scale == "w":
            self._week_based_schedule(active_worksheet, custom_ordered_dict, proc_table, schedule_setup)
        
    def _day_based_schedule(self, active_worksheet, custom_ordered_dict:dict, proc_table, 
                                 schedule_setup:dict) -> None:
        ws = active_worksheet
        start_ovr_date = schedule_setup.get("start_ovr_date")
        starting_point_row = schedule_setup.get("starting_point_row")
        starting_point_col = schedule_setup.get("starting_point_col")
        ref_lead = schedule_setup.get("ref_lead")
        occupied_rows = schedule_setup.get("occupied_rows", {})
        completed_tasks = schedule_setup.get("completed_tasks", set())

        for idx, entry in enumerate(proc_table.index.get_level_values("entry")):
            obj = custom_ordered_dict[entry]
            current_lead = self._generate_compound_category_name(obj)
            dates = obj["dates"].get("processed", [])

            if not dates:
                continue

            initial_date = datetime.strptime(dates[0], "%d-%b-%Y")

            if current_lead != ref_lead:
                starting_point_row = max(starting_point_row, self.wbs_start_row + idx + 1)
                ref_lead = current_lead
                occupied_rows[starting_point_row] = []

            date_offset = self._determine_date_offset(start_ovr_date, initial_date)
            starting_col = column_index_from_string(starting_point_col) + len(date_offset)
            task_range = range(starting_col, starting_col + len(dates))

            target_row = self._assign_range_to_occupied(occupied_rows, starting_point_row, task_range)
            occupied_rows.setdefault(target_row, []).extend(task_range)

            original_sequence = []
            for i in task_range:
                cell = ws.cell(row=target_row, column=i)

                if obj.get("predecessor"):
                    self._style_cell(cell, obj, True)
                else:
                    self._style_cell(cell, obj)

                if not original_sequence:
                    self._add_comment(cell, obj)

                original_sequence.append(cell.coordinate)

            obj["cell_sequence"] = {
                "original": original_sequence,
                "reworked": None
            }
            completed_tasks.add(entry)

        missing_tasks = set(proc_table.index.get_level_values("entry")) - completed_tasks
        if missing_tasks:
            self.module_data["logs"]["status"].append(dict(
                Warning= f"{ScheduleFramework.__name__}| The following tasks were not painted: {missing_tasks}."
            ))
            print(f"Warning: The following tasks were not painted: {missing_tasks}.")

    def _week_based_schedule(self, active_worksheet, custom_ordered_dict:dict, proc_table, 
                                 schedule_setup:dict) -> None:
        ws = active_worksheet
        starting_point_row = schedule_setup.get("starting_point_row")
        starting_point_col = schedule_setup.get("starting_point_col")
        ref_lead = schedule_setup.get("ref_lead")
        occupied_rows = schedule_setup.get("occupied_rows", {})
        completed_tasks = schedule_setup.get("completed_tasks", set())

        for idx, entry in enumerate(proc_table.index.get_level_values("entry")):
            item = custom_ordered_dict[entry]
            current_lead = self._generate_compound_category_name(item)
            dates = item["dates"].get("processed", [])

            if not dates:
                continue

            initial_date = datetime.strptime(dates[0], "%d-%b-%Y")
            final_date = datetime.strptime(dates[-1], "%d-%b-%Y")

            if current_lead != ref_lead:
                starting_point_row = max(starting_point_row, self.wbs_start_row + idx + 1)
                ref_lead = current_lead
                occupied_rows[starting_point_row] = []

            date_offset, week_coverage = self._determine_date_coverage(initial_date, final_date)
            starting_col = column_index_from_string(starting_point_col) + date_offset
            task_range = range(starting_col, starting_col + week_coverage)

            target_row = self._assign_range_to_occupied(occupied_rows, starting_point_row, task_range)
            occupied_rows.setdefault(target_row, []).extend(task_range)

            original_sequence = []
            for i in task_range:
                cell = ws.cell(row=target_row, column=i)

                if item.get("predecessor"):
                    self._style_cell(cell, item, True)
                else:
                    self._style_cell(cell, item)

                if not original_sequence:
                    self._add_comment(cell, item)

                original_sequence.append(cell.coordinate)

            item["cell_sequence"] = {
                "original": original_sequence,
                "reworked": None
            }
            completed_tasks.add(entry)

        missing_tasks = set(proc_table.index.get_level_values("entry")) - completed_tasks
        if missing_tasks:
            self.module_data["logs"]["status"].append(dict(
                Warning= f"{ScheduleFramework.__name__}| The following tasks were not painted: {missing_tasks}."
            ))
            print(f"Warning: The following tasks were not painted: {missing_tasks}.")
    
    def _generate_compound_category_name(self, json_obj:dict) -> str:
        category_names = []

        for category in self.wbs_final_categories.keys():
            cat_name = json_obj.get(category)
            if cat_name:
                category_names.append(cat_name)

        return "|".join(category_names)

    def _determine_date_offset(self, project_start:datetime, task_start:datetime) -> list:
        date_range = []

        if task_start < project_start:
            self.module_data["logs"]["status"].append(dict(
                Warning= f"{ScheduleFramework.__name__}| Task start date is before project start date. Returning an empty range"
            ))
            print("Warning: Task start date is before project start date. Returning an empty range.")
            return date_range

        date_offset = (task_start - project_start).days

        for i in range(date_offset):
            current_date = project_start + timedelta(days=i)
            if self.calendar_weekdays[current_date.weekday()] in self.input_workweek:
                date_range.append(current_date)

        return date_range

    def _determine_date_coverage(self, task_start:datetime, task_finish:datetime) -> tuple[int, int]:
        if task_start > task_finish:
            self.module_data["logs"]["status"].append(dict(
                Error= f"{ScheduleFramework.__name__}| Task start ({task_start}) cannot be after finish ({task_finish})."
            ))
            raise ValueError(f"Task start ({task_start}) cannot be after finish ({task_finish}).")
        
        if not self.schedule_scale_based:
            return (0, 0)
        
        for i in range(len(self.schedule_scale_based) - 1):
            if self.schedule_scale_based[i][0] > self.schedule_scale_based[i+1][0]:
                self.module_data["logs"]["status"].append(dict(
                    Error= f"{ScheduleFramework.__name__}| Schedule periods must be in chronological order."
                ))
                raise ValueError("Schedule periods must be in chronological order.")
        
        offset = 0
        for period_start, period_end in self.schedule_scale_based:
            if task_start <= period_end:
                break
            offset += 1
        
        coverage = 0
        for period_start, period_end in self.schedule_scale_based[offset:]:
            if task_finish < period_start:
                break
            coverage += 1
    
        return (offset, coverage)

    def _assign_range_to_occupied(self, occupied_rows:dict, target_row:int, range_to_check:list) -> int:
        while any(target_col in occupied_rows.get(target_row, []) for target_col in range_to_check):
            target_row += 1

        return target_row

    def _style_cell(self, cell, current_item:dict, paint_border:bool=False) -> None:
        try:
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if paint_border:
                cell.border = Border(
                    top=Side(border_style="double", color="00FF0000"), 
                    left=Side(border_style="double", color="00FF0000"), 
                    right=Side(border_style="double", color="00FF0000"), 
                    bottom=Side(border_style="double", color="00FF0000"))
            
            color = current_item.get("color")
            if color:
                cell.fill = PatternFill(start_color=re.sub('#', "00", current_item.get("color")), 
                                    end_color=re.sub('#', "00", current_item.get("color")), 
                                    fill_type='solid')
            else:
                print(f"Color hex not found for: {current_item['entry']}.")
                self.module_data["logs"]["status"].append(dict(
                    Warning= f"{ScheduleFramework.__name__}| Color hex not found for: {current_item['entry']}."
                ))
                cell.fill = PatternFill(start_color=self.default_hex_fill_color, 
                                    end_color=self.default_hex_fill_color, 
                                    fill_type='solid')
            
            rgb = ScheduleFramework.hex2rgb(color)
            lum_rgb = ScheduleFramework.calculateLuminance(rgb[0], rgb[1], rgb[2])
            if lum_rgb > 0.5:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.dark_default_hex_font)
            else:
                cell.font = Font(name="Calibri", size="12", bold=True, color=self.light_default_hex_font)

            cell.value = current_item.get("activity_code") if current_item.get("activity_code") else "NaN"
            
        except Exception as e:
            self.module_data["logs"]["status"].append(dict(
                Error= f"{ScheduleFramework.__name__}| While painting cell: {e}."
            ))
            print(f"Error. While painting cell: {e}.")

    def _add_comment(self, cell, current_item:dict) -> None:
        msg = f"""
            Activity Name: {current_item.get('activity_name', "NaN")}
            Phase: {current_item.get("phase", "NaN")} 
            Location: {current_item.get("area", "NaN")}
            Trade: {current_item.get("trade", "NaN")} 

            Start Date: {current_item.get("start", "NaN")}
            End Date: {current_item.get("finish", "NaN")}
        """

        comment = Comment(msg, "AMP")
        comment.width = 300
        comment.height = 200

        cell.comment = comment


if __name__ == "__main__":
    project = ScheduleFramework.main(False)
